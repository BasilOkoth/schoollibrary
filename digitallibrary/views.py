# ==============================
# file: digitallibrary/views.py
# ==============================
from __future__ import annotations
from django.db import models
import re
from django.db.models import Sum
from digitallibrary.decorators import role_required
from django.http import HttpResponseRedirect

import logging
from django.contrib.auth import authenticate, login, logout
from tenants.models import School
from .decorators import parent_session_required
from django.db import models
from .models import ParentOTP
from .forms import ParentLoginForm, ParentOTPForm
from django.db import models
from django.conf import settings
from .decorators import tenant_app_view
from collections import defaultdict
from statistics import mean, pstdev
from django.db import connection
from digitallibrary.decorators import fees_access
from digitallibrary.decorators import sms_access
from django.db.models import Sum
from .sms_utils import send_sms, send_bulk_sms, send_to_teachers, send_to_students, send_to_all_users
from .sms_utils import format_phone_number
from django.views.decorators.http import require_http_methods
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from .models import Student, FeePayment, FeeBalance, HistoricalArrears, Term, FeeStructure
import json
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Exam, TeacherGradingPreference, GradingSystem, CBEGradingPathway

from django.http import HttpResponse
from django.template.loader import get_template
from django.template import Context
from xhtml2pdf import pisa
import io
from decimal import Decimal

import os
import openpyxl
import mimetypes
import random
from digitallibrary.models import UserProfile
import csv
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail
from django.conf import settings
import pandas as pd
from django.contrib import messages
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import redirect
from django.urls import reverse
from django.core.validators import ValidationError
from .forms import BulkStudentUploadForm
from .models import Student, Class
from datetime import datetime, timedelta
from django.core.cache import cache
from django.db.models import Count, Sum, Q, Avg, Max, Min
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.http import HttpResponse, Http404, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib.auth.views import LoginView
from django.urls import reverse
from django.contrib.admin.views.decorators import staff_member_required
from django.core.mail import send_mail
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
# ReportLab imports for PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from django.contrib.auth.decorators import login_required, user_passes_test
# Tenant imports for central dashboard
from tenants.models import School
from django_tenants.utils import tenant_context
from digitallibrary.decorators import role_required
from .forms import ResourceForm, AnnouncementForm, AnnouncementFilterForm, FeedbackForm
from .models import (
    Resource,
    Category,
    Subject,
    PrintJob,
    SchoolSetting,
    UserProfile,
    Announcement,
    AnnouncementRead,
    ActivityLog,
    Notification,
    FeeStructure,
    Student,
    FeePayment,
    FeeBalance,
    Feedback,
    Class,
    
    FeeComponent,
    Subject,
    Class as ClassModel,
    Exam,
    StudentResult,
    PerformanceSummary,
    Grade,
    PaperSet,
    PaperResource,
)
from .forms import (
    FeeStructureForm, 
    StudentForm, 
    FeePaymentForm, 
    StudentSearchForm,
    ExamForm,
    StudentResultForm,
    BulkResultForm,
    PaperResourceForm,
    PaperSetFilterForm,
)
from . import models

# ========== HELPER FUNCTIONS ==========

MOCK_SMS_MODE = getattr(settings, 'MOCK_SMS_MODE', True)


def generate_receipt_number():
    """Generate a unique receipt number"""
    prefix = "RCP"
    date_str = datetime.now().strftime("%Y%m%d")
    random_num = str(random.randint(1000, 9999))
    
    receipt_number = f"{prefix}{date_str}{random_num}"
    
    while FeePayment.objects.filter(receipt_number=receipt_number).exists():
        random_num = str(random.randint(1000, 9999))
        receipt_number = f"{prefix}{date_str}{random_num}"
    
    return receipt_number


def update_fee_balance_after_payment(payment):
    """Update fee balance after a payment is recorded"""
    from django.db.models import Sum
    from digitallibrary.models import FeeBalance, FeePayment
    
    # Get or create fee balance
    balance, created = FeeBalance.objects.get_or_create(
        student=payment.student,
        term=payment.term,
        academic_year=payment.academic_year,
        defaults={
            'total_expected': 0,
            'total_paid': 0,
            'balance': 0
        }
    )
    
    # Calculate total paid
    total_paid = FeePayment.objects.filter(
        student=payment.student,
        term=payment.term,
        academic_year=payment.academic_year
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Just update total paid - don't query fee structure
    balance.total_paid = total_paid
    balance.save()


def get_year_choices():
    """Generate year choices from 2000 to next year"""
    current_year = timezone.now().year
    years = [str(year) for year in range(current_year + 1, 1999, -1)]
    years.append('N/A')
    return years


def get_grade_from_score(score):
    """Helper function to get grade from score"""
    grade_obj = Grade.objects.filter(min_score__lte=score, max_score__gte=score).first()
    return grade_obj.grade if grade_obj else 'N/A'


def update_performance_summary(student, academic_year, term):
    """Update or create performance summary for a student"""
    results = StudentResult.objects.filter(
        student=student,
        exam__academic_year=academic_year,
        exam__term=term
    ).select_related('exam')
    
    if not results:
        return
    
    total_score = sum(float(r.score) for r in results)
    average_score = total_score / len(results)
    
    total_points = 0
    subjects_passed = 0
    subjects_failed = 0
    
    for result in results:
        grade_obj = Grade.objects.filter(min_score__lte=result.score, max_score__gte=result.score).first()
        if grade_obj:
            total_points += float(grade_obj.points)
            if grade_obj.grade in ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-']:
                subjects_passed += 1
            else:
                subjects_failed += 1
    
    average_points = total_points / len(results) if results else 0
    
    overall_grade_obj = Grade.objects.filter(min_score__lte=average_score, max_score__gte=average_score).first()
    overall_grade = overall_grade_obj.grade if overall_grade_obj else None
    
    class_summaries = PerformanceSummary.objects.filter(
        student__current_class=student.current_class,
        academic_year=academic_year,
        term=term
    ).order_by('-average_score')
    
    rank = 1
    for i, summary in enumerate(class_summaries, 1):
        if summary.student_id == student.id:
            rank = i
            break
    
    PerformanceSummary.objects.update_or_create(
        student=student,
        academic_year=academic_year,
        term=term,
        defaults={
            'total_score': total_score,
            'average_score': average_score,
            'total_points': total_points,
            'average_points': average_points,
            'overall_grade': overall_grade,
            'rank_in_class': rank,
            'subjects_passed': subjects_passed,
            'subjects_failed': subjects_failed,
        }
    )

@staff_member_required
def search_students_ajax(request):
    """AJAX endpoint for searching students"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'results': []})
    
    students = Student.objects.filter(is_active=True).filter(
        Q(admission_number__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(parent_phone__icontains=query) |
        Q(upi_number__icontains=query)
    )[:20]
    
    results = []
    for student in students:
        results.append({
            'id': student.id,
            'admission_number': student.admission_number,
            'name': student.get_full_name(),
            'class': student.current_class.name if student.current_class else 'N/A',
            'parent_phone': student.parent_phone,
        })
    
    return JsonResponse({'results': results})
# ========== AJAX SEARCH FUNCTIONS ==========

@staff_member_required
def search_students_ajax(request):
    """AJAX endpoint for searching students"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'results': []})
    
    students = Student.objects.filter(is_active=True).filter(
        Q(admission_number__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(parent_phone__icontains=query) |
        Q(upi_number__icontains=query)
    )[:20]
    
    results = []
    for student in students:
        results.append({
            'id': student.id,
            'admission_number': student.admission_number,
            'name': student.get_full_name(),
            'class': student.current_class.name if student.current_class else 'N/A',
            'parent_phone': student.parent_phone,
        })
    
    return JsonResponse({'results': results})


@login_required
def search_students_for_payment(request):
    """AJAX endpoint to search students for payment selection"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'results': []})
    
    students = Student.objects.filter(is_active=True).filter(
        Q(admission_number__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(parent_phone__icontains=query) |
        Q(upi_number__icontains=query)
    )[:20]
    
    results = []
    for student in students:
        results.append({
            'id': student.id,
            'text': f"{student.admission_number} - {student.first_name} {student.last_name} ({student.current_class.name if student.current_class else 'No Class'})",
            'admission': student.admission_number,
            'name': f"{student.first_name} {student.last_name}",
            'class': student.current_class.name if student.current_class else 'N/A',
            'parent_phone': student.parent_phone,
        })
    
    return JsonResponse({'results': results})


@login_required
def get_students_by_class(request, class_id):
    """API to get students for a specific class"""
    students = Student.objects.filter(
        current_class_id=class_id,
        is_active=True
    ).values('id', 'admission_number', 'first_name', 'last_name')
    
    return JsonResponse({
        'success': True,
        'students': list(students)
    })


@login_required
def get_all_students(request):
    """API to get all active students"""
    students = Student.objects.filter(
        is_active=True
    ).values('id', 'admission_number', 'first_name', 'last_name')
    
    return JsonResponse({
        'success': True,
        'students': list(students)
    })


@login_required
def user_management(request):
    """Manage all users in the system"""
    if request.user.profile.role not in ['admin', 'principal']:
        messages.error(request, "Access Denied. Only administrators can manage users.")
        return redirect('digitallibrary:home')
    
    # Get all users with their profiles - ADDED order_by to fix pagination warning
    users = User.objects.all().select_related('profile').order_by('-date_joined')
    
    # Filter by role if specified
    role_filter = request.GET.get('role', '')
    if role_filter:
        users = users.filter(profile__role=role_filter)
    
    # Search
    search = request.GET.get('search', '')
    if search:
        users = users.filter(
            Q(username__icontains=search) |
            Q(email__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page', 1)
    users_page = paginator.get_page(page_number)
    
    # Calculate statistics
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    inactive_users = User.objects.filter(is_active=False).count()
    
    # Role counts
    role_counts = {}
    for role_code, role_name in UserProfile.ROLE_CHOICES:
        count = UserProfile.objects.filter(role=role_code).count()
        role_counts[role_code] = count
    
    context = {
        'users': users_page,
        'role_filter': role_filter,
        'search': search,
        'roles': UserProfile.ROLE_CHOICES,
        'school': SchoolSetting.objects.first(),
        # Statistics
        'total_users': total_users,
        'active_users': active_users,
        'inactive_users': inactive_users,
        'role_counts': role_counts,
    }
    return render(request, 'digitallibrary/user_management.html', context)
@login_required
@role_required(['admin', 'principal'])
def assign_class_teachers(request):
    """Assign class teachers (homeroom teachers) - Admin and Principal only"""
    
    classes = Class.objects.all().order_by('name')
    
    # Teachers that can be class teachers (including class_teacher role and regular teachers)
    teachers = User.objects.filter(
        profile__role__in=['class_teacher', 'teacher', 'admin', 'principal'],
        profile__is_approved=True,
        is_active=True
    ).order_by('first_name', 'last_name')
    
    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        teacher_id = request.POST.get('teacher_id')
        
        class_obj = get_object_or_404(Class, id=class_id)
        
        if teacher_id:
            teacher = get_object_or_404(User, id=teacher_id)
            class_obj.class_teacher = teacher
            class_obj.save()
            
            # Update user's role to class_teacher if they are a regular teacher
            if teacher.profile.role == 'teacher':
                teacher.profile.role = 'class_teacher'
                teacher.profile.save()
            
            messages.success(request, f"{teacher.get_full_name()} assigned as class teacher for {class_obj.name}")
        else:
            # Remove class teacher
            old_teacher = class_obj.class_teacher
            class_obj.class_teacher = None
            class_obj.save()
            
            # Optionally revert role back to teacher (if they have no other classes)
            if old_teacher:
                other_classes = Class.objects.filter(class_teacher=old_teacher).exclude(id=class_obj.id)
                if not other_classes.exists() and old_teacher.profile.role == 'class_teacher':
                    old_teacher.profile.role = 'teacher'
                    old_teacher.profile.save()
            
            messages.success(request, f"Class teacher removed for {class_obj.name}")
        
        return redirect('digitallibrary:assign_class_teachers')
    
    context = {
        'classes': classes,
        'teachers': teachers,
        'title': 'Assign Class Teachers',
    }
    return render(request, 'digitallibrary/assign_class_teachers.html', context)
@login_required
@role_required(['admin', 'principal'])
def class_teacher_dashboard(request):
    """Dashboard for class teachers to see their assigned class"""
    
    if request.user.profile.role == 'class_teacher':
        # Get the class this teacher is assigned to
        assigned_class = Class.objects.filter(class_teacher=request.user).first()
        
        if assigned_class:
            # Get students in this class
            students = Student.objects.filter(current_class=assigned_class, is_active=True)
            
            context = {
                'assigned_class': assigned_class,
                'students': students,
                'total_students': students.count(),
                'title': f'Class Teacher Dashboard - {assigned_class.name}',
            }
            return render(request, 'digitallibrary/class_teacher_dashboard.html', context)
        else:
            messages.warning(request, "You are not assigned to any class yet.")
            return redirect('digitallibrary:home')
    
    # For admin/principal viewing all classes
    classes = Class.objects.all().order_by('name')
    context = {
        'classes': classes,
        'title': 'Class Teacher Overview',
    }
    return render(request, 'digitallibrary/class_teacher_overview.html', context)

def add_user(request):
    """Add a new user to the system"""
    from django.contrib.auth.models import User
    from digitallibrary.models import UserProfile, SchoolSetting
    from django.core.mail import send_mail
    from django.conf import settings
    
    if request.user.profile.role not in ['admin', 'principal']:
        messages.error(request, "Access Denied.")
        return redirect('digitallibrary:home')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        role = request.POST.get('role')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validation
        errors = []
        if User.objects.filter(username=username).exists():
            errors.append(f"Username '{username}' already exists.")
        if User.objects.filter(email=email).exists():
            errors.append(f"Email '{email}' already exists.")
        if not password:
            errors.append("Password is required.")
        elif password != confirm_password:
            errors.append("Passwords do not match.")
        elif len(password) < 6:
            errors.append("Password must be at least 6 characters.")
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            # Create user
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            
            # Create profile
            profile, created = UserProfile.objects.get_or_create(user=user)
            profile.role = role
            profile.is_approved = True
            profile.save()
            
            messages.success(request, f"User '{username}' created successfully! Password: {password}")
            return redirect('digitallibrary:user_management')
    
    school = SchoolSetting.objects.first()
    context = {
        'roles': UserProfile.ROLE_CHOICES,
        'title': 'Add New User',
        'school': school,
    }
    return render(request, 'digitallibrary/user_form.html', context)
from django.contrib.auth.decorators import login_required

@login_required
def edit_user(request, user_id):
    """Edit user details"""
    from django.contrib.auth.models import User
    from digitallibrary.models import UserProfile, SchoolSetting
    
    # Debug: Check user profile
    try:
        current_user_role = request.user.profile.role
        print(f"Current user: {request.user.username}, Role: {current_user_role}")
    except Exception as e:
        print(f"Error getting user profile: {e}")
        # Create profile for current user if missing
        profile, created = UserProfile.objects.get_or_create(user=request.user)
        if created:
            profile.role = 'admin' if request.user.is_superuser else 'staff'
            profile.save()
            current_user_role = profile.role
    
    # Check permission
    if current_user_role not in ['admin', 'principal']:
        messages.error(request, "Access Denied. You don't have permission to edit users.")
        return redirect('digitallibrary:home')
    
    edit_user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        try:
            # Update basic fields
            edit_user.first_name = request.POST.get('first_name', '')
            edit_user.last_name = request.POST.get('last_name', '')
            edit_user.email = request.POST.get('email', '')
            edit_user.is_active = request.POST.get('is_active') == 'on'
            
            # Update role
            new_role = request.POST.get('role')
            if new_role:
                profile, created = UserProfile.objects.get_or_create(user=edit_user)
                profile.role = new_role
                profile.is_approved = True
                profile.save()
                print(f"Updated role for {edit_user.username} to {new_role}")
            
            # Check if password reset is requested
            new_password = request.POST.get('new_password')
            if new_password and len(new_password) >= 6:
                edit_user.set_password(new_password)
                messages.success(request, f"Password for '{edit_user.username}' has been reset successfully!")
                print(f"Password reset for {edit_user.username}")
            elif new_password and len(new_password) < 6:
                messages.warning(request, "Password not changed. Must be at least 6 characters.")
            
            edit_user.save()
            messages.success(request, f"User '{edit_user.username}' updated successfully!")
            return redirect('digitallibrary:user_management')
            
        except Exception as e:
            messages.error(request, f"Error updating user: {str(e)}")
            print(f"Error in edit_user: {e}")
            return redirect('digitallibrary:user_management')
    
    # Get user's current role
    try:
        user_profile = UserProfile.objects.get(user=edit_user)
        current_role = user_profile.role
    except:
        current_role = 'user'
    
    school = SchoolSetting.objects.first()
    context = {
        'edit_user': edit_user,
        'roles': UserProfile.ROLE_CHOICES,
        'current_role': current_role,
        'title': f'Edit User - {edit_user.username}',
        'school': school,
    }
    return render(request, 'digitallibrary/user_form.html', context)
@login_required
def reset_user_password(request, user_id):
    """Reset user password"""
    if request.user.profile.role not in ['admin', 'principal']:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
        messages.error(request, "Access Denied.")
        return redirect('digitallibrary:user_management')
    
    if request.method == 'POST':
        try:
            # Handle JSON request from fetch API
            if request.headers.get('Content-Type') == 'application/json':
                data = json.loads(request.body)
                new_password = data.get('password')
            else:
                new_password = request.POST.get('new_password')
            
            user = get_object_or_404(User, id=user_id)
            
            if not new_password or len(new_password) < 6:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': 'Password must be at least 6 characters'})
                messages.error(request, "Password must be at least 6 characters.")
                return redirect('digitallibrary:user_management')
            
            # Set new password
            user.set_password(new_password)
            user.save()
            
            # Send email notification (optional)
            try:
                send_mail(
                    subject="Your Password Has Been Reset",
                    message=f"""
                    Hello {user.first_name} {user.last_name},
                    
                    Your password has been reset by an administrator.
                    
                    New Login Details:
                    Username: {user.username}
                    Password: {new_password}
                    
                    Please change your password after logging in.
                    
                    Regards,
                    Administration
                    """,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[user.email] if user.email else [],
                    fail_silently=True,
                )
            except:
                pass
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': f'Password reset to: {new_password}'})
            
            messages.success(request, f"Password for '{user.username}' has been reset to: {new_password}")
            return redirect('digitallibrary:user_management')
            
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(request, f"Error: {str(e)}")
            return redirect('digitallibrary:user_management')
    
    # For GET requests, show a simple form
    context = {
        'user': get_object_or_404(User, id=user_id),
        'title': 'Reset Password',
    }
    return render(request, 'digitallibrary/reset_password.html', context)
@login_required
def get_user_json(request, user_id):
    """Get user data as JSON for modal forms"""
    if request.user.profile.role not in ['admin', 'principal']:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    user = get_object_or_404(User, id=user_id)
    return JsonResponse({
        'id': user.id,
        'username': user.username,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'email': user.email,
        'role': user.profile.role,
        'is_active': user.is_active,
    })
@login_required
def toggle_user_status(request, user_id):
    """Activate/Deactivate user"""
    if request.user.profile.role not in ['admin', 'principal']:
        messages.error(request, "Access Denied.")
        return redirect('digitallibrary:user_management')
    
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('digitallibrary:user_management')
    
    user = get_object_or_404(User, id=user_id)
    user.is_active = not user.is_active
    user.save()
    
    status = "activated" if user.is_active else "deactivated"
    messages.success(request, f"User '{user.username}' has been {status}.")
    
    return redirect('digitallibrary:user_management')

@login_required
def delete_user(request, user_id):
    """Delete user (soft delete or hard delete)"""
    if request.user.profile.role != 'admin':
        messages.error(request, "Access Denied. Only administrators can delete users.")
        return redirect('digitallibrary:user_management')
    
    # Only allow POST requests for security
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('digitallibrary:user_management')
    
    user = get_object_or_404(User, id=user_id)
    
    # Don't allow deleting yourself
    if user.id == request.user.id:
        messages.error(request, "You cannot delete your own account.")
        return redirect('digitallibrary:user_management')
    
    username = user.username
    user.delete()
    messages.success(request, f"User '{username}' has been deleted.")
    
    return redirect('digitallibrary:user_management')
# ========== PERFORMANCE VIEWS ==========

@staff_member_required
@tenant_app_view
def exam_list(request):
    """List all exams with class filtering and results count"""
    exams = Exam.objects.all().select_related('student_class').order_by('-academic_year', '-term', 'name')
    
    year = request.GET.get('year')
    term = request.GET.get('term')
    class_id = request.GET.get('class')
    
    if year:
        exams = exams.filter(academic_year=year)
    if term:
        exams = exams.filter(term=term)
    if class_id:
        exams = exams.filter(student_class_id=class_id)
    
    # Add results count and student count for each exam
    for exam in exams:
        # Count distinct students who have results for this exam
        exam.results_count = StudentResult.objects.filter(exam=exam).values('student').distinct().count()
        # Get total students for this exam
        exam.total_students = exam.get_students_for_exam().count()
    
    # Get available years for filter
    available_years = Exam.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
    
    context = {
        'exams': exams,
        'current_year': request.GET.get('year', str(timezone.now().year)),
        'current_term': request.GET.get('term', ''),
        'selected_class': request.GET.get('class', ''),
        'classes': Class.objects.all().order_by('name'),
        'available_years': available_years,
        'school': SchoolSetting.objects.first(),
    }
    return render(request, 'performance/exam_list.html', context)

@staff_member_required
def exam_create(request):
    """Create a new exam"""
    if request.method == 'POST':
        form = ExamForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Exam created successfully!')
            return redirect('digitallibrary:exam_list')
    else:
        form = ExamForm()
    
    return render(request, 'performance/exam_form.html', {'form': form, 'title': 'Create Exam'})
@tenant_app_view
def bulk_select(request):
    """Step 1: Select exam and subject for bulk entry"""
    from .models import Exam, Subject, Student
    
    exams = Exam.objects.all().order_by('-academic_year', '-created_at')
    subjects = Subject.objects.all().order_by('name')
    
    selected_exam_id = None
    selected_subject_id = None
    selected_exam = None
    total_students = 0
    
    if request.method == 'POST':
        exam_id = request.POST.get('exam')
        subject_id = request.POST.get('subject')
        
        if exam_id and subject_id:
            try:
                exam = Exam.objects.get(id=exam_id)
                subject = Subject.objects.get(id=subject_id)
                return redirect('digitallibrary:bulk_results_entry', exam_id=exam.id, subject_id=subject.id)
            except (Exam.DoesNotExist, Subject.DoesNotExist):
                messages.error(request, 'Invalid selection')
        
        selected_exam_id = exam_id
        selected_subject_id = subject_id
        if selected_exam_id:
            try:
                selected_exam = Exam.objects.get(id=selected_exam_id)
            except Exam.DoesNotExist:
                pass
    else:
        exam_id = request.GET.get('exam')
        if exam_id:
            try:
                selected_exam = Exam.objects.get(id=exam_id)
                selected_exam_id = exam_id
            except Exam.DoesNotExist:
                pass
    
    if selected_exam:
        if selected_exam.student_class:
            total_students = selected_exam.student_class.students.filter(is_active=True).count()
        else:
            total_students = Student.objects.filter(is_active=True).count()
    
    context = {
        'exams': exams,
        'subjects': subjects,
        'selected_exam_id': selected_exam_id,
        'selected_subject_id': selected_subject_id,
        'selected_exam': selected_exam,
        'total_students': total_students,
    }
    
    return render(request, 'digitallibrary/bulk_select.html', context)

@staff_member_required
def bulk_results_entry(request, exam_id, subject_id):
    """
    Enter results for all students in a class for a specific subject
    """
    from .models import Exam, Subject, Student, StudentResult, Class
    from django.db import connection
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect, render
    
    exam = get_object_or_404(Exam, id=exam_id)
    subject = get_object_or_404(Subject, id=subject_id)
    
    # Get class_id from session
    class_id = request.session.get('bulk_class_id')
    if not class_id:
        class_id = request.GET.get('class_id')
    
    if not class_id:
        messages.error(request, 'Please select a class first')
        return redirect('digitallibrary:bulk_enter_results')
    
    student_class = get_object_or_404(Class, id=class_id)
    
    # Get students
    students = Student.objects.filter(
        current_class=student_class,
        is_active=True
    ).order_by('first_name', 'last_name')
    
    if not students.exists():
        students = Student.objects.filter(
            student_class=student_class,
            is_active=True
        ).order_by('first_name', 'last_name')
    
    # Check grading system from session
    session_grading = request.session.get('active_grading_system_id')
    use_cbe = session_grading == 'cbe'
    
    # Get existing results as a dictionary for easy lookup
    existing_results = {}
    if students.exists():
        results = StudentResult.objects.filter(
            exam=exam, 
            subject=subject,
            student__in=students
        ).select_related('student')
        for r in results:
            existing_results[r.student_id] = r
    
    # Handle POST request
    if request.method == 'POST':
        saved_count = 0
        error_count = 0
        
        for key, value in request.POST.items():
            if key.startswith('score_') and value.strip():
                student_id = key.replace('score_', '')
                try:
                    score = float(value)
                    
                    # Validate score range
                    max_score = exam.max_score if exam.max_score else 100
                    if score < 0 or score > max_score:
                        error_count += 1
                        continue
                    
                    student = Student.objects.get(id=student_id)
                    
                    # Calculate grade based on system
                    if use_cbe:
                        # Get CBE grade from kneccbegrade table
                        with connection.cursor() as cursor:
                            cursor.execute("""
                                SELECT id, grade, points 
                                FROM digitallibrary_kneccbegrade 
                                WHERE min_score <= %s AND max_score >= %s
                                LIMIT 1
                            """, [score, score])
                            grade_result = cursor.fetchone()
                            
                            if grade_result:
                                grade_id, grade_name, points = grade_result
                                StudentResult.objects.update_or_create(
                                    student=student,
                                    exam=exam,
                                    subject=subject,
                                    defaults={
                                        'score': score, 
                                        'grade_id': grade_id,
                                        'grade': grade_name,
                                        'points': points,
                                        'entered_by': request.user
                                    }
                                )
                                saved_count += 1
                            else:
                                error_count += 1
                    else:
                        # Traditional grading
                        max_possible = exam.max_score if exam.max_score else 100
                        percentage = (score / max_possible) * 100
                        
                        if percentage >= 80:
                            grade_name = 'A'
                            points = 12
                        elif percentage >= 75:
                            grade_name = 'A-'
                            points = 11
                        elif percentage >= 70:
                            grade_name = 'B+'
                            points = 10
                        elif percentage >= 65:
                            grade_name = 'B'
                            points = 9
                        elif percentage >= 60:
                            grade_name = 'B-'
                            points = 8
                        elif percentage >= 55:
                            grade_name = 'C+'
                            points = 7
                        elif percentage >= 50:
                            grade_name = 'C'
                            points = 6
                        elif percentage >= 45:
                            grade_name = 'C-'
                            points = 5
                        elif percentage >= 40:
                            grade_name = 'D+'
                            points = 4
                        elif percentage >= 35:
                            grade_name = 'D'
                            points = 3
                        elif percentage >= 30:
                            grade_name = 'D-'
                            points = 2
                        else:
                            grade_name = 'E'
                            points = 1
                        
                        # Try to find grade_id from digitallibrary_grade
                        with connection.cursor() as cursor:
                            cursor.execute("""
                                SELECT id FROM digitallibrary_grade 
                                WHERE grade = %s LIMIT 1
                            """, [grade_name])
                            grade_result = cursor.fetchone()
                            grade_id = grade_result[0] if grade_result else None
                        
                        StudentResult.objects.update_or_create(
                            student=student,
                            exam=exam,
                            subject=subject,
                            defaults={
                                'score': score, 
                                'grade_id': grade_id,
                                'grade': grade_name,
                                'points': points,
                                'entered_by': request.user
                            }
                        )
                        saved_count += 1
                        
                except (ValueError, Student.DoesNotExist) as e:
                    error_count += 1
                    continue
                except Exception as e:
                    error_count += 1
                    continue
        
        if saved_count > 0:
            messages.success(request, f'✅ Successfully saved {saved_count} results for {subject.name}')
        if error_count > 0:
            messages.warning(request, f'⚠️ Failed to save {error_count} results. Please check your scores.')
        
        return redirect('digitallibrary:bulk_results_entry', exam_id=exam.id, subject_id=subject.id)
    
    # Build a simple dictionary for template to avoid custom filters
    existing_scores = {}
    existing_grades = {}
    existing_points = {}
    
    for student in students:
        result = existing_results.get(student.id)
        if result:
            existing_scores[student.id] = result.score
            existing_grades[student.id] = result.grade
            existing_points[student.id] = result.points
        else:
            existing_scores[student.id] = ''
            existing_grades[student.id] = '—'
            existing_points[student.id] = '—'
    
    context = {
        'exam': exam,
        'subject': subject,
        'class': student_class,
        'students': students,
        'existing_scores': existing_scores,
        'existing_grades': existing_grades,
        'existing_points': existing_points,
        'student_count': students.count(),
        'use_cbe': use_cbe,
        'max_score': exam.max_score if exam.max_score else 100,
        'title': f'Enter Results - {exam.name} - {subject.name} - {student_class.name}',
    }
    return render(request, 'digitallibrary/bulk_results_entry.html', context)

@staff_member_required
def download_excel_template(request):
    """Download Excel template for bulk upload"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="exam_results_template.xlsx"'
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Results Template"
    
    # Headers
    headers = ['Admission Number', 'Score', 'Student Name (Optional)']
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a4d8c", end_color="1a4d8c", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Sample data
    sample_data = [
        ['2024001', 85.5, 'John Doe'],
        ['2024002', 92.0, 'Jane Smith'],
        ['2024003', 76.5, 'Mike Johnson'],
        ['2024004', 68.0, 'Sarah Williams'],
        ['2024005', 94.5, 'David Brown'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    
    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    instructions_ws['A1'] = "Instructions for Bulk Upload"
    instructions_ws['A1'].font = Font(bold=True, size=14)
    
    instructions = [
        "",
        "1. Admission Number column is REQUIRED - must match existing student admission numbers",
        "2. Score column is REQUIRED - must be a number between 0 and the exam's max score",
        "3. Student Name column is OPTIONAL - for reference only",
        "4. The first row contains column headers - do not delete or modify",
        "5. You can add as many rows as needed",
        "6. Scores will be created or updated for the selected exam and subject",
        "7. Make sure students already exist in the system before uploading",
    ]
    
    for i, instruction in enumerate(instructions, 1):
        instructions_ws.cell(row=i, column=1, value=instruction)
    
    instructions_ws.column_dimensions['A'].width = 60
    
    wb.save(response)
    return response
@staff_member_required
def exam_edit(request, pk):
    """Edit an exam"""
    exam = get_object_or_404(Exam, pk=pk)
    
    if request.method == 'POST':
        form = ExamForm(request.POST, instance=exam)
        if form.is_valid():
            form.save()
            messages.success(request, 'Exam updated successfully!')
            return redirect('digitallibrary:exam_list')
    else:
        form = ExamForm(instance=exam)
    
    return render(request, 'performance/exam_form.html', {'form': form, 'title': 'Edit Exam'})

def system_dashboard(request):
    """Executive dashboard with filtering by term, class, year, and subject"""
    from .models import Exam, Class, Subject, Student, StudentResult
    from django.db.models import Avg, Sum, Count, Q
    
    # Get filter parameters
    current_year = request.GET.get('year', '')
    current_term = request.GET.get('term', '')
    selected_class = request.GET.get('class', '')
    selected_subject = request.GET.get('subject', '')
    
    # Build exams queryset with filters
    exams_qs = Exam.objects.all().order_by('-academic_year', '-created_at')
    
    # Apply year filter
    if current_year:
        exams_qs = exams_qs.filter(academic_year=current_year)
    
    # Apply term filter
    if current_term:
        exams_qs = exams_qs.filter(term=current_term)
    
    # Apply class filter
    if selected_class:
        exams_qs = exams_qs.filter(student_class_id=selected_class)
    
    # Debug: Print the count
    print(f"Exams found: {exams_qs.count()}")
    print(f"Filters - Year: {current_year}, Term: {current_term}, Class: {selected_class}")
    
    # Get available years for filter dropdown
    available_years = Exam.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
    
    # Get filtered exams with completion rates
    recent_exams = []
    for exam in exams_qs:
        # Get total students for this exam
        if exam.student_class:
            total_students = exam.student_class.students.filter(is_active=True).count()
        else:
            total_students = Student.objects.filter(is_active=True).count()
        
        # Get results count based on filters
        results_qs = StudentResult.objects.filter(exam=exam)
        
        # Apply subject filter to results count
        if selected_subject:
            results_qs = results_qs.filter(subject_id=selected_subject)
            results_count = results_qs.values('student').distinct().count()
        else:
            results_count = results_qs.values('student').distinct().count()
        
        # Calculate completion rate
        completion_rate = (results_count / total_students * 100) if total_students > 0 else 0
        
        recent_exams.append({
            'id': exam.id,
            'name': exam.name,
            'academic_year': exam.academic_year,
            'term': exam.term,
            'class_name': exam.student_class.name if exam.student_class else 'All Classes',
            'completion_rate': completion_rate,
            'total_students': total_students,
            'results_count': results_count,
        })
    
    # Calculate overall metrics with filters
    results_qs = StudentResult.objects.all()
    
    # Apply all filters to results
    if current_year:
        results_qs = results_qs.filter(exam__academic_year=current_year)
    if current_term:
        results_qs = results_qs.filter(exam__term=current_term)
    if selected_class:
        results_qs = results_qs.filter(student__current_class_id=selected_class)
    if selected_subject:
        results_qs = results_qs.filter(subject_id=selected_subject)
    
    # Calculate averages
    avg_data = results_qs.aggregate(avg=Avg('score'))
    avg_score = avg_data['avg'] or 0
    
    # Calculate pass rate
    total_results = results_qs.count()
    passed_results = results_qs.filter(score__gte=50).count()
    pass_rate = (passed_results / total_results * 100) if total_results > 0 else 0
    
    # Get total students
    students_qs = Student.objects.filter(is_active=True)
    if selected_class:
        students_qs = students_qs.filter(current_class_id=selected_class)
    total_students = students_qs.count()
    
    # Get total exams count after filters
    total_exams = exams_qs.count()
    
    # Get classes and subjects for filter dropdowns
    classes = Class.objects.all().order_by('name')
    subjects = Subject.objects.all().order_by('name')
    
    # Get selected class/subject names for display
    selected_class_name = None
    if selected_class:
        class_obj = Class.objects.filter(id=selected_class).first()
        selected_class_name = class_obj.name if class_obj else None
    
    selected_subject_name = None
    if selected_subject:
        subject_obj = Subject.objects.filter(id=selected_subject).first()
        selected_subject_name = subject_obj.name if subject_obj else None
    
    context = {
        'total_students': total_students,
        'total_exams': total_exams,
        'avg_score': avg_score,
        'pass_rate': pass_rate,
        'recent_exams': recent_exams,
        'current_year': current_year,
        'current_term': current_term,
        'selected_class': selected_class,
        'selected_subject': selected_subject,
        'selected_class_name': selected_class_name,
        'selected_subject_name': selected_subject_name,
        'classes': classes,
        'subjects': subjects,
        'available_years': available_years,
    }
    
    return render(request, 'performance/system_dashboard.html', context)
@staff_member_required
def student_performance(request, student_id):
    """View individual student performance"""
    student = get_object_or_404(Student, pk=student_id)
    results = StudentResult.objects.filter(student=student).select_related('exam', 'subject').order_by('-exam__academic_year', '-exam__term', 'subject__name')
    summaries = PerformanceSummary.objects.filter(student=student).order_by('-academic_year', '-term')
    
    all_scores = results.values_list('score', flat=True)
    overall_avg = sum(all_scores) / len(all_scores) if all_scores else 0
    
    context = {
        'student': student,
        'results': results,
        'summaries': summaries,
        'overall_avg': overall_avg,
        'school': SchoolSetting.objects.first(),
    }
    return render(request, 'performance/student_performance.html', context)


@staff_member_required
def enter_results(request):
    """Page for entering exam results with grading system selection"""
    from .models import GradingSystem, Exam, Subject, Student, StudentResult, TeacherGradingPreference, Class, SchoolSetting
    
    # Get all active grading systems
    all_grading_systems = GradingSystem.objects.filter(is_active=True).order_by('-is_default', 'name')
    
    print(f"\n📊 enter_results view - Found {all_grading_systems.count()} grading systems")
    
    # Get exam from GET parameter or session
    exam_id = request.GET.get('exam')
    
    # If no exam in GET, try to get from session
    if not exam_id:
        exam_id = request.session.get('exam_id')
    
    # If still no exam, redirect to exam list
    if not exam_id:
        messages.info(request, 'Please select an exam first')
        return redirect('digitallibrary:exam_list')
    
    try:
        exam = Exam.objects.get(id=exam_id)
        
        # Store exam in session for later
        request.session['exam_id'] = exam.id
        
        # Get students for this exam
        if exam.student_class:
            students = exam.student_class.students.filter(is_active=True).order_by('admission_number')
        else:
            students = Student.objects.filter(is_active=True).order_by('admission_number')
        
        # Get existing results
        existing_results = StudentResult.objects.filter(exam=exam)
        existing_scores = {r.student_id: {'score': r.score, 'grade': r.grade, 'points': r.points} for r in existing_results}
        
        # Get active grading system from session
        active_grading_system_id = request.session.get('active_grading_system_id')
        active_grading_system = None
        if active_grading_system_id and active_grading_system_id != 'traditional':
            try:
                active_grading_system = GradingSystem.objects.get(id=active_grading_system_id)
            except GradingSystem.DoesNotExist:
                pass
        
        # Get teacher's preference
        teacher_preference = TeacherGradingPreference.objects.filter(
            teacher=request.user,
            is_global=True
        ).first()
        
        # Get subjects for the dropdown (for navigation)
        subjects = Subject.objects.filter(is_active=True).order_by('name')
        selected_subject_id = request.GET.get('subject')
        selected_subject = None
        if selected_subject_id:
            try:
                selected_subject = Subject.objects.get(id=selected_subject_id)
                request.session['subject_id'] = selected_subject_id
            except Subject.DoesNotExist:
                pass
        
        context = {
            'exam': exam,
            'students': students,
            'existing_results': existing_results,
            'existing_scores': existing_scores,
            'all_grading_systems': all_grading_systems,
            'active_grading_system': active_grading_system,
            'teacher_preference': teacher_preference,
            'subjects': subjects,
            'selected_subject': selected_subject,
            'school': SchoolSetting.objects.first(),
        }
        return render(request, 'performance/enter_results_form.html', context)
        
    except Exam.DoesNotExist:
        messages.error(request, 'Exam not found')
        return redirect('digitallibrary:exam_list')

@staff_member_required
def enter_results_grid(request):
    """Redirect to the working enter_results_form with all parameters preserved"""
    from django.shortcuts import redirect
    from django.urls import reverse
    from django.contrib import messages
    from .models import Exam, Subject, TeacherGradingPreference, GradingSystem
    
    print("\n" + "="*60)
    print("🔍 DEBUG: enter_results_grid view called - Redirecting to form")
    print("="*60)
    
    # Get exam_id and subject_id from session
    exam_id = request.session.get('exam_id')
    subject_id = request.session.get('subject_id')
    grading_system_id = request.session.get('active_grading_system_id')
    
    print(f"   exam_id from session: {exam_id}")
    print(f"   subject_id from session: {subject_id}")
    print(f"   grading_system_id from session: {grading_system_id}")
    
    if not exam_id:
        messages.error(request, 'Please select an exam first.')
        return redirect('digitallibrary:enter_results')
    
    exam = get_object_or_404(Exam, id=exam_id)
    subject = None
    if subject_id:
        subject = get_object_or_404(Subject, id=subject_id)
    
    # Handle POST - Save grading preference before redirecting
    if request.method == 'POST':
        grading_choice = request.POST.get('grading_choice')
        
        if grading_choice == 'traditional':
            # Clear any saved preferences
            TeacherGradingPreference.objects.filter(
                teacher=request.user, exam=exam, subject=subject
            ).delete()
            request.session['active_grading_system_id'] = None
            messages.success(request, '✓ Using Traditional Grading System (A-E)')
            
        elif grading_choice == 'cbe':
            # Save CBE preference
            preference, created = TeacherGradingPreference.objects.get_or_create(
                teacher=request.user,
                exam=exam,
                subject=subject
            )
            preference.use_cbe_pathways = True
            preference.use_custom_grading = False
            preference.custom_grading_system = None
            preference.save()
            request.session['active_grading_system_id'] = 'cbe'
            messages.success(request, '✓ Using CBE (Competency-Based) Grading System')
            
        elif grading_choice == 'custom':
            custom_system_id = request.POST.get('custom_grading_system_id')
            if custom_system_id:
                custom_system = get_object_or_404(GradingSystem, id=custom_system_id)
                preference, created = TeacherGradingPreference.objects.get_or_create(
                    teacher=request.user,
                    exam=exam,
                    subject=subject
                )
                preference.use_cbe_pathways = False
                preference.use_custom_grading = True
                preference.custom_grading_system = custom_system
                preference.save()
                request.session['active_grading_system_id'] = custom_system.id
                messages.success(request, f'✓ Using {custom_system.name} Grading System')
        
        # After saving preference, redirect to enter_results_form
        url = reverse('digitallibrary:enter_results_form')
        params = f"?exam={exam_id}"
        if subject_id:
            params += f"&subject={subject_id}"
        
        full_url = f"{url}{params}"
        print(f"   POST redirect to: {full_url}")
        return redirect(full_url)
    
    # For GET requests, preserve the grading system in session and redirect
    if grading_system_id:
        request.session['active_grading_system_id'] = grading_system_id
    
    # Build redirect URL to enter_results_form
    url = reverse('digitallibrary:enter_results_form')
    params = f"?exam={exam_id}"
    
    if subject_id:
        params += f"&subject={subject_id}"
    
    full_url = f"{url}{params}"
    print(f"   GET redirect to: {full_url}")
    print("="*60 + "\n")
    
    messages.info(request, 'Redirecting to results entry form...')
    return redirect(full_url)
# digitallibrary/views.py - Add these views

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta
from .models import TVDisplay, TVContent, Announcement
from .forms import TVContentForm

# At the top of your views.py file, ensure these imports
from django.contrib.auth.decorators import user_passes_test, login_required
from django_tenants.utils import get_tenant
from django.utils import timezone
from datetime import timedelta
from .models import TVDisplay, TVContent, Announcement, SchoolSetting


def is_admin_or_principal(user):
    """Check if user is admin or principal"""
    if user.is_authenticated and (user.is_superuser or user.is_staff):
        return True
    if hasattr(user, 'profile'):
        return user.profile.role in ['admin', 'principal']
    return False

@login_required
@user_passes_test(is_admin_or_principal)
def tv_display(request):
    """Display the school TV interface - Professional news-style layout"""
    
    from django_tenants.utils import get_tenant
    from django.utils import timezone
    from datetime import timedelta
    from django.db import models  # Import Django's models here for Q
    from .models import TVDisplay, TVContent, Announcement, SchoolSetting
    
    school = get_tenant(request)
    
    # Get school settings for logo and branding
    school_settings = SchoolSetting.objects.first()
    
    # Get or create TV display for this school
    tv, created = TVDisplay.objects.get_or_create(
        school=school,
        defaults={
            'name': f"{school.name} TV",
            'is_active': True,
            'layout': 'split',
            'accent_color': '#bb1919',
            'background_color': '#0a0a0a',
            'refresh_interval': 30,
            'display_duration': 10,
            'show_clock': True,
            'show_weather': True,
            'show_news_ticker': True,
            'show_events': True,
            'show_exam_schedule': True,
            'show_noticeboard': True,
        }
    )
    
    # Use school logo from SchoolSetting if TV doesn't have its own
    if not tv.school_logo and school_settings and school_settings.logo:
        tv.school_logo = school_settings.logo
        tv.save()
    
    # Get motto from school settings with fallback
    school_motto = getattr(school_settings, 'motto', None) or "Excellence in Education"
    
    if not tv.is_active:
        return render(request, 'digitallibrary/tv/offline.html', {'school': school})
    
    # Get current content (next 30 days)
    now = timezone.now()
    future_date = now + timedelta(days=30)
    
    # Get TV-specific content
    tv_contents = TVContent.objects.filter(
        tv_display=tv,
        start_date__lte=future_date,
        is_active=True
    ).filter(
        models.Q(end_date__isnull=True) | models.Q(end_date__gte=now)  # Now models.Q works
    ).order_by('-priority', '-created_at')
    
    # Get breaking news (high priority or featured)
    breaking_news = tv_contents.filter(priority__gte=4, is_featured=True).first()
    
    # Get content from noticeboard if enabled
    noticeboard_contents = []
    if tv.show_noticeboard:
        noticeboard_contents = Announcement.objects.filter(
            expires_at__isnull=True
        ).order_by('-created_at')[:10]
    
    # Featured content (highest priority)
    featured = tv_contents.filter(is_featured=True, priority__gte=2).first()
    
    # If no featured content, get the most recent announcement
    if not featured:
        featured = tv_contents.filter(content_type='announcement').first()
    
    # Separate content by type
    announcements = tv_contents.filter(content_type='announcement')[:12]
    events = tv_contents.filter(content_type='event')[:8]
    exams = tv_contents.filter(content_type='exam')[:6]
    achievements = tv_contents.filter(content_type='achievement')[:6]
    
    # Ticker messages (all active content titles)
    ticker_messages = list(tv_contents.values_list('title', flat=True)[:15])
    
    # Add noticeboard titles to ticker
    for ann in noticeboard_contents[:5]:
        ticker_messages.append(ann.title)
    
    context = {
        'tv': tv,
        'school': school,
        'school_settings': school_settings,
        'school_motto': school_motto,  # Pass motto separately
        'layout': tv.layout,
        'accent_color': tv.accent_color,
        'background_color': tv.background_color,
        'text_color': tv.text_color,
        'refresh_interval': tv.refresh_interval,
        'display_duration': tv.display_duration,
        'show_clock': tv.show_clock,
        'show_weather': tv.show_weather,
        'show_news_ticker': tv.show_news_ticker,
        'footer_text': tv.footer_text,
        'breaking_news': breaking_news,
        'featured_content': featured,
        'announcements': announcements,
        'events': events,
        'exams': exams,
        'achievements': achievements,
        'noticeboard_contents': noticeboard_contents,
        'ticker_messages': ticker_messages,
        'tv_url': f"https://{request.get_host()}/app/tv/",
    }
    
    return render(request, 'digitallibrary/tv/display.html', context)

from django.contrib.auth.decorators import user_passes_test

# Define the permission check function
def is_admin_or_principal(user):
    """Check if user is Administrator or Principal"""
    if not user.is_authenticated:
        return False
    
    # Check by group
    if user.groups.filter(name__in=['Administrator', 'Principal']).exists():
        return True
    
    # Check by role field if your User model has it
    if hasattr(user, 'profile') and user.profile.role in ['administrator', 'principal']:
        return True
    
    # Check if user is superuser (optional - they can access everything)
    if user.is_superuser:
        return True
    
    return False

# Optional: Add a dashboard view for TV management
@login_required
@user_passes_test(is_admin_or_principal)
def tv_dashboard(request):
    """Admin/Principal dashboard for managing TV content"""

    from django_tenants.utils import get_tenant
    from django.contrib import messages
    from .models import TVDisplay, TVContent, SchoolSetting

    school = get_tenant(request)

    # Get or create TV display safely
    tv, created = TVDisplay.objects.get_or_create(
        school=school,
        defaults={
            "name": f"{school.name} TV",
            "is_active": True,
            "layout": "split",
            "accent_color": "#bb1919",
            "background_color": "#0a0a0a",
            "refresh_interval": 30,
            "show_clock": True,
            "show_news_ticker": True,
            "show_events": True,
            "show_exam_schedule": True,
        },
    )

    if created:
        messages.info(request, f"TV display created for {school.name}")

    # Important safety check
    if not tv.pk:
        tv.save()

    # Use tv_display_id to avoid unsaved-object related filter errors
    contents = TVContent.objects.filter(
        tv_display_id=tv.pk
    ).order_by("-created_at")

    user_role = "Admin"
    if hasattr(request.user, "profile"):
        user_role = request.user.profile.role.capitalize()
    elif request.user.groups.filter(name='Administrator').exists():
        user_role = "Administrator"
    elif request.user.groups.filter(name='Principal').exists():
        user_role = "Principal"

    host = request.get_host()
    tv_url = f"http://{host}/app/tv/"
    if request.is_secure():
        tv_url = f"https://{host}/app/tv/"

    context = {
        "tv": tv,
        "contents": contents,
        "recent_content": contents[:10],
        "contents_count": contents.count(),
        "active_contents": contents.filter(is_active=True).count(),
        "content_counts": {
            "total": contents.count(),
            "announcements": contents.filter(content_type="announcement").count(),
            "events": contents.filter(content_type="event").count(),
            "exams": contents.filter(content_type="exam").count(),
            "achievements": contents.filter(content_type="achievement").count(),
        },
        "tv_url": tv_url,
        "embed_code": f'<iframe src="{tv_url}" style="width:100%; height:100vh; border:none;"></iframe>',
        "user_role": user_role,
        "school": school,
        "school_settings": SchoolSetting.objects.first(),
    }

    return render(request, "digitallibrary/tv/dashboard.html", context)

@login_required
@user_passes_test(is_admin_or_principal, login_url="/app/login/")
def tv_content_add(request):
    """Add content to TV display"""

    from django_tenants.utils import get_tenant
    from django.contrib import messages
    from .models import TVDisplay, TVContent
    from .forms import TVContentForm

    school = get_tenant(request)

    tv = TVDisplay.objects.filter(
        school=school
    ).order_by("id").first()

    if tv is None:
        tv = TVDisplay.objects.create(
            school=school,
            name=f"{school.name} TV",
            is_active=True,
        )

    if request.method == "POST":
        form = TVContentForm(request.POST, request.FILES)

        if form.is_valid():
            content = form.save(commit=False)
            content.tv_display = tv
            content.created_by = request.user
            content.save()

            messages.success(
                request,
                f'✅ "{content.title}" added to TV successfully!'
            )

            return redirect("digitallibrary:tv_dashboard")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = TVContentForm()

    context = {
        "form": form,
        "tv": tv,
    }

    return render(request, "digitallibrary/tv/content_form.html", context)


@login_required
@user_passes_test(is_admin_or_principal)
def tv_content_edit(request, pk):
    """Edit TV content"""
    
    from django_tenants.utils import get_tenant
    from django.contrib import messages
    from .models import TVDisplay, TVContent
    from .forms import TVContentForm
    
    school = get_tenant(request)
    tv = TVDisplay.objects.get(school=school)
    
    content = get_object_or_404(TVContent, id=pk, tv_display=tv)
    
    if request.method == 'POST':
        form = TVContentForm(request.POST, request.FILES, instance=content)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ "{content.title}" updated successfully!')
            return redirect('digitallibrary:tv_dashboard')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = TVContentForm(instance=content)
    
    context = {
        'form': form,
        'content': content,
        'tv': tv,
    }
    return render(request, 'digitallibrary/tv/content_form.html', context)


@login_required
@user_passes_test(is_admin_or_principal, login_url='/app/login/')
def tv_content_delete(request, pk):
    """Delete TV content"""
    
    from django_tenants.utils import get_tenant
    from django.contrib import messages
    from .models import TVDisplay, TVContent
    
    school = get_tenant(request)
    tv = TVDisplay.objects.get(school=school)
    
    content = get_object_or_404(TVContent, id=pk, tv_display=tv)
    
    if request.method == 'POST':
        content_title = content.title
        content.delete()
        messages.success(request, f'🗑️ "{content_title}" deleted successfully!')
        return redirect('digitallibrary:tv_dashboard')
    
    return redirect('digitallibrary:tv_dashboard')


@login_required
@user_passes_test(is_admin_or_principal, login_url='/app/login/')
def tv_settings(request):
    """Update TV settings"""
    
    from django_tenants.utils import get_tenant
    from django.contrib import messages
    from .models import TVDisplay
    
    school = get_tenant(request)
    tv = TVDisplay.objects.get(school=school)
    
    if request.method == 'POST':
        tv.layout = request.POST.get('layout', tv.layout)
        tv.accent_color = request.POST.get('accent_color', tv.accent_color)
        tv.background_color = request.POST.get('background_color', tv.background_color)
        tv.refresh_interval = int(request.POST.get('refresh_interval', tv.refresh_interval))
        tv.display_duration = int(request.POST.get('display_duration', tv.display_duration))
        tv.show_clock = request.POST.get('show_clock') == 'on'
        tv.show_weather = request.POST.get('show_weather') == 'on'
        tv.show_news_ticker = request.POST.get('show_news_ticker') == 'on'
        tv.show_events = request.POST.get('show_events') == 'on'
        tv.show_exam_schedule = request.POST.get('show_exam_schedule') == 'on'
        tv.show_noticeboard = request.POST.get('show_noticeboard') == 'on'
        tv.footer_text = request.POST.get('footer_text', tv.footer_text)
        tv.save()
        
        messages.success(request, '✅ TV settings updated successfully!')
        return redirect('digitallibrary:tv_dashboard')
    
    return redirect('digitallibrary:tv_dashboard')


@login_required
@user_passes_test(is_admin_or_principal, login_url='/app/login/')
def tv_upload_logo(request):
    """Upload school logo for TV"""
    
    from django_tenants.utils import get_tenant
    from django.contrib import messages
    from .models import TVDisplay
    
    school = get_tenant(request)
    tv = TVDisplay.objects.get(school=school)
    
    if request.method == 'POST' and request.FILES.get('logo'):
        tv.school_logo = request.FILES['logo']
        tv.save()
        messages.success(request, '✅ School logo uploaded successfully!')
    else:
        messages.error(request, 'Please select a valid image file.')
    
    return redirect('digitallibrary:tv_dashboard')


@login_required
@user_passes_test(is_admin_or_principal, login_url='/app/login/')
def tv_remove_logo(request):
    """Remove school logo from TV"""
    
    from django_tenants.utils import get_tenant
    from django.contrib import messages
    from .models import TVDisplay
    
    school = get_tenant(request)
    tv = TVDisplay.objects.get(school=school)
    
    if request.method == 'POST':
        if tv.school_logo:
            tv.school_logo.delete()
            tv.school_logo = None
            tv.save()
        messages.success(request, '✅ School logo removed.')
    
    return redirect('digitallibrary:tv_dashboard')

def api_tv_content(request):
    """API endpoint for TV content (for AJAX refresh)"""
    
    from django_tenants.utils import get_tenant
    from django.http import JsonResponse
    
    school = get_tenant(request)
    tv = TVDisplay.objects.get(school=school)
    
    now = timezone.now()
    contents = TVContent.objects.filter(
        tv_display=tv,
        start_date__lte=now,
        is_active=True
    ).filter(
        models.Q(end_date__isnull=True) | models.Q(end_date__gte=now)
    ).order_by('-priority', '-created_at')
    
    data = {
        'contents': [
            {
                'id': c.id,
                'title': c.title,
                'message': c.message,
                'content_type': c.content_type,
                'priority': c.priority,
                'image_url': c.image.url if c.image else None,
            } for c in contents[:20]
        ],
        'refresh_interval': tv.refresh_interval,
    }
    
    return JsonResponse(data)


@staff_member_required
def set_grading_preference(request, exam_id):
    """Set the grading system preference for this exam session"""
    from django.shortcuts import redirect
    from django.contrib import messages
    from .models import GradingSystem, TeacherGradingPreference
    
    print(f"\n{'='*60}")
    print(f"🔧 set_grading_preference called")
    print(f"   exam_id: {exam_id}")
    print(f"   Method: {request.method}")
    print(f"   POST params: {dict(request.POST)}")
    print(f"{'='*60}")
    
    if request.method == 'POST':
        grading_system_id = request.POST.get('grading_system_id')
        subject_id = request.GET.get('subject')
        
        print(f"   grading_system_id: '{grading_system_id}'")
        print(f"   subject_id: '{subject_id}'")
        
        # Get or create teacher preference
        preference, created = TeacherGradingPreference.objects.get_or_create(
            teacher=request.user,
            exam_id=exam_id
        )
        
        if grading_system_id == 'cbe':
            # Use CBE grading
            preference.use_cbe_pathways = True
            preference.use_custom_grading = False
            preference.custom_grading_system = None
            request.session['active_grading_system_id'] = 'cbe'
            messages.success(request, '✓ CBE Grading System Activated (EE1, EE2, ME1, ME2, AE2, AE1, BE2, BE1)')
            print(f"   Set session: active_grading_system_id = 'cbe'")
            
        elif grading_system_id == 'traditional':
            # Use traditional grading
            preference.use_cbe_pathways = False
            preference.use_custom_grading = False
            preference.custom_grading_system = None
            request.session['active_grading_system_id'] = None
            messages.success(request, '✓ Traditional Grading System (KCSE) Activated')
            print(f"   Set session: active_grading_system_id = None")
            
        elif grading_system_id:
            try:
                # Try to get custom grading system by ID
                grading_system = GradingSystem.objects.get(id=int(grading_system_id), is_active=True)
                preference.use_cbe_pathways = False
                preference.use_custom_grading = True
                preference.custom_grading_system = grading_system
                request.session['active_grading_system_id'] = grading_system.id
                messages.success(request, f'✓ {grading_system.name} Grading System Activated')
                print(f"   Set session: active_grading_system_id = {grading_system.id}")
            except (GradingSystem.DoesNotExist, ValueError) as e:
                messages.error(request, f'Selected grading system not found')
                print(f"   ERROR: {e}")
        else:
            messages.error(request, 'Please select a grading system')
            print(f"   ERROR: No grading_system_id provided")
        
        preference.save()
        
        # Redirect back to the results entry form
        if subject_id and subject_id != 'None':
            redirect_url = f'/app/enter-results-form/?exam={exam_id}&subject={subject_id}'
        else:
            redirect_url = f'/app/enter-results-form/?exam={exam_id}'
        
        print(f"   Redirecting to: {redirect_url}")
        print(f"{'='*60}\n")
        return redirect(redirect_url)
    
    # For GET requests, just redirect to the form
    subject_id = request.GET.get('subject')
    if subject_id and subject_id != 'None':
        return redirect(f'/app/enter-results-form/?exam={exam_id}&subject={subject_id}')
    return redirect(f'/app/enter-results-form/?exam={exam_id}')
# ========== ENHANCED STUDENT ANALYTICS VIEWS ==========

@staff_member_required
def student_analytics(request, student_id):
    """Comprehensive student performance analytics with trends, recommendations, and parent summary"""
    from collections import defaultdict
    from statistics import mean, pstdev
    
    student = get_object_or_404(Student, pk=student_id)

    summaries = list(
        PerformanceSummary.objects.filter(student=student)
        .select_related("student")
        .order_by("academic_year", "term")
    )

    results = list(
        StudentResult.objects.filter(student=student)
        .select_related("exam", "subject", "student")
        .order_by("exam__academic_year", "exam__term", "exam__name", "subject__name")
    )

    latest_summary = summaries[-1] if summaries else None
    previous_summary = summaries[-2] if len(summaries) > 1 else None

    latest_avg = round(float(latest_summary.average_score), 1) if latest_summary else 0.0
    previous_avg = round(float(previous_summary.average_score), 1) if previous_summary else 0.0
    avg_change = round(latest_avg - previous_avg, 1) if previous_summary else 0.0
    performance_status = _status_from_change(avg_change)
    overall_grade = _grade_from_score_analytics(latest_avg)

    latest_rank = latest_summary.rank_in_class if latest_summary and latest_summary.rank_in_class else None
    previous_rank = previous_summary.rank_in_class if previous_summary and previous_summary.rank_in_class else None
    rank_change = (previous_rank - latest_rank) if latest_rank and previous_rank else None

    class_average = 0.0
    class_rank = latest_rank

    if latest_summary and getattr(student, "current_class", None):
        class_average_value = (
            PerformanceSummary.objects.filter(
                student__current_class=student.current_class,
                academic_year=latest_summary.academic_year,
                term=latest_summary.term,
            ).aggregate(avg=Avg("average_score"))["avg"]
            or 0
        )
        class_average = round(float(class_average_value), 1)

    trend_labels = []
    trend_scores = []
    overall_scores = []

    terms_data = {}
    for summary in summaries:
        label = f"Term {summary.term} - {summary.academic_year}"
        score = round(float(summary.average_score), 1)
        trend_labels.append(label)
        trend_scores.append(score)
        overall_scores.append(score)

        terms_data[label] = {
            "average": score,
            "term": summary.term,
            "year": summary.academic_year,
            "rank": summary.rank_in_class,
            "grade": summary.overall_grade or _grade_from_score_analytics(score),
            "points": round(float(summary.average_points), 1) if summary.average_points is not None else 0.0,
            "subjects": [],
        }

    consistency_label, score_deviation = _consistency_label(overall_scores)

    subject_history = defaultdict(list)
    exams_data = defaultdict(lambda: {"exam": None, "subjects": [], "average": 0.0})

    for result in results:
        term_label = f"T{result.exam.term} {result.exam.academic_year}"
        score = round(float(result.score), 1)

        subject_history[result.subject.name].append(
            {
                "term": result.exam.term,
                "year": result.exam.academic_year,
                "term_label": term_label,
                "exam_name": result.exam.name,
                "score": score,
            }
        )

        exam_key = result.exam.id
        exams_data[exam_key]["exam"] = result.exam
        exams_data[exam_key]["subjects"].append(
            {
                "name": result.subject.name,
                "score": score,
            }
        )

        summary_key = f"Term {result.exam.term} - {result.exam.academic_year}"
        if summary_key in terms_data:
            terms_data[summary_key]["subjects"].append(
                {
                    "name": result.subject.name,
                    "score": score,
                }
            )

    for exam_id, exam_block in exams_data.items():
        subject_scores = [item["score"] for item in exam_block["subjects"]]
        exam_block["average"] = _safe_mean(subject_scores)

    subject_analysis = []
    subject_comparison = []
    subject_trends = {}
    subjects_data = {}
    improving_subjects = []
    declining_subjects = []
    risk_subjects = []
    strong_subjects = []

    for subject_name, items in subject_history.items():
        items = sorted(items, key=lambda x: (int(x["year"]), int(x["term"]), x["exam_name"]))
        subjects_data[subject_name] = items

        scores = [item["score"] for item in items]
        first_score = round(scores[0], 1)
        latest_score = round(scores[-1], 1)
        best_score = round(max(scores), 1)
        lowest_score = round(min(scores), 1)
        average_score = _safe_mean(scores)
        subject_change = round(latest_score - first_score, 1) if len(scores) >= 2 else 0.0
        subject_status = _status_from_change(subject_change) if len(scores) >= 2 else "Stable"

        latest_item = items[-1]
        latest_term = latest_item["term"]
        latest_year = latest_item["year"]

        class_avg_value = (
            StudentResult.objects.filter(
                subject__name=subject_name,
                exam__term=latest_term,
                exam__academic_year=latest_year,
                student__current_class=student.current_class,
            ).aggregate(avg=Avg("score"))["avg"]
            or 0
        )
        class_avg = round(float(class_avg_value), 1)
        difference = round(latest_score - class_avg, 1)

        subject_analysis.append(
            {
                "name": subject_name,
                "first_score": first_score,
                "latest_score": latest_score,
                "best_score": best_score,
                "lowest_score": lowest_score,
                "average_score": average_score,
                "change": subject_change,
                "status": subject_status,
                "grade": _grade_from_score_analytics(latest_score),
                "class_average": class_avg,
                "class_diff": difference,
                "better_than_class": difference >= 0,
                "at_risk": latest_score < 50,
                "strong": latest_score >= 70,
            }
        )

        subject_comparison.append(
            {
                "name": subject_name,
                "student_avg": latest_score,
                "class_avg": class_avg,
                "difference": difference,
                "better": difference >= 0,
            }
        )

        subject_trends[subject_name] = {
            "first": first_score,
            "last": latest_score,
            "improvement": subject_change,
        }

        if len(scores) >= 2:
            if subject_change >= 3:
                improving_subjects.append(subject_name)
            elif subject_change <= -3:
                declining_subjects.append(subject_name)

        if latest_score < 50:
            risk_subjects.append(subject_name)
        if latest_score >= 70:
            strong_subjects.append(subject_name)

    subject_analysis.sort(key=lambda x: x["latest_score"], reverse=True)
    subject_comparison.sort(key=lambda x: x["difference"], reverse=True)

    teacher_recommendations = []
    parent_recommendations = []

    if risk_subjects:
        teacher_recommendations.append(
            f"Provide immediate remediation in {', '.join(risk_subjects[:3])} and review recent assessment errors."
        )
        parent_recommendations.append(
            f"Set aside extra weekly revision time for {', '.join(risk_subjects[:3])} using short and regular practice."
        )

    if declining_subjects:
        teacher_recommendations.append(
            f"Closely monitor declining performance in {', '.join(declining_subjects[:3])} and compare classwork against exam performance."
        )
        parent_recommendations.append(
            f"Discuss challenges in {', '.join(declining_subjects[:3])} and track homework completion more closely."
        )

    if strong_subjects:
        teacher_recommendations.append(
            f"Extend learning in {', '.join(strong_subjects[:3])} through more challenging class activities."
        )
        parent_recommendations.append(
            f"Maintain motivation in {', '.join(strong_subjects[:3])} with praise and consistent revision routines."
        )

    if not teacher_recommendations:
        teacher_recommendations.append("Maintain current support, continue tracking progress, and reinforce good study habits.")

    if not parent_recommendations:
        parent_recommendations.append("Maintain a consistent study routine and review school feedback regularly.")

    parent_summary = _build_parent_summary(
        student_name=f"{student.first_name} {student.last_name}",
        latest_avg=latest_avg,
        avg_change=avg_change,
        performance_status=performance_status,
        strong_subjects=strong_subjects,
        risk_subjects=risk_subjects,
        declining_subjects=declining_subjects,
    )

    context = {
        "student": student,
        "summaries": list(reversed(summaries)),
        "results": results,
        "total_terms": len(summaries),
        "overall_avg": latest_avg,
        "overall_grade": overall_grade,
        "class_rank": class_rank,
        "class_average": class_average,
        "latest_avg": latest_avg,
        "previous_avg": previous_avg,
        "avg_change": avg_change,
        "performance_status": performance_status,
        "latest_rank": latest_rank,
        "previous_rank": previous_rank,
        "rank_change": rank_change,
        "trend_labels": trend_labels,
        "trend_scores": trend_scores,
        "terms_data": terms_data,
        "exams_data": dict(exams_data),
        "subjects_data": dict(subjects_data),
        "subject_analysis": subject_analysis,
        "subject_comparison": subject_comparison,
        "subject_trends": subject_trends,
        "improving_subjects": improving_subjects,
        "declining_subjects": declining_subjects,
        "risk_subjects": risk_subjects,
        "strong_subjects": strong_subjects,
        "consistency_label": consistency_label,
        "score_deviation": score_deviation,
        "teacher_recommendations": teacher_recommendations,
        "parent_recommendations": parent_recommendations,
        "parent_summary": parent_summary,
        "school": SchoolSetting.objects.first(),
    }
    return render(request, "digitallibrary/student_analytics.html", context)


# ========== CLASS PERFORMANCE ANALYTICS VIEW ==========

@staff_member_required
def class_performance_analytics(request, class_id):
    """Class-level performance analytics with metrics and graphs"""
    import json
    
    student_class = get_object_or_404(Class, pk=class_id)
    academic_year = request.GET.get('year', str(timezone.now().year))
    
    students = Student.objects.filter(current_class=student_class, is_active=True)
    summaries = PerformanceSummary.objects.filter(
        student__in=students,
        academic_year=academic_year
    ).select_related('student').order_by('term', 'rank_in_class')
    
    term_data = {}
    for term in [1, 2, 3]:
        term_summaries = summaries.filter(term=term)
        if term_summaries.exists():
            term_data[f'Term {term}'] = {
                'average': term_summaries.aggregate(Avg('average_score'))['average_score__avg'] or 0,
                'pass_rate': term_summaries.filter(average_score__gte=50).count() / term_summaries.count() * 100 if term_summaries.count() > 0 else 0,
                'top_student': term_summaries.order_by('-average_score').first().student.get_full_name() if term_summaries.exists() else 'N/A',
                'top_score': term_summaries.aggregate(Max('average_score'))['average_score__max'] or 0,
            }
    
    latest_term = summaries.aggregate(Max('term'))['term__max'] or 1
    latest_summaries = summaries.filter(term=latest_term)
    grade_distribution = {
        'A': latest_summaries.filter(overall_grade='A').count(),
        'A-': latest_summaries.filter(overall_grade='A-').count(),
        'B+': latest_summaries.filter(overall_grade='B+').count(),
        'B': latest_summaries.filter(overall_grade='B').count(),
        'B-': latest_summaries.filter(overall_grade='B-').count(),
        'C+': latest_summaries.filter(overall_grade='C+').count(),
        'C': latest_summaries.filter(overall_grade='C').count(),
        'C-': latest_summaries.filter(overall_grade='C-').count(),
        'D+': latest_summaries.filter(overall_grade='D+').count(),
        'D': latest_summaries.filter(overall_grade='D').count(),
        'E': latest_summaries.filter(overall_grade='E').count(),
    }
    
    top_students = summaries.filter(term=latest_term).order_by('rank_in_class')[:10]
    subject_performance = []
    subjects = Subject.objects.filter(is_active=True)
    
    for subject in subjects:
        results = StudentResult.objects.filter(
            exam__subject=subject,
            exam__academic_year=academic_year,
            student__in=students
        )
        if results.exists():
            avg_score = results.aggregate(Avg('score'))['score__avg'] or 0
            pass_count = results.filter(score__gte=50).values('student').distinct().count()
            subject_performance.append({
                'name': subject.name,
                'average': round(avg_score, 1),
                'pass_rate': round(pass_count / students.count() * 100, 1) if students.count() > 0 else 0,
                'students': results.values('student').distinct().count()
            })
    
    term_labels = list(term_data.keys())
    term_averages = [term_data[t]['average'] for t in term_labels]
    term_pass_rates = [term_data[t]['pass_rate'] for t in term_labels]
    
    context = {
        'class': student_class,
        'students': students,
        'summaries': summaries,
        'term_data': term_data,
        'term_labels': json.dumps(term_labels),
        'term_averages': json.dumps(term_averages),
        'term_pass_rates': json.dumps(term_pass_rates),
        'subject_performance': subject_performance,
        'grade_distribution': grade_distribution,
        'top_students': top_students,
        'academic_year': academic_year,
        'total_students': students.count(),
        'latest_term': latest_term,
        'title': f'Class Performance - {student_class.name}',
        'school': SchoolSetting.objects.first(),
    }
    return render(request, 'performance/class_analytics.html', context)


# ========== STUDENT REPORT CARD VIEW ==========

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django_tenants.utils import get_tenant
from datetime import datetime
from .models import Student, Exam, StudentResult, PerformanceSummary, SchoolSetting

# ========== STUDENT REPORT CARD VIEW ==========

def student_report_card(request, student_id, exam_id=None):
    """Generate a printable report card for a student"""
    
    # Get the student
    student = get_object_or_404(Student, pk=student_id, is_active=True)
    
    # Get exam
    if exam_id:
        exam = get_object_or_404(Exam, pk=exam_id)
    else:
        exam = Exam.objects.filter(
            student_class=student.current_class,
            is_active=True
        ).order_by('-academic_year', '-term').first()
    
    if not exam:
        messages.error(request, "No exam results available for this student.")
        return redirect('digitallibrary:student_performance', student_id=student.id)
    
    # Get results
    results = StudentResult.objects.filter(
        student=student,
        exam=exam
    ).select_related('subject').order_by('subject__name')
    
    if not results.exists():
        messages.warning(request, "No subject results found for this exam.")
        return redirect('digitallibrary:student_performance', student_id=student.id)
    
    # Calculate totals
    total_marks = sum(float(r.score) for r in results)
    overall_average = total_marks / len(results) if results else 0
    
    # Get tenant
    tenant = get_tenant(request)
    
    context = {
        'student': student,
        'exam': exam,
        'results': results,
        'total_marks': total_marks,
        'overall_average': overall_average,
        'tenant': tenant,
        'current_date': timezone.now(),
    }
    
    return render(request, 'performance/student_report_card.html', context)
# ========== BULK RESULTS ENTRY VIEWS ==========



@staff_member_required
def bulk_excel_process(request):
    """Process the uploaded Excel file and save results"""
    from .models import Exam, Subject, Student, StudentResult
    from django.db import connection
    import pandas as pd
    import os
    import tempfile
    
    exam_id = request.session.get('bulk_exam_id')
    subject_id = request.session.get('bulk_subject_id')
    grading_system = request.session.get('bulk_grading_system', 'cbe')
    file_path = request.session.get('bulk_file_path')
    
    if not all([exam_id, subject_id, file_path]):
        messages.error(request, 'Missing required data. Please start over.')
        return redirect('digitallibrary:bulk_enter_results')
    
    try:
        exam = Exam.objects.get(id=exam_id)
        subject = Subject.objects.get(id=subject_id)
        use_cbe = grading_system == 'cbe'
        
        # Read file
        try:
            df = pd.read_excel(file_path)
        except:
            df = pd.read_csv(file_path)
        
        # Clean up temp file
        try:
            os.remove(file_path)
        except:
            pass
        
        # Clear session data
        del request.session['bulk_exam_id']
        del request.session['bulk_subject_id']
        del request.session['bulk_grading_system']
        del request.session['bulk_file_path']
        
        # Normalize columns
        df.columns = df.columns.str.strip().str.lower()
        
        # Find admission column
        admission_col = None
        score_col = None
        
        for col in df.columns:
            if 'admission' in col or 'adm' in col or 'reg' in col:
                admission_col = col
            elif 'score' in col or 'mark' in col or 'result' in col:
                score_col = col
        
        if admission_col is None or score_col is None:
            messages.error(request, 'Excel file must have "Admission Number" and "Score" columns')
            return redirect('digitallibrary:bulk_enter_results')
        
        results_processed = 0
        errors = []
        max_score = float(exam.max_score) if exam.max_score else 100.0
        
        with connection.cursor() as cursor:
            for index, row in df.iterrows():
                admission_number = str(row[admission_col]).strip() if pd.notna(row[admission_col]) else None
                score_value = row[score_col] if pd.notna(row[score_col]) else None
                
                if not admission_number or score_value is None:
                    continue
                
                try:
                    score = float(score_value)
                    
                    if score < 0 or score > max_score:
                        errors.append(f"Row {index + 2}: Score {score} is outside valid range (0-{max_score})")
                        continue
                    
                    # Get student
                    student = Student.objects.filter(admission_number=admission_number, is_active=True).first()
                    if not student:
                        errors.append(f"Row {index + 2}: Student with admission number '{admission_number}' not found")
                        continue
                    
                    if use_cbe:
                        # Get CBE grade
                        cursor.execute("""
                            SELECT id, points FROM digitallibrary_kneccbegrade 
                            WHERE min_score <= %s AND max_score >= %s
                            LIMIT 1
                        """, [score, score])
                        grade = cursor.fetchone()
                        if grade:
                            grade_id, points = grade
                            cursor.execute("""
                                INSERT INTO digitallibrary_studentresult 
                                (student_id, exam_id, subject_id, score, grade_id, points, entered_by_id, entered_at, updated_at)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                                ON CONFLICT (student_id, exam_id, subject_id) 
                                DO UPDATE SET 
                                    score = EXCLUDED.score,
                                    grade_id = EXCLUDED.grade_id,
                                    points = EXCLUDED.points,
                                    updated_at = NOW()
                            """, [student.id, exam.id, subject.id, score, grade_id, points, request.user.id])
                            results_processed += 1
                    else:
                        # Traditional grading
                        percentage = (score / max_score) * 100
                        if percentage >= 80: grade_name = 'A'; points = 12
                        elif percentage >= 75: grade_name = 'A-'; points = 11
                        elif percentage >= 70: grade_name = 'B+'; points = 10
                        elif percentage >= 65: grade_name = 'B'; points = 9
                        elif percentage >= 60: grade_name = 'B-'; points = 8
                        elif percentage >= 55: grade_name = 'C+'; points = 7
                        elif percentage >= 50: grade_name = 'C'; points = 6
                        elif percentage >= 45: grade_name = 'C-'; points = 5
                        elif percentage >= 40: grade_name = 'D+'; points = 4
                        elif percentage >= 35: grade_name = 'D'; points = 3
                        elif percentage >= 30: grade_name = 'D-'; points = 2
                        else: grade_name = 'E'; points = 1
                        
                        cursor.execute("""
                            SELECT id FROM digitallibrary_grade WHERE grade = %s LIMIT 1
                        """, [grade_name])
                        grade = cursor.fetchone()
                        grade_id = grade[0] if grade else None
                        
                        cursor.execute("""
                            INSERT INTO digitallibrary_studentresult 
                            (student_id, exam_id, subject_id, score, grade_id, points, entered_by_id, entered_at, updated_at)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                            ON CONFLICT (student_id, exam_id, subject_id) 
                            DO UPDATE SET 
                                score = EXCLUDED.score,
                                grade_id = EXCLUDED.grade_id,
                                points = EXCLUDED.points,
                                updated_at = NOW()
                        """, [student.id, exam.id, subject.id, score, grade_id, points, request.user.id])
                        results_processed += 1
                        
                except ValueError:
                    errors.append(f"Row {index + 2}: Invalid score value '{score_value}'")
                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
        
        if results_processed > 0:
            messages.success(request, f'✅ Successfully processed {results_processed} results for {exam.name} - {subject.name}')
        else:
            messages.warning(request, '⚠️ No valid results were found in the file.')
        
        if errors:
            for error in errors[:10]:
                messages.warning(request, error)
            if len(errors) > 10:
                messages.info(request, f'... and {len(errors) - 10} more errors')
                
        return redirect('digitallibrary:exam_list')
        
    except Exam.DoesNotExist:
        messages.error(request, 'Selected exam not found')
    except Subject.DoesNotExist:
        messages.error(request, 'Selected subject not found')
    except Exception as e:
        messages.error(request, f'Error processing file: {str(e)}')
    
    return redirect('digitallibrary:bulk_enter_results')


def bulk_results_entry(request, exam_id, subject_id):
    """Step 2: Enter results for all students in a table"""
    from .models import Exam, Subject, Student
    
    exam = Exam.objects.get(id=exam_id)
    subject = Subject.objects.get(id=subject_id)
    
    # Get students
    if exam.student_class:
        students = exam.student_class.students.filter(is_active=True)
    else:
        students = Student.objects.filter(is_active=True)
    
    students = students.order_by('first_name', 'last_name')
    
    # Get existing results
    existing_results = {}
    try:
        from .models import Result
        results = Result.objects.filter(exam=exam, subject=subject, student__in=students)
        existing_results = {r.student_id: r for r in results}
    except ImportError:
        pass
    
    if request.method == 'POST':
        saved_count = 0
        for key, value in request.POST.items():
            if key.startswith('score_') and value:
                student_id = key.replace('score_', '')
                try:
                    score = float(value)
                    student = Student.objects.get(id=student_id)
                    
                    try:
                        from .models import Result
                        result, created = Result.objects.update_or_create(
                            exam=exam,
                            subject=subject,
                            student=student,
                            defaults={'score': score}
                        )
                        saved_count += 1
                    except ImportError:
                        saved_count += 1
                except (ValueError, Student.DoesNotExist):
                    continue
        
        messages.success(request, f'Successfully saved {saved_count} results for {subject.name}')
        return redirect('digitallibrary:bulk_results_entry', exam_id=exam.id, subject_id=subject.id)
    
    context = {
        'exam': exam,
        'subject': subject,
        'students': students,
        'existing_results': existing_results,
    }
    
    return render(request, 'digitallibrary/bulk_results_entry.html', context)

# ========== BULK EXCEL UPLOAD VIEW ==========

@staff_member_required
def bulk_excel_upload(request):
    """Upload Excel file with results for bulk entry"""
    from .forms import ExcelResultsUploadForm
    
    if request.method == 'POST':
        form = ExcelResultsUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            exam = form.cleaned_data['exam']
            subject = form.cleaned_data['subject']
            
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
                
                # Find column indices
                admission_col = None
                score_col = None
                
                for idx, cell in enumerate(sheet[1], 1):
                    header = str(cell.value).lower().strip() if cell.value else ''
                    if 'admission' in header or 'adm' in header or 'reg' in header:
                        admission_col = idx
                    elif 'score' in header or 'mark' in header or 'result' in header:
                        score_col = idx
                
                if admission_col is None or score_col is None:
                    messages.error(request, 'Excel file must have "Admission Number" and "Score" columns')
                    return redirect('digitallibrary:bulk_excel_upload')
                
                results_processed = 0
                errors = []
                
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not row:
                        continue
                    
                    admission_number = str(row[admission_col - 1]).strip() if row[admission_col - 1] else None
                    score_value = row[score_col - 1] if score_col - 1 < len(row) else None
                    
                    if not admission_number or score_value is None:
                        continue
                    
                    try:
                        score = float(score_value)
                        max_score = float(exam.max_score) if exam.max_score else 100.0
                        
                        if score < 0 or score > max_score:
                            errors.append(f"Row {row_idx}: Score {score} is outside valid range (0-{max_score})")
                            continue
                        
                        student = Student.objects.get(admission_number=admission_number, is_active=True)
                        
                        StudentResult.objects.update_or_create(
                            student=student,
                            exam=exam,
                            subject=subject,
                            defaults={'score': score, 'entered_by': request.user}
                        )
                        results_processed += 1
                        
                    except Student.DoesNotExist:
                        errors.append(f"Row {row_idx}: Student with admission number '{admission_number}' not found")
                    except ValueError:
                        errors.append(f"Row {row_idx}: Invalid score value '{score_value}'")
                    except Exception as e:
                        errors.append(f"Row {row_idx}: {str(e)}")
                
                if results_processed > 0:
                    messages.success(request, f'Successfully processed {results_processed} results for {exam.name} - {subject.name}')
                    return redirect('digitallibrary:exam_results_entry', exam_id=exam.id)
                
                if errors:
                    for error in errors[:5]:
                        messages.warning(request, error)
                    if len(errors) > 5:
                        messages.warning(request, f'... and {len(errors) - 5} more errors')
                
                if results_processed == 0:
                    messages.warning(request, 'No valid results were found in the file.')
                    
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
                return redirect('digitallibrary:bulk_excel_upload')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ExcelResultsUploadForm()
    
    context = {
        'form': form,
        'title': 'Bulk Excel Upload',
        'school': SchoolSetting.objects.first(),
    }
    return render(request, 'performance/bulk_excel_upload.html', context)
# ========== BULK RESULTS ENTRY VIEWS ==========

@staff_member_required
def bulk_enter_results(request):
    """
    Step 1: Select exam, subject, and upload file for bulk entry
    """
    from .models import Exam, Subject, SchoolSetting
    import pandas as pd
    import os
    import tempfile
    from django.db import connection
    
    print("\n" + "="*60)
    print("🔵 bulk_enter_results view called")
    print(f"   Method: {request.method}")
    print("="*60)
    
    exams = Exam.objects.filter(is_active=True).order_by('-academic_year', '-created_at')
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    
    if request.method == 'POST':
        exam_id = request.POST.get('exam')
        subject_id = request.POST.get('subject')
        grading_system = request.POST.get('grading_system', 'cbe')
        excel_file = request.FILES.get('excel_file')
        
        print(f"📝 POST data received:")
        print(f"   exam_id: {exam_id}")
        print(f"   subject_id: {subject_id}")
        print(f"   grading_system: {grading_system}")
        print(f"   file: {excel_file.name if excel_file else 'None'}")
        
        if not exam_id or not subject_id or not excel_file:
            messages.error(request, 'Please select exam, subject and upload a file')
            return redirect('digitallibrary:bulk_enter_results')
        
        try:
            exam = Exam.objects.get(id=exam_id)
            subject = Subject.objects.get(id=subject_id)
            use_cbe = grading_system == 'cbe'
            
            print(f"✅ Found exam: {exam.name}")
            print(f"✅ Found subject: {subject.name}")
            
            # Read file
            ext = excel_file.name.split('.')[-1].lower()
            try:
                if ext == 'csv':
                    df = pd.read_csv(excel_file)
                else:
                    df = pd.read_excel(excel_file)
            except Exception as e:
                messages.error(request, f'Error reading file: {str(e)}')
                return redirect('digitallibrary:bulk_enter_results')
            
            # Normalize columns
            df.columns = df.columns.str.strip().str.lower()
            print(f"📊 Columns found: {list(df.columns)}")
            
            # Find admission column
            admission_col = None
            score_col = None
            
            for col in df.columns:
                if 'admission' in col or 'adm' in col or 'reg' in col or 'student' in col:
                    admission_col = col
                elif 'score' in col or 'mark' in col or 'result' in col:
                    score_col = col
            
            if admission_col is None or score_col is None:
                messages.error(request, 'Excel file must have "Admission Number" and "Score" columns')
                return redirect('digitallibrary:bulk_enter_results')
            
            results_processed = 0
            errors = []
            max_score = float(exam.max_score) if exam.max_score else 100.0
            
            with connection.cursor() as cursor:
                for index, row in df.iterrows():
                    admission_number = str(row[admission_col]).strip() if pd.notna(row[admission_col]) else None
                    score_value = row[score_col] if pd.notna(row[score_col]) else None
                    
                    if not admission_number or score_value is None:
                        continue
                    
                    try:
                        score = float(score_value)
                        
                        if score < 0 or score > max_score:
                            errors.append(f"Row {index + 2}: Score {score} out of range (0-{max_score})")
                            continue
                        
                        # Get student
                        student = Student.objects.filter(admission_number=admission_number, is_active=True).first()
                        if not student:
                            errors.append(f"Row {index + 2}: Student '{admission_number}' not found")
                            continue
                        
                        if use_cbe:
                            # Get CBE grade
                            cursor.execute("""
                                SELECT id, points FROM digitallibrary_kneccbegrade 
                                WHERE min_score <= %s AND max_score >= %s
                                LIMIT 1
                            """, [score, score])
                            grade = cursor.fetchone()
                            if grade:
                                grade_id, points = grade
                                cursor.execute("""
                                    INSERT INTO digitallibrary_studentresult 
                                    (student_id, exam_id, subject_id, score, grade_id, points, entered_by_id, entered_at, updated_at)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                                    ON CONFLICT (student_id, exam_id, subject_id) 
                                    DO UPDATE SET 
                                        score = EXCLUDED.score,
                                        grade_id = EXCLUDED.grade_id,
                                        points = EXCLUDED.points,
                                        updated_at = NOW()
                                """, [student.id, exam.id, subject.id, score, grade_id, points, request.user.id])
                                results_processed += 1
                        else:
                            # Traditional grading
                            percentage = (score / max_score) * 100
                            if percentage >= 80: grade_name = 'A'; points = 12
                            elif percentage >= 75: grade_name = 'A-'; points = 11
                            elif percentage >= 70: grade_name = 'B+'; points = 10
                            elif percentage >= 65: grade_name = 'B'; points = 9
                            elif percentage >= 60: grade_name = 'B-'; points = 8
                            elif percentage >= 55: grade_name = 'C+'; points = 7
                            elif percentage >= 50: grade_name = 'C'; points = 6
                            elif percentage >= 45: grade_name = 'C-'; points = 5
                            elif percentage >= 40: grade_name = 'D+'; points = 4
                            elif percentage >= 35: grade_name = 'D'; points = 3
                            elif percentage >= 30: grade_name = 'D-'; points = 2
                            else: grade_name = 'E'; points = 1
                            
                            cursor.execute("""
                                SELECT id FROM digitallibrary_grade WHERE grade = %s LIMIT 1
                            """, [grade_name])
                            grade = cursor.fetchone()
                            grade_id = grade[0] if grade else None
                            
                            cursor.execute("""
                                INSERT INTO digitallibrary_studentresult 
                                (student_id, exam_id, subject_id, score, grade_id, points, entered_by_id, entered_at, updated_at)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                                ON CONFLICT (student_id, exam_id, subject_id) 
                                DO UPDATE SET 
                                    score = EXCLUDED.score,
                                    grade_id = EXCLUDED.grade_id,
                                    points = EXCLUDED.points,
                                    updated_at = NOW()
                            """, [student.id, exam.id, subject.id, score, grade_id, points, request.user.id])
                            results_processed += 1
                            
                    except Exception as e:
                        errors.append(f"Row {index + 2}: {str(e)}")
            
            if results_processed > 0:
                messages.success(request, f'✅ Successfully processed {results_processed} results for {exam.name} - {subject.name}')
            else:
                messages.warning(request, '⚠️ No valid results were found in the file.')
            
            if errors:
                for error in errors[:5]:
                    messages.warning(request, error)
                if len(errors) > 5:
                    messages.info(request, f'... and {len(errors) - 5} more errors')
                
            return redirect('digitallibrary:exam_list')
            
        except Exam.DoesNotExist:
            messages.error(request, 'Selected exam not found')
        except Subject.DoesNotExist:
            messages.error(request, 'Selected subject not found')
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            print(f"❌ Exception: {e}")
        
        return redirect('digitallibrary:bulk_enter_results')
    
    # GET request - show form
    context = {
        'exams': exams,
        'subjects': subjects,
        'title': 'Bulk Results Upload',
        'school': SchoolSetting.objects.first(),
    }
    
    print(f"📊 Context: exams={exams.count()}, subjects={subjects.count()}")
    return render(request, 'performance/bulk_excel_upload.html', context)
# ========== TEACHER DASHBOARD ==========
@login_required
@tenant_app_view
def teacher_dashboard(request):
    """Teacher dashboard showing class teacher and subject teacher responsibilities"""
    
    if request.user.profile.role != 'teacher':
        messages.error(request, "Access Denied. Only teachers can access this page.")
        return redirect('digitallibrary:home')
    
    # Check if teacher is a class teacher (homeroom)
    my_class = Class.objects.filter(class_teacher=request.user).first()
    
    # Get subjects they teach
    from .models import TeacherSubject, StudentResult, Exam, TeacherGradingPreference
    my_subjects = TeacherSubject.objects.filter(
        teacher=request.user
    ).select_related('subject', 'class_assigned')
    
    # Get teacher's grading preference
    teacher_preference = TeacherGradingPreference.objects.filter(
        teacher=request.user,
        is_global=True
    ).first()
    
    # If no preference set, create a default one
    if not teacher_preference:
        teacher_preference = TeacherGradingPreference.objects.create(
            teacher=request.user,
            grading_choice='traditional',
            is_global=True
        )
    
    # Get recent exams for subjects they teach
    subject_ids = my_subjects.values_list('subject_id', flat=True)
    recent_exams = Exam.objects.filter(
        subject__id__in=subject_ids
    ).order_by('-created_at')[:5] if subject_ids else []
    
    # Get recent results they entered
    recent_results = StudentResult.objects.filter(
        entered_by=request.user
    ).select_related('student', 'exam', 'subject')[:10]
    
    # Get student count in class (if class teacher)
    student_count = my_class.students.count() if my_class else 0
    
    # Get grading systems available for teacher
    from .models import GradingSystem, CBEGradingPathway
    available_grading_systems = GradingSystem.objects.filter(is_active=True)[:5]
    available_cbe_pathways = CBEGradingPathway.objects.filter(is_active=True)[:5]
    
    context = {
        'my_class': my_class,
        'my_subjects': my_subjects,
        'recent_exams': recent_exams,
        'recent_results': recent_results,
        'has_class': my_class is not None,
        'has_subjects': my_subjects.exists(),
        'has_results': recent_results.exists(),
        'student_count': student_count,
        'teacher_preference': teacher_preference,
        'available_grading_systems': available_grading_systems,
        'available_cbe_pathways': available_cbe_pathways,
        'title': 'Teacher Dashboard',
    }
    return render(request, 'digitallibrary/teacher_dashboard.html', context)


def class_teacher_dashboard(request):
    """Class teacher dashboard"""
    from .models import Exam, Result, Student
    
    exams = Exam.objects.all().order_by('-academic_year', '-created_at')
    total_students = Student.objects.filter(is_active=True).count()
    
    exams_data = []
    for exam in exams:
        results_count_total = Result.objects.filter(exam=exam).values('student').distinct().count()
        
        exams_data.append({
            'id': exam.id,
            'name': exam.name,
            'academic_year': exam.academic_year,
            'term': exam.term,
            'results_count': results_count_total,
            'total_students': total_students,
            'subject_progress': [],
        })
    
    context = {
        'assigned_class': None,
        'total_exams': exams.count(),
        'completed_exams': 0,
        'in_progress_exams': 0,
        'total_students': total_students,
        'exams': exams_data,
        'top_students': [],
    }
    
    return render(request, 'performance/class_teacher_dashboard.html', context)


def compile_results_overview(request):
    """Overview of all exams ready for compilation"""
    from .models import Exam
    
    exams = Exam.objects.all().order_by('-academic_year', '-created_at')
    
    context = {
        'exams': exams,
    }
    
    return render(request, 'performance/compile_results_overview.html', context)


def exam_compilation(request, exam_id):
    """Compile results for a specific exam"""
    from .models import Exam, Result, Student
    
    exam = Exam.objects.get(id=exam_id)
    total_students = Student.objects.filter(is_active=True).count()
    
    # Get subjects for this exam
    from .models import Subject
    subjects = Subject.objects.all()
    
    subjects_data = []
    subjects_completed = 0
    
    for subject in subjects:
        results_count = Result.objects.filter(exam=exam, subject=subject).count()
        completion_rate = (results_count / total_students * 100) if total_students > 0 else 0
        
        if completion_rate == 100:
            subjects_completed += 1
        
        subjects_data.append({
            'subject': subject,
            'results_entered': results_count,
            'completion_rate': completion_rate,
            'teacher': 'Not assigned',
        })
    
    completion_percentage = (subjects_completed / len(subjects) * 100) if subjects else 0
    all_subjects_complete = subjects_completed == len(subjects) if subjects else False
    
    if request.method == 'POST' and all_subjects_complete:
        # Compile results - calculate total scores and ranks
        from .models import ExamResultSummary
        
        students = Student.objects.filter(is_active=True)
        
        for student in students:
            # Get all results for this student in this exam
            results = Result.objects.filter(exam=exam, student=student)
            total_score = sum(r.score for r in results)
            avg_score = total_score / results.count() if results.count() > 0 else 0
            
            # Calculate grade
            if avg_score >= 80:
                grade = 'A'
            elif avg_score >= 70:
                grade = 'B'
            elif avg_score >= 60:
                grade = 'C'
            elif avg_score >= 50:
                grade = 'D'
            else:
                grade = 'E'
            
            summary, created = ExamResultSummary.objects.update_or_create(
                exam=exam,
                student=student,
                defaults={
                    'total_score': total_score,
                    'average_score': avg_score,
                    'overall_grade': grade,
                }
            )
        
        # Calculate ranks
        summaries = ExamResultSummary.objects.filter(exam=exam).order_by('-average_score')
        for idx, summary in enumerate(summaries, 1):
            summary.rank = idx
            summary.save()
        
        messages.success(request, f'Results compiled successfully for {exam.name}!')
        return redirect('digitallibrary:exam_ranking', exam_id=exam.id)
    
    context = {
        'exam': exam,
        'subjects_data': subjects_data,
        'total_subjects': len(subjects),
        'subjects_completed': subjects_completed,
        'completion_percentage': completion_percentage,
        'all_subjects_complete': all_subjects_complete,
        'total_students': total_students,
        'subjects_incomplete': len(subjects) - subjects_completed,
    }
    
    return render(request, 'performance/exam_compilation.html', context)


def exam_ranking(request, exam_id):
    """View rankings for a compiled exam"""
    from .models import Exam, ExamResultSummary, Student
    
    exam = Exam.objects.get(id=exam_id)
    rankings = ExamResultSummary.objects.filter(exam=exam).select_related('student').order_by('rank')
    
    # Calculate class average and pass rate
    class_average = rankings.aggregate(avg=models.Avg('average_score'))['avg'] or 0
    pass_count = rankings.filter(average_score__gte=50).count()
    pass_rate = (pass_count / rankings.count() * 100) if rankings.count() > 0 else 0
    
    top_student = rankings.first()
    
    context = {
        'exam': exam,
        'rankings': rankings,
        'class_average': class_average,
        'pass_rate': pass_rate,
        'top_student': top_student.student if top_student else None,
    }
    
    return render(request, 'performance/exam_ranking.html', context)


def class_ranking(request, class_id):
    """View rankings for a class across all exams"""
    from .models import Class, Student, ExamResultSummary
    
    class_obj = Class.objects.get(id=class_id)
    students = Student.objects.filter(current_class=class_obj, is_active=True)
    
    student_summaries = []
    for student in students:
        summaries = ExamResultSummary.objects.filter(student=student)
        if summaries.exists():
            avg_overall = summaries.aggregate(avg=models.Avg('average_score'))['avg'] or 0
            student_summaries.append({
                'student': student,
                'average_score': avg_overall,
                'exams_taken': summaries.count(),
            })
    
    student_summaries.sort(key=lambda x: x['average_score'], reverse=True)
    
    context = {
        'class_obj': class_obj,
        'rankings': student_summaries,
    }
    
    return render(request, 'performance/class_ranking.html', context)


def export_ranking_csv(request, exam_id):
    """Export exam rankings to CSV"""
    import csv
    from django.http import HttpResponse
    from .models import Exam, ExamResultSummary
    
    exam = Exam.objects.get(id=exam_id)
    rankings = ExamResultSummary.objects.filter(exam=exam).select_related('student').order_by('rank')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{exam.name}_rankings.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Rank', 'Admission Number', 'Student Name', 'Total Score', 'Average Score', 'Grade'])
    
    for ranking in rankings:
        writer.writerow([
            ranking.rank,
            ranking.student.admission_number,
            f"{ranking.student.first_name} {ranking.student.last_name}",
            ranking.total_score,
            ranking.average_score,
            ranking.overall_grade,
        ])
    
    return response


def subject_exam_performance(request, subject_id, exam_id):
    """View performance for a specific subject in an exam"""
    from .models import Subject, Exam, Result, Student
    
    subject = Subject.objects.get(id=subject_id)
    exam = Exam.objects.get(id=exam_id)
    
    results = Result.objects.filter(exam=exam, subject=subject).select_related('student')
    
    total_students = Student.objects.filter(is_active=True).count()
    avg_score = results.aggregate(avg=models.Avg('score'))['avg'] or 0
    top_score = results.aggregate(max=models.Max('score'))['max'] or 0
    lowest_score = results.aggregate(min=models.Min('score'))['min'] or 0
    
    context = {
        'subject': subject,
        'exam': exam,
        'results': results,
        'total_students': total_students,
        'avg_score': avg_score,
        'top_score': top_score,
        'lowest_score': lowest_score,
    }
    
    return render(request, 'performance/subject_exam_performance.html', context)


def view_subject_results(request, exam_id, subject_id):
    """View all results for a subject in an exam (for class teacher)"""
    from .models import Exam, Subject, Result
    
    exam = Exam.objects.get(id=exam_id)
    subject = Subject.objects.get(id=subject_id)
    results = Result.objects.filter(exam=exam, subject=subject).select_related('student').order_by('-score')
    
    context = {
        'exam': exam,
        'subject': subject,
        'results': results,
    }
    
    return render(request, 'performance/view_subject_results.html', context)


def student_performance_tracking(request, student_id):
    """Track student performance across different exams"""
    from .models import Student, Result, Exam
    
    student = Student.objects.get(id=student_id)
    
    # Get all results grouped by subject and exam
    results = Result.objects.filter(student=student).select_related('exam', 'subject').order_by('-exam__academic_year', '-exam__created_at')
    
    # Group by subject
    subjects_data = {}
    for result in results:
        subject_name = result.subject.name
        if subject_name not in subjects_data:
            subjects_data[subject_name] = []
        
        subjects_data[subject_name].append({
            'exam_name': result.exam.name,
            'exam_term': result.exam.term,
            'exam_year': result.exam.academic_year,
            'score': result.score,
            'grade': result.grade if hasattr(result, 'grade') else 'C',
        })
    
    # Get overall averages per subject
    subject_averages = {}
    for subject_name, scores in subjects_data.items():
        avg = sum(s['score'] for s in scores) / len(scores) if scores else 0
        subject_averages[subject_name] = avg
    
    context = {
        'student': student,
        'subjects_data': subjects_data,
        'subject_averages': subject_averages,
        'total_exams': results.values('exam').distinct().count(),
    }
    
    return render(request, 'performance/student_performance_tracking.html', context)

# ========== SUBJECT PERFORMANCE VIEW ==========

@staff_member_required
def subject_performance(request, subject_id):
    """View performance for a specific subject"""
    subject = get_object_or_404(Subject, pk=subject_id)
    academic_year = request.GET.get('year', str(timezone.now().year))
    term = request.GET.get('term', '1')
    
    exams = Exam.objects.filter(academic_year=academic_year, term=term)
    results = StudentResult.objects.filter(
        subject=subject,
        exam__in=exams
    ).select_related('student', 'exam')
    
    avg_score = results.aggregate(Avg('score'))['score__avg'] or 0
    top_score = results.aggregate(Max('score'))['score__max'] or 0
    lowest_score = results.aggregate(Min('score'))['score__min'] or 0
    
    context = {
        'subject': subject,
        'exams': exams,
        'results': results,
        'avg_score': avg_score,
        'top_score': top_score,
        'lowest_score': lowest_score,
        'academic_year': academic_year,
        'term': term,
        'school': SchoolSetting.objects.first(),
    }
    return render(request, 'performance/subject_performance.html', context)


# ========== CLASS PERFORMANCE VIEW ==========

@staff_member_required
def class_performance(request, class_id):
    """View performance for a specific class"""
    student_class = get_object_or_404(Class, pk=class_id)
    academic_year = request.GET.get('year', str(timezone.now().year))
    term = request.GET.get('term', '1')
    
    students = Student.objects.filter(current_class=student_class, is_active=True)
    summaries = PerformanceSummary.objects.filter(
        student__in=students,
        academic_year=academic_year,
        term=term
    ).select_related('student').order_by('rank_in_class')
    
    avg_class_score = summaries.aggregate(Avg('average_score'))['average_score__avg'] or 0
    total_passed = summaries.filter(average_score__gte=50).count()
    
    context = {
        'class': student_class,
        'summaries': summaries,
        'academic_year': academic_year,
        'term': term,
        'avg_class_score': avg_class_score,
        'total_students': students.count(),
        'total_passed': total_passed,
        'school': SchoolSetting.objects.first(),
    }
    return render(request, 'performance/class_performance.html', context)


# ========== PERFORMANCE REPORTS VIEW ==========

@staff_member_required
@tenant_app_view
def performance_reports(request):
    """Comprehensive performance reports with grade distribution and summaries"""
    from .models import Exam, Class, Subject, Student, StudentResult
    from django.db.models import Avg, Count, Q, Sum
    from collections import defaultdict
    
    # Get filter parameters
    academic_year = request.GET.get('year', '')
    term = request.GET.get('term', '')
    selected_class = request.GET.get('class', '')
    selected_exam = request.GET.get('exam', '')
    
    # Base queryset for results
    results_qs = StudentResult.objects.all()
    
    # Apply filters
    if academic_year:
        results_qs = results_qs.filter(exam__academic_year=academic_year)
    if term:
        results_qs = results_qs.filter(exam__term=term)
    if selected_class:
        results_qs = results_qs.filter(student__current_class_id=selected_class)
    if selected_exam:
        results_qs = results_qs.filter(exam_id=selected_exam)
    
    # Get total students
    students_qs = Student.objects.filter(is_active=True)
    if selected_class:
        students_qs = students_qs.filter(current_class_id=selected_class)
    total_students = students_qs.count()
    
    # Calculate overall metrics
    avg_data = results_qs.aggregate(avg=Avg('score'))
    avg_score = avg_data['avg'] or 0
    
    total_results = results_qs.count()
    passed_results = results_qs.filter(score__gte=50).count()
    pass_rate = (passed_results / total_results * 100) if total_results > 0 else 0
    
    # Calculate passed students count
    passed_students = results_qs.filter(score__gte=50).values('student').distinct().count()
    
    # Overall Grade Distribution
    grade_ranges = {
        'A': (80, 100),
        'A-': (75, 79),
        'B+': (70, 74),
        'B': (65, 69),
        'B-': (60, 64),
        'C+': (55, 59),
        'C': (50, 54),
        'C-': (45, 49),
        'D+': (40, 44),
        'D': (35, 39),
        'E': (0, 34),
    }
    
    overall_grade_distribution = {}
    total_grade_count = 0
    
    for grade, (min_score, max_score) in grade_ranges.items():
        count = results_qs.filter(score__gte=min_score, score__lte=max_score).count()
        overall_grade_distribution[grade] = {
            'count': count,
            'percentage': (count / total_results * 100) if total_results > 0 else 0
        }
        total_grade_count += count
    
    # Exams Summary
    exams_summary = []
    exams = Exam.objects.all()
    if academic_year:
        exams = exams.filter(academic_year=academic_year)
    if term:
        exams = exams.filter(term=term)
    if selected_class:
        exams = exams.filter(student_class_id=selected_class)
    if selected_exam:
        exams = exams.filter(id=selected_exam)
    
    for exam in exams:
        exam_results = results_qs.filter(exam=exam)
        if exam_results.exists():
            avg = exam_results.aggregate(avg=Avg('score'))['avg'] or 0
            passed = exam_results.filter(score__gte=50).count()
            pass_rate_exam = (passed / exam_results.count() * 100) if exam_results.count() > 0 else 0
            
            # Get top grade in this exam
            top_score = exam_results.aggregate(max=Avg('score'))['max'] or 0
            if top_score >= 80:
                top_grade = 'A'
            elif top_score >= 75:
                top_grade = 'A-'
            elif top_score >= 70:
                top_grade = 'B+'
            elif top_score >= 65:
                top_grade = 'B'
            elif top_score >= 60:
                top_grade = 'B-'
            else:
                top_grade = 'C'
            
            exams_summary.append({
                'id': exam.id,
                'name': exam.name,
                'academic_year': exam.academic_year,
                'term': exam.term,
                'students': exam_results.values('student').distinct().count(),
                'avg_score': avg,
                'pass_rate': pass_rate_exam,
                'top_grade': top_grade,
            })
    
    # Exam Grade Distribution
    exam_grade_distribution = []
    for exam in exams[:10]:  # Limit to 10 exams for performance
        exam_results = results_qs.filter(exam=exam)
        if exam_results.exists():
            distribution = {}
            for grade, (min_score, max_score) in grade_ranges.items():
                count = exam_results.filter(score__gte=min_score, score__lte=max_score).count()
                if count > 0:
                    distribution[grade] = count
            
            exam_grade_distribution.append({
                'id': exam.id,
                'name': exam.name,
                'academic_year': exam.academic_year,
                'term': exam.term,
                'distribution': distribution,
                'total_results': exam_results.count(),
            })
    
    # Top Students
    top_students_data = results_qs.values('student').annotate(
        avg=Avg('score'),
        exams_taken=Count('exam', distinct=True)
    ).order_by('-avg')[:20]
    
    top_students = []
    for ts in top_students_data:
        student = Student.objects.filter(id=ts['student']).first()
        if student:
            avg = ts['avg']
            if avg >= 80:
                grade = 'A'
            elif avg >= 75:
                grade = 'A-'
            elif avg >= 70:
                grade = 'B+'
            elif avg >= 65:
                grade = 'B'
            elif avg >= 60:
                grade = 'B-'
            elif avg >= 55:
                grade = 'C+'
            elif avg >= 50:
                grade = 'C'
            else:
                grade = 'D'
            
            top_students.append({
                'student': student,
                'average': avg,
                'grade': grade,
                'exams_taken': ts['exams_taken'],
            })
    
    # Subject Performance
    subject_performance = []
    subjects = Subject.objects.all()
    for subject in subjects:
        subject_results = results_qs.filter(subject=subject)
        if subject_results.exists():
            avg = subject_results.aggregate(avg=Avg('score'))['avg'] or 0
            passed = subject_results.filter(score__gte=50).count()
            pass_rate_subj = (passed / subject_results.count() * 100) if subject_results.count() > 0 else 0
            
            if avg >= 80:
                grade = 'A'
            elif avg >= 70:
                grade = 'B'
            elif avg >= 60:
                grade = 'C'
            elif avg >= 50:
                grade = 'D'
            else:
                grade = 'E'
            
            subject_performance.append({
                'name': subject.name,
                'students': subject_results.values('student').distinct().count(),
                'average': avg,
                'pass_rate': pass_rate_subj,
                'grade': grade,
            })
    
    # Class Performance
    class_performance = []
    classes = Class.objects.all()
    for class_obj in classes:
        class_results = results_qs.filter(student__current_class=class_obj)
        if class_results.exists():
            avg = class_results.aggregate(avg=Avg('score'))['avg'] or 0
            passed = class_results.filter(score__gte=50).count()
            pass_rate_class = (passed / class_results.count() * 100) if class_results.count() > 0 else 0
            
            if avg >= 80:
                grade = 'A'
            elif avg >= 70:
                grade = 'B'
            elif avg >= 60:
                grade = 'C'
            elif avg >= 50:
                grade = 'D'
            else:
                grade = 'E'
            
            class_performance.append({
                'name': class_obj.name,
                'students': class_results.values('student').distinct().count(),
                'average': avg,
                'pass_rate': pass_rate_class,
                'grade': grade,
            })
    
    # Year choices for filter
    year_choices = Exam.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
    
    # Classes for filter
    classes = Class.objects.all().order_by('name')
    
    # Exams for filter
    exam_choices = Exam.objects.all().order_by('-academic_year', '-created_at')
    
    context = {
        # Existing context
        'total_students': total_students,
        'avg_score': avg_score,
        'pass_rate': pass_rate,
        'pass_count': passed_students,
        'top_students': top_students,
        'subject_performance': subject_performance,
        'class_performance': class_performance,
        'year_choices': year_choices,
        'academic_year': academic_year,
        'term': term,
        'selected_class': selected_class,
        'classes': classes,
        
        # NEW context variables
        'total_exams': exams.count(),
        'total_results_count': total_results,
        'overall_grade_distribution': overall_grade_distribution,
        'exams_summary': exams_summary,
        'exam_grade_distribution': exam_grade_distribution,
        'exam_choices': exam_choices,
        'selected_exam': selected_exam,
    }
    
    return render(request, 'performance/performance_reports.html', context)
def export_performance_report(request):
    """Export performance report to CSV"""
    import csv
    from django.http import HttpResponse
    from .models import StudentResult, Student, Exam, Subject
    from django.db.models import Avg
    
    # Get filter parameters
    academic_year = request.GET.get('year', '')
    term = request.GET.get('term', '')
    selected_class = request.GET.get('class', '')
    selected_exam = request.GET.get('exam', '')
    
    # Base queryset
    results_qs = StudentResult.objects.all()
    
    if academic_year:
        results_qs = results_qs.filter(exam__academic_year=academic_year)
    if term:
        results_qs = results_qs.filter(exam__term=term)
    if selected_class:
        results_qs = results_qs.filter(student__current_class_id=selected_class)
    if selected_exam:
        results_qs = results_qs.filter(exam_id=selected_exam)
    
    # Get top students
    top_students = results_qs.values('student').annotate(
        avg=Avg('score')
    ).order_by('-avg')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="performance_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Rank', 'Admission Number', 'Student Name', 'Class', 'Average Score', 'Grade'])
    
    for idx, ts in enumerate(top_students, 1):
        student = Student.objects.filter(id=ts['student']).first()
        if student:
            avg = ts['avg']
            if avg >= 80:
                grade = 'A'
            elif avg >= 70:
                grade = 'B'
            elif avg >= 60:
                grade = 'C'
            elif avg >= 50:
                grade = 'D'
            else:
                grade = 'E'
            
            writer.writerow([
                idx,
                student.admission_number,
                f"{student.first_name} {student.last_name}",
                student.current_class.name if student.current_class else 'N/A',
                f"{avg:.1f}",
                grade
            ])
    
    return response
# ========== API SUBMIT RESULTS ==========

@csrf_exempt
@login_required
def submit_results_api(request):
    """API endpoint to submit exam results"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST method required'}, status=405)
    
    try:
        data = json.loads(request.body)
        results = data.get('results', [])
        
        if not results:
            return JsonResponse({'success': False, 'error': 'No results provided'}, status=400)
        
        saved_count = 0
        errors = []
        
        for result_data in results:
            try:
                student_id = result_data.get('student_id')
                exam_id = result_data.get('exam_id')
                score = result_data.get('score')
                
                if not all([student_id, exam_id, score is not None]):
                    errors.append(f"Missing data for student {student_id}")
                    continue
                
                exam = Exam.objects.get(id=exam_id)
                student = Student.objects.get(id=student_id)
                
                StudentResult.objects.update_or_create(
                    student=student,
                    exam=exam,
                    defaults={'score': float(score), 'entered_by': request.user}
                )
                saved_count += 1
                
            except Exam.DoesNotExist:
                errors.append(f"Exam {exam_id} not found")
            except Student.DoesNotExist:
                errors.append(f"Student {student_id} not found")
            except Exception as e:
                errors.append(str(e))
        
        return JsonResponse({
            'success': True,
            'saved': saved_count,
            'errors': errors if errors else None
        })
        
    except json.JSONDecodeError as e:
        return JsonResponse({'success': False, 'error': f'Invalid JSON: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
# ========== SMS DASHBOARD VIEWS ==========



@sms_access  # Allows admin, principal, and bursar
def sms_dashboard(request):
    """SMS management dashboard"""
    from django.conf import settings
    from django.db import connection
    from django.core.paginator import Paginator
    from .models import Student, Class, SchoolSetting
    
    # Removed manual role check since decorator handles it
    
    # Get teacher count using SQL (no UserProfile)
    total_teachers = 0
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM auth_user WHERE is_staff = true AND is_superuser = false AND is_active = true")
            result = cursor.fetchone()
            total_teachers = result[0] if result else 0
    except Exception as e:
        print(f"Error counting teachers: {e}")
    
    # Get student counts
    total_students = Student.objects.filter(is_active=True).count()
    students_with_phone = Student.objects.filter(is_active=True, parent_phone__isnull=False).exclude(parent_phone='').count()
    
    # Get paginated students
    all_students = Student.objects.filter(is_active=True).select_related('current_class').order_by('first_name', 'last_name')
    paginator = Paginator(all_students, 20)
    page_number = request.GET.get('page', 1)
    students = paginator.get_page(page_number)
    
    # Get classes
    classes = Class.objects.all().order_by('name')
    
    # Get SMS mode status
    mock_sms_mode = getattr(settings, 'MOCK_SMS_MODE', True)
    africastalking_username = getattr(settings, 'AFRICASTALKING_USERNAME', 'sandbox')
    
    # Determine SMS mode
    if not mock_sms_mode and africastalking_username != 'sandbox':
        sms_mode = 'LIVE'
        sms_mode_color = 'green'
        sms_mode_text = 'LIVE MODE'
        sms_mode_subtext = 'Real SMS - Charges apply'
        sms_status_icon = '✅'
    else:
        sms_mode = 'TEST'
        sms_mode_color = 'yellow'
        sms_mode_text = 'TEST MODE'
        sms_mode_subtext = 'Mock SMS - No charges'
        sms_status_icon = '🔧'
    
    context = {
        'title': 'SMS Dashboard',
        'total_teachers': total_teachers,
        'total_students': total_students,
        'students_with_phone': students_with_phone,
        'students': students,
        'classes': classes,
        'messages_sent': 0,
        'school': SchoolSetting.objects.first(),
        # SMS Mode variables
        'sms_mode': sms_mode,
        'sms_mode_color': sms_mode_color,
        'sms_mode_text': sms_mode_text,
        'sms_mode_subtext': sms_mode_subtext,
        'sms_status_icon': sms_status_icon,
        'mock_sms_mode': mock_sms_mode,
    }
    return render(request, "digitallibrary/sms/dashboard.html", context)


@sms_access  # Allows admin, principal, and bursar
@require_http_methods(["POST"])
def send_bulk_sms_view(request):
    """Send bulk SMS to selected recipients"""
    # Removed manual role check since decorator handles it
    
    recipient_type = request.POST.get('recipient_type')
    message = request.POST.get('message', '').strip()
    student_ids = request.POST.get('student_ids', '')
    class_id = request.POST.get('class_id')
    
    if not message:
        return JsonResponse({'error': 'Message cannot be empty'}, status=400)
    
    if len(message) > 160:
        return JsonResponse({'error': 'Message exceeds 160 characters'}, status=400)
    
    phone_numbers = []
    recipients_info = []
    
    if recipient_type == 'selected' and student_ids:
        student_id_list = [int(id) for id in student_ids.split(',') if id]
        students = Student.objects.filter(id__in=student_id_list, is_active=True)
        for student in students:
            if student.parent_phone:
                formatted = format_phone_number(student.parent_phone)
                if formatted:
                    phone_numbers.append(formatted)
                    recipients_info.append({
                        'name': student.get_full_name(),
                        'phone': formatted,
                        'admission': student.admission_number
                    })
    
    elif recipient_type == 'all':
        students = Student.objects.filter(is_active=True, parent_phone__isnull=False).exclude(parent_phone='')
        for student in students:
            formatted = format_phone_number(student.parent_phone)
            if formatted:
                phone_numbers.append(formatted)
                recipients_info.append({
                    'name': student.get_full_name(),
                    'phone': formatted,
                    'admission': student.admission_number
                })
    
    elif recipient_type == 'class' and class_id:
        students = Student.objects.filter(current_class_id=class_id, is_active=True, parent_phone__isnull=False).exclude(parent_phone='')
        for student in students:
            formatted = format_phone_number(student.parent_phone)
            if formatted:
                phone_numbers.append(formatted)
                recipients_info.append({
                    'name': student.get_full_name(),
                    'phone': formatted,
                    'admission': student.admission_number
                })
    
    if not phone_numbers:
        return JsonResponse({'success': False, 'error': 'No valid phone numbers found.'}, status=400)
    
    if MOCK_SMS_MODE:
        for info in recipients_info[:5]:
            print(f"[TEST MODE] Would send SMS to {info['name']} ({info['phone']})")
        
        return JsonResponse({
            'success': True,
            'successful': len(phone_numbers),
            'failed': 0,
            'total': len(phone_numbers),
            'message': f"✓ Test Mode: {len(phone_numbers)} SMS message(s) would be sent to {len(recipients_info)} recipient(s)."
        })
    else:
        try:
            result = send_bulk_sms(phone_numbers, message)
            return JsonResponse({
                'success': True,
                'successful': result.get('successful', 0),
                'failed': result.get('failed', 0),
                'total': result.get('total', 0),
                'message': f"✓ Successfully sent to {result.get('successful', 0)} out of {result.get('total', 0)} recipients"
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


@sms_access  # Allows admin, principal, and bursar
@require_http_methods(["POST"])
def send_test_sms(request):
    """Send a test SMS to a single number"""
    # Removed manual role check since decorator handles it
    
    phone_number = request.POST.get('phone_number', '').strip()
    message = request.POST.get('message', '').strip()
    
    if not phone_number or not message:
        return JsonResponse({'error': 'Phone number and message are required'}, status=400)
    
    if not phone_number.startswith('+'):
        if phone_number.startswith('0'):
            phone_number = '+254' + phone_number[1:]
        elif phone_number.startswith('254'):
            phone_number = '+' + phone_number
    
    result = send_sms(phone_number, message)
    
    if result['success']:
        ActivityLog.objects.create(
            user=request.user,
            action="sms_test",
            description=f"Test SMS sent to {phone_number}"
        )
        return JsonResponse({'success': True, 'response': result['response']})
    else:
        return JsonResponse({'success': False, 'error': result['error']}, status=500)

# ========== API ENDPOINTS FOR STUDENTS ==========

@login_required
def get_students_by_class_api(request, class_id):
    """API to get students by class"""
    if request.user.profile.role not in ['admin', 'principal', 'teacher']:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    students = Student.objects.filter(current_class_id=class_id, is_active=True)
    
    data = {
        'students': [
            {
                'id': s.id,
                'admission_number': s.admission_number,
                'first_name': s.first_name,
                'last_name': s.last_name,
                'class_name': s.current_class.name if s.current_class else None
            }
            for s in students
        ]
    }
    return JsonResponse(data)


@login_required
def get_all_students_api(request):
    """API to get all active students"""
    if request.user.profile.role not in ['admin', 'principal', 'teacher']:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    students = Student.objects.filter(is_active=True).select_related('current_class')
    
    data = {
        'students': [
            {
                'id': s.id,
                'admission_number': s.admission_number,
                'first_name': s.first_name,
                'last_name': s.last_name,
                'class_name': s.current_class.name if s.current_class else None
            }
            for s in students
        ]
    }
    return JsonResponse(data)


# ========== FEEDBACK SYSTEM VIEWS ==========

@login_required
def share_feedback(request):
    """Submit user feedback with beautiful email template"""
    if request.method == 'POST':
        form = FeedbackForm(request.POST, request.FILES)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.user = request.user
            feedback.page_url = request.META.get('HTTP_REFERER', '')
            
            # Get school info from tenant if available
            if hasattr(request, 'tenant') and request.tenant:
                feedback.school_id = request.tenant.schema_name
                feedback.school_name = request.tenant.name
                feedback.school_location = getattr(request.tenant, 'location', 'Kenya')
            else:
                feedback.school_id = getattr(settings, 'SCHOOL_ID', 'unknown')
                feedback.school_name = getattr(settings, 'SCHOOL_NAME', 'ShuleHub')
                feedback.school_location = getattr(settings, 'SCHOOL_LOCATION', 'Kenya')
            
            feedback.save()
            
            # Handle screenshot if uploaded
            if 'screenshot' in request.FILES:
                screenshot = request.FILES['screenshot']
                # You can save screenshot logic here
                pass
            
            # Prepare email context for beautiful template
            from django.template.loader import render_to_string
            from django.core.mail import send_mail
            from django.utils.html import strip_tags
            
            # Get user info
            user_name = feedback.user.get_full_name() or feedback.user.username
            user_email = feedback.user.email
            user_role = feedback.user.profile.role if hasattr(feedback.user, 'profile') else 'User'
            
            # Get admin URL
            admin_url = request.build_absolute_uri('/admin/digitallibrary/feedback/')
            
            # Email context
            context = {
                'school_name': feedback.school_name or 'ShuleHub',
                'school_location': feedback.school_location or 'Kenya',
                'school_id': feedback.school_id or 'N/A',
                'user_name': user_name,
                'user_role': user_role.upper() if user_role else 'USER',
                'user_email': user_email or 'Not provided',
                'feedback_type': feedback.feedback_type,
                'priority': feedback.priority,
                'rating': feedback.rating if feedback.rating else None,
                'subject': feedback.subject,
                'message': feedback.message,
                'admin_url': admin_url,
                'has_rating': feedback.rating is not None,
            }
            
            # Render HTML email
            html_message = render_to_string('emails/feedback_notification.html', context)
            plain_message = strip_tags(html_message)
            
            subject = f"[Feedback] {feedback.school_name} - {feedback.subject}"
            
            try:
                send_mail(
                    subject=subject,
                    message=plain_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=settings.ADMIN_EMAILS,
                    html_message=html_message,
                    fail_silently=False,
                )
                print(f"✅ Feedback email sent to {settings.ADMIN_EMAILS}")
            except Exception as e:
                print(f"❌ Email error: {e}")
                # Fallback to simple email if HTML fails
                try:
                    simple_message = f"""
                    School: {feedback.school_name}
                    From: {user_name}
                    Email: {user_email}
                    Role: {user_role}
                    Type: {feedback.get_feedback_type_display()}
                    Rating: {feedback.rating or 'No rating'}/5
                    Priority: {feedback.priority}
                    Subject: {feedback.subject}
                    Message: {feedback.message}
                    """
                    send_mail(
                        subject=subject,
                        message=simple_message,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=settings.ADMIN_EMAILS,
                        fail_silently=False,
                    )
                except Exception as e2:
                    print(f"❌ Fallback email also failed: {e2}")
            
            messages.success(request, 'Thank you for your feedback! Our team has been notified.')
            return redirect('/app/feedback/success/')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = FeedbackForm()
    
    # Add try/except for SchoolSetting to prevent errors
    from .models import SchoolSetting
    try:
        school = SchoolSetting.objects.first()
    except:
        school = None
    
    context = {
        'form': form,
        'title': 'Share Feedback',
        'school': school,
    }
    return render(request, 'digitallibrary/feedback.html', context)
def feedback_success(request):
    """Feedback submission success page"""
    from .models import SchoolSetting
    try:
        school = SchoolSetting.objects.first()
    except:
        school = None
    
    context = {
        'school': school,
        'title': 'Thank You for Your Feedback',
    }
    return render(request, 'digitallibrary/feedback_success.html', context)
@login_required
def feedback_list(request):
    """List all feedback (admin only)"""
    if request.user.profile.role not in ['admin', 'principal']:
        messages.error(request, "Access Denied.")
        return redirect('digitallibrary:home')
    
    feedbacks = Feedback.objects.all().order_by('-created_at')
    
    status = request.GET.get('status')
    if status == 'resolved':
        feedbacks = feedbacks.filter(is_resolved=True)
    elif status == 'pending':
        feedbacks = feedbacks.filter(is_resolved=False)
    
    ftype = request.GET.get('type')
    if ftype:
        feedbacks = feedbacks.filter(feedback_type=ftype)
    
    paginator = Paginator(feedbacks, 20)
    page = request.GET.get('page', 1)
    feedbacks = paginator.get_page(page)
    
    context = {
        'feedbacks': feedbacks,
        'school': SchoolSetting.objects.first(),
    }
    return render(request, 'digitallibrary/feedback_list.html', context)
def update_performance_summary_for_exam(exam):
    """Update performance summaries for all students in an exam"""
    from django.db.models import Avg, Sum
    
    # Get all results for this exam
    results = StudentResult.objects.filter(exam=exam).select_related('student', 'subject')
    
    # Group by student
    student_scores = {}
    for result in results:
        student_id = result.student.id
        if student_id not in student_scores:
            student_scores[student_id] = {
                'student': result.student,
                'total_score': 0,
                'subject_count': 0,
                'scores': []
            }
        student_scores[student_id]['total_score'] += float(result.score)
        student_scores[student_id]['subject_count'] += 1
        student_scores[student_id]['scores'].append(float(result.score))
    
    # Update or create PerformanceSummary for each student
    for student_id, data in student_scores.items():
        if data['subject_count'] > 0:
            average_score = data['total_score'] / data['subject_count']
            
            # Determine grade
            if average_score >= 80:
                overall_grade = 'A'
            elif average_score >= 75:
                overall_grade = 'A-'
            elif average_score >= 70:
                overall_grade = 'B+'
            elif average_score >= 65:
                overall_grade = 'B'
            elif average_score >= 60:
                overall_grade = 'B-'
            elif average_score >= 55:
                overall_grade = 'C+'
            elif average_score >= 50:
                overall_grade = 'C'
            elif average_score >= 45:
                overall_grade = 'C-'
            elif average_score >= 40:
                overall_grade = 'D+'
            elif average_score >= 35:
                overall_grade = 'D'
            else:
                overall_grade = 'E'
            
            # Count passed subjects (score >= 50)
            subjects_passed = sum(1 for s in data['scores'] if s >= 50)
            subjects_failed = data['subject_count'] - subjects_passed
            
            # Calculate rank in class
            student_class = data['student'].current_class
            if student_class:
                class_summaries = PerformanceSummary.objects.filter(
                    student__current_class=student_class,
                    academic_year=exam.academic_year,
                    term=exam.term
                ).order_by('-average_score')
                
                rank = 1
                for i, s in enumerate(class_summaries, 1):
                    if s.student_id == student_id:
                        rank = i
                        break
            else:
                rank = 0
            
            PerformanceSummary.objects.update_or_create(
                student=data['student'],
                academic_year=exam.academic_year,
                term=exam.term,
                defaults={
                    'total_score': data['total_score'],
                    'average_score': average_score,
                    'overall_grade': overall_grade,
                    'rank_in_class': rank,
                    'subjects_passed': subjects_passed,
                    'subjects_failed': subjects_failed,
                }
            )


@staff_member_required
def exam_results_entry(request, exam_id):
    """Enter results for an exam - by subject, filtered by registered student subjects"""
    exam = get_object_or_404(Exam, pk=exam_id)

    # Show all active subjects for selection
    subjects = Subject.objects.filter(is_active=True).order_by('name')

    selected_subject_id = request.GET.get('subject')
    selected_subject = None
    students = []
    existing_results = {}

    if selected_subject_id:
        try:
            selected_subject = Subject.objects.get(pk=selected_subject_id, is_active=True)

            # Get students for this exam
            students_qs = exam.get_students_for_exam()

            # IMPORTANT:
            # Only show students registered for the selected subject
            students = students_qs.filter(
                subjects=selected_subject,
                is_active=True
            ).distinct().order_by('admission_number')

            existing_results_qs = StudentResult.objects.filter(
                exam=exam,
                subject=selected_subject,
                student__in=students
            ).select_related('student')

            existing_results = {
                result.student_id: result
                for result in existing_results_qs
            }

        except Subject.DoesNotExist:
            messages.error(request, "Selected subject does not exist.")

    if request.method == 'POST':
        subject_id = request.POST.get('subject_id')

        if subject_id:
            selected_subject = get_object_or_404(Subject, pk=subject_id, is_active=True)

            students = exam.get_students_for_exam().filter(
                subjects=selected_subject,
                is_active=True
            ).distinct()

            saved_count = 0

            for student in students:
                score_key = f'score_{student.id}'
                if score_key in request.POST:
                    score = request.POST.get(score_key)

                    if score and score.strip():
                        try:
                            score_value = float(score)

                            if 0 <= score_value <= float(exam.max_score):
                                StudentResult.objects.update_or_create(
                                    student=student,
                                    exam=exam,
                                    subject=selected_subject,
                                    defaults={
                                        'score': score_value,
                                        'entered_by': request.user
                                    }
                                )
                                saved_count += 1

                        except ValueError:
                            pass

            if saved_count > 0:
                update_performance_summary_for_exam(exam)
                messages.success(
                    request,
                    f'Results for {exam.name} - {selected_subject.name} saved successfully! '
                    f'{saved_count} records updated.'
                )
            else:
                messages.warning(
                    request,
                    'No results were saved. Confirm that students are registered for this subject.'
                )

            return redirect(f'{request.path}?subject={subject_id}')

    context = {
        'exam': exam,
        'subjects': subjects,
        'selected_subject': selected_subject,
        'students': students,
        'existing_results': existing_results,
        'title': f'Enter Results - {exam.name}',
        'school': SchoolSetting.objects.first(),
    }

    return render(request, 'performance/exam_results_entry.html', context)
# ========== TEMPORARY PDF DOWNLOAD (Disabled) ==========
# Comment out the original download_fee_structure function and use this one temporarily

def download_fee_structure(request, fee_structure_id):
    """Download fee structure - Currently disabled due to ReportLab"""
    messages.warning(request, "PDF download is temporarily unavailable. Please check back later.")
    return redirect('digitallibrary:fee_structure_list')
# ========== PERFORMANCE ANALYTICS HELPER FUNCTIONS ==========

def _safe_mean(values):
    return round(mean(values), 1) if values else 0.0


def _grade_from_score_analytics(score: float) -> str:
    if score >= 80:
        return "A"
    if score >= 75:
        return "A-"
    if score >= 70:
        return "B+"
    if score >= 65:
        return "B"
    if score >= 60:
        return "B-"
    if score >= 55:
        return "C+"
    if score >= 50:
        return "C"
    if score >= 45:
        return "C-"
    if score >= 40:
        return "D+"
    if score >= 35:
        return "D"
    return "E"


def _status_from_change(change: float) -> str:
    if change >= 3:
        return "Improving"
    if change <= -3:
        return "Declining"
    return "Stable"


def _consistency_label(scores):
    if len(scores) < 2:
        return "Insufficient Data", 0.0
    deviation = round(pstdev(scores), 1)
    if deviation <= 5:
        return "Very Consistent", deviation
    if deviation <= 10:
        return "Moderately Consistent", deviation
    return "Needs Monitoring", deviation


def _build_parent_summary(student_name, latest_avg, avg_change, performance_status, strong_subjects, risk_subjects, declining_subjects):
    summary = []
    summary.append(f"{student_name} currently has an overall average of {latest_avg:.1f}% and is classified as {performance_status.lower()}.")
    if avg_change > 0:
        summary.append(f"This is an improvement of {avg_change:.1f}% from the previous term.")
    elif avg_change < 0:
        summary.append(f"This is a drop of {abs(avg_change):.1f}% from the previous term.")
    else:
        summary.append("Performance is stable compared to the previous term.")
    if strong_subjects:
        summary.append(f"Strong performance is seen in {', '.join(strong_subjects[:3])}.")
    if declining_subjects:
        summary.append(f"The main subjects needing closer monitoring are {', '.join(declining_subjects[:3])}.")
    if risk_subjects:
        summary.append(f"Immediate support is recommended in {', '.join(risk_subjects[:3])}.")
    if not risk_subjects and not declining_subjects:
        summary.append("The learner is maintaining a healthy academic pattern and should continue with the current study routine.")
    return " ".join(summary)

# ========== CUSTOM LOGIN VIEW ==========


class CustomLoginView(LoginView):
    template_name = 'digitallibrary/login.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            context['school'] = SchoolSetting.objects.first()
        except:
            context['school'] = None
        return context
    
    def get_success_url(self):
        # Get tenant from the URL path
        path = self.request.path
        match = re.match(r'^/tenant/([^/]+)/app/login/', path)
        if match:
            tenant_schema = match.group(1)
            return f'/tenant/{tenant_schema}/app/dashboard/'
        return '/app/dashboard/'
        
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Q
import logging

logger = logging.getLogger(__name__)


from django.shortcuts import render, redirect
from django.db import connection
from django.utils import timezone
from django.db.models import Q, Sum
import logging

logger = logging.getLogger(__name__)


def home(request):
    """
    Home page - Shows landing page for public schema, dashboard for tenants
    """
    from django.db import connection
    from django.shortcuts import render, redirect
    from .models import Resource, Announcement, SchoolSetting, Student
    from django.utils import timezone
    from django.db.models import Q
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Get current schema name and host
    current_schema = connection.schema_name
    host = request.get_host().split(':')[0]
    
    print(f"\n{'='*60}")
    print(f"HOME VIEW - Schema: {current_schema}")
    print(f"Host: {host}")
    print(f"Path: {request.path}")
    print(f"{'='*60}\n")
    
    # If this is public schema, show public landing page
    if current_schema == 'public':
        # If accessing /app/ on public schema, redirect to root
        if request.path.startswith('/app/'):
            print(f"🔄 Redirecting /app/ to / for public schema")
            return redirect('/')
        
        # ========== PUBLIC LANDING PAGE ==========
        print("📌 Showing PUBLIC landing page")
        
        # Get platform metrics - SAFE VERSION without UserProfile
        from tenants.models import School
        from django.core.cache import cache
        from django_tenants.utils import tenant_context
        from django.db import connection as db_connection
        
        metrics = cache.get('platform_metrics')
        if not metrics:
            all_schools = School.objects.filter(is_active=True)
            total_schools = all_schools.count()
            
            total_teachers = 0
            total_students = 0
            total_resources = 0
            total_views = 0
            
            for school_tenant in all_schools:
                try:
                    with tenant_context(school_tenant):
                        # Use direct SQL instead of UserProfile to avoid errors
                        with db_connection.cursor() as cursor:
                            # Count teachers (staff users)
                            cursor.execute("SELECT COUNT(*) FROM auth_user WHERE is_staff = true AND is_superuser = false AND is_active = true")
                            result = cursor.fetchone()
                            total_teachers += result[0] if result else 0
                            
                            # Count students (non-staff users)
                            cursor.execute("SELECT COUNT(*) FROM auth_user WHERE is_staff = false AND is_active = true")
                            result = cursor.fetchone()
                            total_students += result[0] if result else 0
                        
                        # Count resources
                        total_resources += Resource.objects.count()
                        
                        # Count views
                        total_views += Resource.objects.aggregate(models.Sum('views'))['views__sum'] or 0
                except Exception as e:
                    logger.error(f"Error processing tenant {school_tenant.schema_name}: {e}")
            
            metrics = {
                'total_schools': total_schools,
                'total_teachers': total_teachers,
                'total_students': total_students,
                'total_resources': total_resources,
                'total_views': total_views,
            }
            cache.set('platform_metrics', metrics, 3600)
        
        # Get public announcements
        announcements = Announcement.objects.filter(
            Q(expires_at__isnull=True) | Q(expires_at__gt=timezone.now()),
            target_audience='all'
        ).order_by('-is_featured', '-created_at')[:5]
        
        # Return landing page template
        return render(request, "digitallibrary/landing_page.html", {
            "is_public_schema": True,
            "metrics": metrics,
            "school": None,
            "school_name": "ShuleHub",
            "latest": [],
            "announcements": announcements,
            "total_resources": metrics['total_resources'],
            "total_teachers": metrics['total_teachers'],
            "user_role": "Guest",
            "unread_count": 0,
            "notification_unread_count": 0,
            "children": [],
        })
    
       
    # ========== TENANT DASHBOARD (PUBLIC VIEW - NO LOGIN REQUIRED) ==========
    print("📌 Showing TENANT dashboard (public view)")
    
    # Get school setting
    school = SchoolSetting.objects.first()
    school_name = school.name if school else "ShuleHub"
    
    if school:
        print(f"✅ School: {school_name}")
    else:
        print(f"⚠️ No SchoolSetting found in {current_schema}")
    
    # Get latest resources (public)
    latest = list(Resource.objects.all().order_by("-created_at")[:8])
    total_resources = Resource.objects.count()
    
    # Get teacher count using SQL (no UserProfile)
    total_teachers = 0
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM auth_user WHERE is_staff = true AND is_superuser = false AND is_active = true")
            result = cursor.fetchone()
            total_teachers = result[0] if result else 0
    except Exception as e:
        print(f"Error counting teachers: {e}")
    
    print(f"📚 Resources: {total_resources}, Teachers: {total_teachers}")
    
    # Get announcements (public)
    announcements = list(Announcement.objects.filter(
        Q(expires_at__isnull=True) | Q(expires_at__gt=timezone.now())
    ).order_by("-is_featured", "-created_at")[:5])
    
    print(f"📢 Announcements: {len(announcements)}")
    
    # User role - only show if authenticated
    user_role = "Guest"
    children = []
    unread_count = 0
    notification_unread_count = 0
    show_admin_panel = False
    
    if request.user.is_authenticated:
        try:
            # Try to get role from profile if it exists
            if hasattr(request.user, 'profile') and request.user.profile:
                user_role = request.user.profile.role
            else:
                # Fallback: determine role from is_staff
                if request.user.is_staff and not request.user.is_superuser:
                    user_role = 'teacher'
                elif request.user.is_superuser:
                    user_role = 'admin'
                else:
                    user_role = 'student'
            
            print(f"👤 User role: {user_role}")
            show_admin_panel = user_role in ['admin', 'principal', 'teacher', 'bursar', 'secretary']
            
            if user_role == 'parent':
                phone = None
                if hasattr(request.user, 'profile') and request.user.profile:
                    phone = request.user.profile.phone_number
                if phone:
                    children = Student.objects.filter(
                        Q(parent_phone=phone) | Q(parent_alternative_phone=phone),
                        is_active=True
                    )
                    print(f"👶 Children: {children.count()}")
            
            # Get unread counts for authenticated users (safe try-except)
            try:
                from .models import AnnouncementRead, Notification
                unread_count = AnnouncementRead.objects.filter(user=request.user, read=False).count()
                notification_unread_count = Notification.objects.filter(recipient=request.user, is_read=False).count()
            except Exception as e:
                print(f"Error getting unread counts: {e}")
                unread_count = 0
                notification_unread_count = 0
                
        except Exception as e:
            print(f"Error getting user data: {e}")
    
    context = {
        "is_public_schema": False,
        "school": school,
        "school_name": school_name,
        "latest": latest,
        "announcements": announcements,
        "total_resources": total_resources,
        "total_teachers": total_teachers,
        "user_role": user_role.capitalize() if user_role != "Guest" else "Guest",
        "unread_count": unread_count,
        "notification_unread_count": notification_unread_count,
        "children": children,
        "show_admin_panel": show_admin_panel,
    }
    
    print(f"\n✅ Returning tenant dashboard with {len(announcements)} announcements")
    return render(request, "digitallibrary/home.html", context)

def logout_view(request):
    """Custom logout view"""
    from django.contrib.auth import logout
    from django.shortcuts import redirect
    from django.contrib import messages
    
    try:
        ActivityLog.objects.create(
            user=request.user, 
            action="logout", 
            description="User logged out"
        )
    except Exception:
        pass
    logout(request)
    messages.success(request, "You have been successfully logged out.")
    return redirect('/login/')
# digitallibrary/views.py

from django.shortcuts import render, get_object_or_404
from django.core.paginator import Paginator
from django.db.models import Q, Count
from .models import Resource, Subject, Category, SchoolSetting
from .forms import ResourceFilterForm

def library_list(request):
    """Display list of library resources with filtering"""
    
    # Start with all resources (not just active - depends on your model)
    # If you have an is_active field, uncomment the filter below
    resources = Resource.objects.all().order_by('-created_at')
    # resources = Resource.objects.filter(is_active=True).order_by('-created_at')
    
    # Get filter parameters from request
    subject_id = request.GET.get('subject')
    grade = request.GET.get('grade')
    year = request.GET.get('year')
    search_query = request.GET.get('q')
    category_id = request.GET.get('category')
    resource_type = request.GET.get('type')
    paper_type = request.GET.get('paper_type')
    
    # Apply filters
    if subject_id and subject_id.isdigit():
        resources = resources.filter(subject_id=int(subject_id))
    
    if grade and grade.strip():
        resources = resources.filter(grade=grade)
    
    if year and year.strip() and year != 'None':
        resources = resources.filter(year=year)
    
    if category_id and category_id.isdigit():
        resources = resources.filter(category_id=int(category_id))
    
    if resource_type and resource_type.strip():
        resources = resources.filter(resource_type=resource_type)
    
    if paper_type and paper_type.strip():
        resources = resources.filter(paper_type=paper_type)
    
    if search_query and search_query.strip():
        resources = resources.filter(
            Q(title__icontains=search_query) |
            Q(author__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(subject__name__icontains=search_query) |
            Q(year__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(resources, 24)
    page = request.GET.get('page', 1)
    resources_page = paginator.get_page(page)
    
    # Get filter options for dropdowns
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    categories = Category.objects.filter(is_active=True).order_by('name')
    
    # Get unique grades (clean and deduplicated)
    grades_raw = Resource.objects.exclude(grade__isnull=True).exclude(grade='').values_list('grade', flat=True).distinct()
    
    # Clean and standardize grades
    grade_mapping = {
        'form1': 'Form 1', 'form 1': 'Form 1', 'Form1': 'Form 1',
        'form2': 'Form 2', 'form 2': 'Form 2', 'Form2': 'Form 2',
        'form3': 'Form 3', 'form 3': 'Form 3', 'Form3': 'Form 3',
        'form4': 'Form 4', 'form 4': 'Form 4', 'Form4': 'Form 4',
        'grade1': 'Grade 1', 'grade 1': 'Grade 1', 'Grade1': 'Grade 1',
        'grade2': 'Grade 2', 'grade 2': 'Grade 2', 'Grade2': 'Grade 2',
        'grade3': 'Grade 3', 'grade 3': 'Grade 3', 'Grade3': 'Grade 3',
        'grade4': 'Grade 4', 'grade 4': 'Grade 4', 'Grade4': 'Grade 4',
        'grade5': 'Grade 5', 'grade 5': 'Grade 5', 'Grade5': 'Grade 5',
        'grade6': 'Grade 6', 'grade 6': 'Grade 6', 'Grade6': 'Grade 6',
        'grade7': 'Grade 7', 'grade 7': 'Grade 7', 'Grade7': 'Grade 7',
        'grade8': 'Grade 8', 'grade 8': 'Grade 8', 'Grade8': 'Grade 8',
        'general': 'General', 'General': 'General', 'GENERAL': 'General',
        'all': 'General', 'All': 'General',
    }
    
    cleaned_grades = set()
    for g in grades_raw:
        grade_lower = str(g).lower().strip()
        if grade_lower in grade_mapping:
            cleaned_grades.add(grade_mapping[grade_lower])
        elif g:
            cleaned_grades.add(str(g).strip())
    
    # Sort grades: Form/Grade 1-4/8 first, then alphabetical
    def grade_sort_key(g):
        order = {
            'Form 1': 1, 'Grade 1': 1,
            'Form 2': 2, 'Grade 2': 2,
            'Form 3': 3, 'Grade 3': 3,
            'Form 4': 4, 'Grade 4': 4,
            'Grade 5': 5, 'Grade 6': 6, 'Grade 7': 7, 'Grade 8': 8,
            'General': 99
        }
        return order.get(g, 100)
    
    grades = sorted(list(cleaned_grades), key=grade_sort_key)
    
    # Get unique years (deduplicated and sorted descending)
    years_raw = Resource.objects.exclude(year__isnull=True).exclude(year='').exclude(year='N/A').values_list('year', flat=True).distinct()
    
    valid_years = set()
    for y in years_raw:
        try:
            year_int = int(y)
            if 2000 <= year_int <= 2030:
                valid_years.add(year_int)
        except (ValueError, TypeError):
            pass
    
    years = sorted(list(valid_years), reverse=True)
    
    # Get unique paper types
    paper_types = Resource.objects.exclude(paper_type__isnull=True).exclude(paper_type='').values_list('paper_type', flat=True).distinct().order_by('paper_type')
    
    # Get resource types
    resource_types = Resource.objects.exclude(resource_type__isnull=True).exclude(resource_type='').values_list('resource_type', flat=True).distinct()
    
    # Get selected subject name for display
    selected_subject_name = None
    if subject_id and subject_id.isdigit():
        try:
            selected_subject = Subject.objects.get(id=int(subject_id))
            selected_subject_name = selected_subject.name
        except Subject.DoesNotExist:
            pass
    
    school = SchoolSetting.objects.first()
    
    context = {
        'resources': resources_page,
        'subjects': subjects,
        'categories': categories,
        'grades': grades,
        'years': years,
        'paper_types': paper_types,
        'resource_types': resource_types,
        'selected_subject': subject_id,
        'selected_grade': grade,
        'selected_year': year,
        'selected_category': category_id,
        'selected_type': resource_type,
        'selected_paper_type': paper_type,
        'selected_subject_name': selected_subject_name,
        'search_query': search_query,
        'total_count': resources.count(),
        'school': school,
    }
    
    return render(request, 'digitallibrary/library_list.html', context)


def resource_detail(request, pk):
    """Display resource details"""
    resource = get_object_or_404(Resource, pk=pk)
    
    # Increment view count (if you have this method)
    if hasattr(resource, 'increment_views'):
        resource.increment_views()
    
    school = SchoolSetting.objects.first()
    
    return render(request, "digitallibrary/resource_detail.html", {
        "resource": resource,
        "school": school
    })

def logout_view(request):
    """Custom logout view"""
    from django.contrib.auth import logout
    from django.shortcuts import redirect
    from django.contrib import messages
    
    try:
        ActivityLog.objects.create(
            user=request.user, 
            action="logout", 
            description="User logged out"
        )
    except Exception:
        pass
    logout(request)
    messages.success(request, "You have been successfully logged out.")
    return redirect('/login/')
# ========== RESOURCE UPLOAD AND MANAGEMENT VIEWS ==========

def can_upload(user):
    """Check if user can upload resources"""
    try:
        profile = user.profile
        return profile.role in ["admin", "teacher", "principal"]
    except Exception:
        return False


@login_required
def upload_resource(request):
    """Upload a new resource"""
    from django.db import connection
    from .forms import ResourceForm
    from .models import Subject, Category, SchoolSetting
    
    if not can_upload(request.user):
        messages.error(request, "Access Denied: Only teachers and administrators can upload resources.")
        return redirect("digitallibrary:library_list")

    if connection.schema_name == 'public':
        return redirect("digitallibrary:home")
    
    school = SchoolSetting.objects.first()

    if request.method == "POST":
        form = ResourceForm(request.POST, request.FILES)
        if form.is_valid():
            resource = form.save(commit=False)
            resource.uploaded_by = request.user
            resource.save()
            try:
                ActivityLog.objects.create(
                    user=request.user,
                    action="upload",
                    description=f"Uploaded resource: {resource.title}",
                )
            except Exception:
                pass
            messages.success(request, "Resource uploaded successfully!")
            if request.user.profile.role == "admin":
                return redirect("digitallibrary:library_admin_resources")
            else:
                return redirect("digitallibrary:my_uploads")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = ResourceForm()

    subjects = Subject.objects.all().order_by("name")
    categories = Category.objects.all().order_by("name")
    year_choices = get_year_choices()
    recent_uploads = Resource.objects.filter(uploaded_by=request.user).order_by("-created_at")[:5]
    
    return render(request, "digitallibrary/upload_resource.html", {
        "form": form, 
        "subjects": subjects,
        "categories": categories,
        "year_choices": year_choices,
        "school": school,
        "recent_uploads": recent_uploads,
        "is_teacher": request.user.profile.role == "teacher"
    })


@login_required
def my_uploads(request):
    """View user's uploaded resources"""
    from .models import SchoolSetting
    from django.core.paginator import Paginator
    
    if not can_upload(request.user):
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    school = SchoolSetting.objects.first()
    all_resources = Resource.objects.filter(uploaded_by=request.user).order_by("-created_at")
    total_uploads = all_resources.count()
    recent_uploads = all_resources.filter(created_at__gte=timezone.now() - timedelta(days=7)).count()
    total_views = all_resources.aggregate(Sum('views'))['views__sum'] or 0
    
    paginator = Paginator(all_resources, 12)
    page = request.GET.get("page", 1)
    resources = paginator.get_page(page)
    
    return render(request, "digitallibrary/my_uploads.html", {
        "resources": resources,
        "school": school,
        "total_uploads": total_uploads,
        "recent_uploads": recent_uploads,
        "total_views": total_views,
        "is_teacher": request.user.profile.role == "teacher"
    })


@login_required
def edit_my_resource(request, pk):
    """Edit user's uploaded resource"""
    from .forms import ResourceForm
    from .models import Subject, Category, SchoolSetting
    
    resource = get_object_or_404(Resource, pk=pk)
    if resource.uploaded_by != request.user and request.user.profile.role != "admin":
        messages.error(request, "You don't have permission to edit this resource.")
        return redirect("digitallibrary:library_list")
    
    school = SchoolSetting.objects.first()
    
    if request.method == "POST":
        form = ResourceForm(request.POST, request.FILES, instance=resource)
        if form.is_valid():
            form.save()
            messages.success(request, "Resource updated successfully!")
            try:
                ActivityLog.objects.create(
                    user=request.user,
                    action="edit",
                    description=f"Edited resource: {resource.title}",
                )
            except Exception:
                pass
            if request.user.profile.role == "admin":
                return redirect("digitallibrary:library_admin_resources")
            else:
                return redirect("digitallibrary:my_uploads")
    else:
        form = ResourceForm(instance=resource)
    
    subjects = Subject.objects.all().order_by("name")
    categories = Category.objects.all().order_by("name")
    year_choices = get_year_choices()
    
    return render(request, "digitallibrary/edit_resource.html", {
        "form": form,
        "resource": resource,
        "subjects": subjects,
        "categories": categories,
        "year_choices": year_choices,
        "school": school,
        "is_teacher": request.user.profile.role == "teacher"
    })


@login_required
@require_POST
def delete_my_resource(request, pk):
    """Delete user's uploaded resource"""
    resource = get_object_or_404(Resource, pk=pk)
    if resource.uploaded_by != request.user and request.user.profile.role != "admin":
        messages.error(request, "You don't have permission to delete this resource.")
        return redirect("digitallibrary:library_list")
    
    title = resource.title
    resource.delete()
    try:
        ActivityLog.objects.create(
            user=request.user,
            action="delete",
            description=f"Deleted resource: {title}",
        )
    except Exception:
        pass
    messages.success(request, f"Resource '{title}' deleted successfully.")
    if request.user.profile.role == "admin":
        return redirect("digitallibrary:library_admin_resources")
    else:
        return redirect("digitallibrary:my_uploads")


def resource_detail(request, pk):
    """Display resource details"""
    resource = get_object_or_404(Resource, pk=pk)
    resource.increment_views()
    school = SchoolSetting.objects.first()
    return render(request, "digitallibrary/resource_detail.html", {
        "resource": resource,
        "school": school
    })


def library_list(request):
    """Display list of library resources"""
    from .models import Resource, Subject, Category, SchoolSetting
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    resources = Resource.objects.all().order_by('-created_at')
    
    subject_id = request.GET.get('subject')
    grade = request.GET.get('grade')
    year = request.GET.get('year')
    q = request.GET.get('q')
    category_id = request.GET.get('category')
    resource_type = request.GET.get('type')
    
    if subject_id:
        resources = resources.filter(subject_id=subject_id)
    if grade:
        resources = resources.filter(grade=grade)
    if year:
        resources = resources.filter(year=year)
    if category_id:
        resources = resources.filter(category_id=category_id)
    if resource_type:
        resources = resources.filter(resource_type=resource_type)
    if q:
        resources = resources.filter(
            Q(title__icontains=q) |
            Q(author__icontains=q) |
            Q(description__icontains=q) |
            Q(subject__name__icontains=q)
        )
    
    paginator = Paginator(resources, 24)
    page = request.GET.get('page', 1)
    resources_page = paginator.get_page(page)
    
    subjects = Subject.objects.all().order_by('name')
    categories = Category.objects.all().order_by('name')
    school = SchoolSetting.objects.first()
    grades = Resource.objects.values_list('grade', flat=True).distinct().exclude(grade='').exclude(grade=None)
    years = Resource.objects.values_list('year', flat=True).distinct().exclude(year='').exclude(year=None).order_by('-year')
    
    context = {
        'resources': resources_page,
        'subjects': subjects,
        'categories': categories,
        'grades': grades,
        'years': years,
        'selected_subject': subject_id,
        'selected_grade': grade,
        'selected_year': year,
        'selected_category': category_id,
        'selected_type': resource_type,
        'search_query': q,
        'school': school,
    }
    return render(request, 'digitallibrary/library_list.html', context)


def logout_view(request):
    """Custom logout view"""
    from django.contrib.auth import logout
    
    try:
        ActivityLog.objects.create(
            user=request.user, 
            action="logout", 
            description="User logged out"
        )
    except Exception:
        pass
    logout(request)
    messages.success(request, "You have been successfully logged out.")
    return redirect('/login/')

# ========== AI SEARCH VIEW ==========

def ai_search_page(request):
    """AI-powered semantic search page"""
    from django.db import connection
    from .models import SchoolSetting
    
    query = request.GET.get("q", "").strip()
    results = []
    if query:
        # results = search_ai(query, k=8)  # Uncomment when AI is available
        pass
    school = SchoolSetting.objects.first() if connection.schema_name != 'public' else None

    return render(request, "digitallibrary/ai_search.html", {
        "q": query,
        "results": results,
        "school": school,
    })


# ========== PRINTING PORTAL VIEWS ==========

@login_required
@tenant_app_view
def printing_portal(request):
    """Printing portal for teachers"""
    from .models import PrintJob, SchoolSetting, UserProfile
    
    profile, _created = UserProfile.objects.get_or_create(user=request.user)
    school = SchoolSetting.objects.first()

    if request.method == "POST":
        file = request.FILES.get("file")
        copies = request.POST.get("copies", 1)
        color = request.POST.get("color", "bw")
        if file:
            job = PrintJob.objects.create(
                file=file,
                teacher=request.user,
                copies=copies,
                color=color,
                status="Pending",
                downloaded=False,
            )
            try:
                ActivityLog.objects.create(
                    user=request.user,
                    action="print_submit",
                    description=f"Submitted print job: {file.name}",
                )
            except Exception as e:
                print(f"Error creating notifications: {e}")
            messages.success(request, "Print request submitted successfully.")
        else:
            messages.error(request, "Please select a file to print.")
        return redirect("digitallibrary:printing_portal")

    if profile.role in ["secretary", "admin"]:
        jobs = PrintJob.objects.all().order_by("-created_at")
        highlight_id = request.GET.get('highlight')
        if highlight_id:
            try:
                highlight_id = int(highlight_id)
            except ValueError:
                highlight_id = None
    else:
        jobs = PrintJob.objects.filter(teacher=request.user).order_by("-created_at")
        highlight_id = None

    pending_count = jobs.filter(status="Pending").count()
    return render(request, "digitallibrary/printing_portal.html", {
        "jobs": jobs,
        "role": profile.role,
        "pending_count": pending_count,
        "school": school,
        "highlight_id": highlight_id,
    })


@login_required
def mark_as_downloaded(request, job_id):
    """Mark print job as downloaded"""
    from .models import PrintJob, ActivityLog, Notification
    
    if request.user.profile.role not in ["secretary", "admin"]:
        messages.error(request, "You don't have permission to do that.")
        return redirect("digitallibrary:printing_portal")
    job = get_object_or_404(PrintJob, id=job_id)
    job.mark_as_downloaded(user=request.user)
    try:
        ActivityLog.objects.create(
            user=request.user,
            action="print_download",
            description=f"Downloaded print job: {job.file.name}",
        )
    except Exception as e:
        print(f"Error creating notification: {e}")
    messages.success(request, f"Job '{job.file.name}' marked as downloaded.")
    return redirect("digitallibrary:printing_portal")


@login_required
def mark_as_completed(request, job_id):
    """Mark print job as completed"""
    from .models import PrintJob, ActivityLog, Notification
    
    if request.user.profile.role not in ["secretary", "admin"]:
        messages.error(request, "You don't have permission to do that.")
        return redirect("digitallibrary:printing_portal")
    job = get_object_or_404(PrintJob, id=job_id)
    job.mark_as_completed(user=request.user)
    try:
        ActivityLog.objects.create(
            user=request.user,
            action="print_complete",
            description=f"Completed print job: {job.file.name}",
        )
    except Exception as e:
        print(f"Error creating notification: {e}")
    messages.success(request, f"Job '{job.file.name}' marked as completed.")
    return redirect("digitallibrary:printing_portal")


@login_required
def download_print_file(request, job_id):
    """Download print job file"""
    from .models import PrintJob, ActivityLog, Notification
    import mimetypes
    import os
    
    print_job = get_object_or_404(PrintJob, id=job_id)
    has_permission = False
    if request.user == print_job.teacher:
        has_permission = True
    try:
        if request.user.profile.role in ["secretary", "admin"]:
            has_permission = True
    except:
        pass
    if not has_permission:
        raise Http404("You don't have permission to download this file")
    if not print_job.file:
        raise Http404("No file associated with this print job")
    file_path = print_job.file.path
    if not os.path.exists(file_path):
        raise Http404("File not found")
    content_type, _ = mimetypes.guess_type(file_path)
    if not content_type:
        content_type = 'application/octet-stream'
    if print_job.status in ["Pending", "Ready"]:
        print_job.status = "Downloaded"
        print_job.downloaded_at = timezone.now()
        print_job.save()
    try:
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            response['Content-Length'] = os.path.getsize(file_path)
            return response
    except Exception as e:
        raise Http404(f"Error reading file: {e}")


@login_required
def print_job_detail(request, job_id):
    """View print job details"""
    job = get_object_or_404(PrintJob, id=job_id)
    try:
        user_role = request.user.profile.role
    except:
        user_role = "teacher"
    if user_role in ["secretary", "admin"] or job.teacher == request.user:
        return redirect(f"{reverse('digitallibrary:printing_portal')}?highlight={job_id}")
    else:
        messages.error(request, "You don't have permission to view this print job.")
        return redirect("digitallibrary:printing_portal")


# ========== LIBRARY ADMIN VIEWS ==========

@login_required
def library_admin_dashboard(request):
    """Library admin dashboard"""
    from .models import Resource, Announcement, SchoolSetting
    
    if request.user.profile.role not in ["admin", "principal"]:
        messages.error(request, "Access Denied. Library Admin access only.")
        return redirect("digitallibrary:home")
    total_resources = Resource.objects.count()
    total_announcements = Announcement.objects.count()
    recent_resources = Resource.objects.order_by("-created_at")[:5]
    recent_announcements = Announcement.objects.order_by("-created_at")[:5]
    school = SchoolSetting.objects.first()
    context = {
        "total_resources": total_resources,
        "total_announcements": total_announcements,
        "recent_resources": recent_resources,
        "recent_announcements": recent_announcements,
        "school": school,
    }
    return render(request, "digitallibrary/library_admin/dashboard.html", context)


@login_required
def library_admin_resources(request):
    """Library admin resource management"""
    from .models import Resource, SchoolSetting
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    if request.user.profile.role not in ["admin", "principal"]:
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    resources = Resource.objects.all().order_by("-created_at")
    q = request.GET.get("q", "")
    if q:
        resources = resources.filter(
            Q(title__icontains=q) | Q(author__icontains=q) | Q(grade__icontains=q) |
            Q(year__icontains=q) | Q(subject__name__icontains=q)
        )
    paginator = Paginator(resources, 20)
    page = request.GET.get("page", 1)
    resources = paginator.get_page(page)
    school = SchoolSetting.objects.first()
    return render(request, "digitallibrary/library_admin/resources.html", {
        "resources": resources,
        "q": q,
        "school": school
    })


@login_required
def library_admin_resource_edit(request, pk=None):
    """Edit or add resource in admin panel"""
    from .forms import ResourceForm
    from .models import Subject, SchoolSetting
    import datetime
    import os
    
    if request.user.profile.role not in ["admin", "principal"]:
        messages.error(request, "Access Denied. Admin access required.")
        return redirect("digitallibrary:home")
    
    if pk:
        resource = get_object_or_404(Resource, id=pk)
        if request.method == 'POST':
            form = ResourceForm(request.POST, request.FILES, instance=resource)
            if form.is_valid():
                updated_resource = form.save()
                messages.success(request, 'Resource updated successfully!')
                return redirect('digitallibrary:library_admin_resources')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        else:
            form = ResourceForm(instance=resource)
        current_file_name = None
        current_file_size = None
        if resource.file:
            current_file_name = os.path.basename(resource.file.name)
            try:
                current_file_size = resource.file.size
            except:
                pass
    else:
        if request.method == 'POST':
            form = ResourceForm(request.POST, request.FILES)
            if form.is_valid():
                resource = form.save(commit=False)
                resource.uploaded_by = request.user
                resource.save()
                messages.success(request, 'Resource created successfully!')
                return redirect('digitallibrary:library_admin_resources')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        else:
            form = ResourceForm()
        current_file_name = None
        current_file_size = None
    
    current_year = datetime.datetime.now().year
    years = list(range(current_year + 5, 1949, -1))
    subjects = Subject.objects.all().order_by('name')
    school = SchoolSetting.objects.first()
    
    context = {
        'form': form,
        'resource': resource if pk else None,
        'years': years,
        'subjects': subjects,
        'title': 'Edit Resource' if pk else 'Add Resource',
        'school': school,
        'current_file_name': current_file_name,
        'current_file_size': current_file_size,
    }
    return render(request, 'digitallibrary/library_admin/resource_form.html', context)


@login_required
def library_admin_resource_delete(request, pk):
    """Delete resource from admin panel"""
    from .models import Resource, SchoolSetting
    
    if request.user.profile.role not in ["admin", "principal"]:
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    resource = get_object_or_404(Resource, pk=pk)
    school = SchoolSetting.objects.first()
    if request.method == "POST":
        resource.delete()
        messages.success(request, "Resource deleted successfully!")
        return redirect("digitallibrary:library_admin_resources")
    return render(request, "digitallibrary/library_admin/resource_confirm_delete.html", {
        "resource": resource,
        "school": school
    })
# ========== LIBRARY ADMIN ANNOUNCEMENTS VIEWS ==========

from django.db import connection

@login_required
def library_admin_announcements(request):
    """Library admin announcements management"""
    from .models import Announcement, SchoolSetting
    from .forms import AnnouncementFilterForm
    from django.db.models import Q
    from django.utils import timezone
    
    # SAFETY CHECK: Prevent access/queries on the public schema
    if connection.schema_name == 'public':
        messages.error(request, "This feature is only available for school tenants.")
        return redirect("digitallibrary:home")

    if request.user.profile.role not in ["admin", "principal"]:
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    filter_form = AnnouncementFilterForm(request.GET)
    announcements = Announcement.objects.all().order_by("-created_at")
    
    if filter_form.is_valid():
        audience = filter_form.cleaned_data.get('audience')
        status = filter_form.cleaned_data.get('status')
        search = filter_form.cleaned_data.get('search')
        date_from = filter_form.cleaned_data.get('date_from')
        date_to = filter_form.cleaned_data.get('date_to')
        
        if audience:
            announcements = announcements.filter(target_audience=audience)
        if status == 'active':
            announcements = announcements.filter(Q(expires_at__isnull=True) | Q(expires_at__gt=timezone.now()))
        elif status == 'expired':
            announcements = announcements.filter(expires_at__lt=timezone.now())
        elif status == 'featured':
            announcements = announcements.filter(is_featured=True)
        if search:
            announcements = announcements.filter(Q(title__icontains=search) | Q(content__icontains=search))
        if date_from:
            announcements = announcements.filter(created_at__date__gte=date_from)
        if date_to:
            announcements = announcements.filter(created_at__date__lte=date_to)
    
    school = SchoolSetting.objects.first()
    return render(request, "digitallibrary/library_admin/announcements.html", {
        "announcements": announcements,
        "filter_form": filter_form,
        "school": school
    })


@login_required
def library_admin_announcement_add(request):
    """Add new announcement from admin panel"""
    from .models import SchoolSetting
    from .forms import AnnouncementForm
    
    # SAFETY CHECK: Prevent access/queries on the public schema
    if connection.schema_name == 'public':
        messages.error(request, "This feature is only available for school tenants.")
        return redirect("digitallibrary:home")

    if request.user.profile.role not in ["admin", "principal"]:
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    school = SchoolSetting.objects.first()
    if request.method == "POST":
        form = AnnouncementForm(request.POST, request.FILES)
        if form.is_valid():
            announcement = form.save(commit=False)
            announcement.author = request.user
            announcement.save()
            messages.success(request, "Announcement created successfully!")
            return redirect("digitallibrary:library_admin_announcements")
    else:
        form = AnnouncementForm()
    return render(request, "digitallibrary/library_admin/announcement_form.html", {
        "form": form,
        "title": "Add New Announcement",
        "school": school
    })


@login_required
def library_admin_announcement_edit(request, pk):
    """Edit announcement from admin panel"""
    from .models import Announcement, SchoolSetting
    from .forms import AnnouncementForm
    
    if request.user.profile.role not in ["admin", "principal"]:
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    announcement = get_object_or_404(Announcement, pk=pk)
    school = SchoolSetting.objects.first()
    if request.method == "POST":
        form = AnnouncementForm(request.POST, request.FILES, instance=announcement)
        if form.is_valid():
            form.save()
            messages.success(request, "Announcement updated successfully!")
            return redirect("digitallibrary:library_admin_announcements")
    else:
        form = AnnouncementForm(instance=announcement)
    return render(request, "digitallibrary/library_admin/announcement_form.html", {
        "form": form,
        "title": "Edit Announcement",
        "announcement": announcement,
        "school": school
    })


@login_required
def library_admin_announcement_delete(request, pk):
    """Delete announcement from admin panel"""
    from .models import Announcement, SchoolSetting
    
    if request.user.profile.role not in ["admin", "principal"]:
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    announcement = get_object_or_404(Announcement, pk=pk)
    school = SchoolSetting.objects.first()
    if request.method == "POST":
        announcement.delete()
        messages.success(request, "Announcement deleted successfully!")
        return redirect("digitallibrary:library_admin_announcements")
    return render(request, "digitallibrary/library_admin/announcement_confirm_delete.html", {
        "announcement": announcement,
        "school": school
    })


# ========== USER PROFILE AND ACTIVITY VIEWS ==========

@login_required
def user_profile(request):
    """View user profile"""
    from .models import SchoolSetting, Resource, PrintJob, ActivityLog
    from datetime import timedelta
    
    profile = request.user.profile
    activities = ActivityLog.objects.filter(user=request.user)[:20]
    print_jobs = PrintJob.objects.filter(teacher=request.user)[:10]
    total_uploads = Resource.objects.filter(uploaded_by=request.user).count()
    recent_uploads = Resource.objects.filter(
        uploaded_by=request.user,
        created_at__gte=timezone.now() - timedelta(days=30)
    ).count()
    recent_resources = Resource.objects.filter(uploaded_by=request.user).order_by('-created_at')[:5]
    school = SchoolSetting.objects.first()
    return render(request, "digitallibrary/profile.html", {
        "profile": profile,
        "activities": activities,
        "print_jobs": print_jobs,
        "total_uploads": total_uploads,
        "recent_uploads": recent_uploads,
        "recent_resources": recent_resources,
        "school": school,
    })


@login_required
def approve_teacher(request, user_id):
    """Approve teacher registration"""
    from django.contrib.auth.models import User
    
    if request.user.profile.role != "admin":
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    teacher = get_object_or_404(User, id=user_id)
    profile = teacher.profile
    if profile.role == "teacher":
        profile.is_approved = True
        profile.save()
        messages.success(request, f"Teacher {teacher.username} approved successfully.")
    return redirect("digitallibrary:manage_users")


@login_required
def activity_log(request):
    """View activity logs"""
    from .models import ActivityLog, SchoolSetting
    
    if request.user.profile.role != "admin":
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    activities = ActivityLog.objects.all()[:100]
    school = SchoolSetting.objects.first()
    return render(request, "digitallibrary/activity_log.html", {
        "activities": activities,
        "school": school
    })


# ========== ANNOUNCEMENT VIEWS ==========

@login_required
def announcement_list(request):
    """List announcements for users"""
    from .models import Announcement, AnnouncementRead, SchoolSetting
    from django.db.models import Q
    from django.utils import timezone
    
    user = request.user
    try:
        user_role = user.profile.role
    except:
        user_role = 'student'
    
    announcements = Announcement.objects.filter(Q(expires_at__isnull=True) | Q(expires_at__gt=timezone.now()))
    
    if user_role == 'admin' or user_role == 'principal':
        pass
    elif user_role == 'teacher':
        announcements = announcements.filter(Q(target_audience='all') | Q(target_audience='teachers') | Q(target_audience='staff'))
    elif user_role == 'student':
        announcements = announcements.filter(Q(target_audience='all') | Q(target_audience='students'))
    elif user_role == 'secretary':
        announcements = announcements.filter(Q(target_audience='all') | Q(target_audience='staff'))
    else:
        announcements = announcements.filter(target_audience='all')
    
    announcements = announcements.order_by("-is_featured", "-created_at")
    for announcement in announcements:
        AnnouncementRead.mark_as_read(announcement, user)
    
    unread_count = AnnouncementRead.get_unread_count(user)
    school = SchoolSetting.objects.first()
    return render(request, "digitallibrary/announcement_list.html", {
        "announcements": announcements,
        "unread_count": unread_count,
        "user_role": user_role,
        "school": school
    })


@login_required
def announcement_detail(request, pk):
    """View announcement details"""
    from .models import Announcement, AnnouncementRead, SchoolSetting
    
    announcement = get_object_or_404(Announcement, pk=pk)
    user = request.user
    
    try:
        user_role = user.profile.role
    except:
        user_role = 'student'
    
    can_view = False
    if user_role in ['admin', 'principal']:
        can_view = True
    elif announcement.target_audience == 'all':
        can_view = True
    elif announcement.target_audience == 'teachers' and user_role == 'teacher':
        can_view = True
    elif announcement.target_audience == 'students' and user_role == 'student':
        can_view = True
    elif announcement.target_audience == 'admin' and user_role in ['admin', 'principal']:
        can_view = True
    elif announcement.target_audience == 'staff' and user_role in ['admin', 'principal', 'teacher', 'secretary']:
        can_view = True
    
    if not can_view:
        messages.error(request, "You don't have permission to view this announcement.")
        return redirect('digitallibrary:announcement_list')
    
    AnnouncementRead.mark_as_read(announcement, user)
    read_stats = None
    if user_role in ['admin', 'principal']:
        read_stats = {
            'total_read': announcement.read_count(),
            'total_target': announcement.target_count(),
            'percentage': announcement.read_percentage(),
            'read_by': announcement.read_receipts.filter(read=True).select_related('user')[:10]
        }
    school = SchoolSetting.objects.first()
    return render(request, "digitallibrary/announcement_detail.html", {
        "announcement": announcement,
        "read_stats": read_stats,
        "school": school
    })


@login_required
def create_announcement(request):
    """Create a new announcement"""
    from .models import SchoolSetting
    from .forms import AnnouncementForm
    
    if request.user.profile.role not in ['admin', 'principal']:
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    school = SchoolSetting.objects.first()
    if request.method == "POST":
        form = AnnouncementForm(request.POST, request.FILES)
        if form.is_valid():
            announcement = form.save(commit=False)
            announcement.author = request.user
            announcement.save()
            messages.success(request, "Announcement created successfully!")
            return redirect("digitallibrary:announcement_list")
    else:
        form = AnnouncementForm()
    return render(request, "digitallibrary/create_announcement.html", {
        "form": form,
        "school": school
    })


@login_required
def edit_announcement(request, pk):
    """Edit an announcement"""
    from .models import Announcement, SchoolSetting
    from .forms import AnnouncementForm
    
    if request.user.profile.role != "admin":
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    announcement = get_object_or_404(Announcement, pk=pk)
    school = SchoolSetting.objects.first()
    if request.method == "POST":
        form = AnnouncementForm(request.POST, request.FILES, instance=announcement)
        if form.is_valid():
            form.save()
            messages.success(request, "Announcement updated successfully.")
            return redirect("digitallibrary:announcement_list")
    else:
        form = AnnouncementForm(instance=announcement)
    return render(request, "digitallibrary/edit_announcement.html", {
        "form": form,
        "announcement": announcement,
        "school": school
    })


@login_required
def delete_announcement(request, pk):
    """Delete an announcement"""
    from .models import Announcement, SchoolSetting
    
    if request.user.profile.role != "admin":
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    announcement = get_object_or_404(Announcement, pk=pk)
    school = SchoolSetting.objects.first()
    announcement.delete()
    messages.success(request, "Announcement deleted successfully.")
    return redirect("digitallibrary:announcement_list")


@login_required
def announcement_read_stats(request, pk):
    """Get announcement read statistics (AJAX)"""
    from .models import Announcement
    
    if request.user.profile.role not in ['admin', 'principal']:
        return JsonResponse({"error": "Unauthorized"}, status=403)
    
    announcement = get_object_or_404(Announcement, pk=pk)
    target_users = announcement.get_target_user_queryset()
    read_records = {
        r.user_id: {
            'read_at': r.read_at,
            'username': r.user.username,
            'role': r.user.profile.role if hasattr(r.user, 'profile') else 'unknown'
        }
        for r in announcement.read_receipts.filter(read=True).select_related('user')
    }
    stats = []
    for user in target_users:
        stats.append({
            'user_id': user.id,
            'username': user.username,
            'role': user.profile.role if hasattr(user, 'profile') else 'unknown',
            'read': user.id in read_records,
            'read_at': read_records.get(user.id, {}).get('read_at', None)
        })
    stats.sort(key=lambda x: (x['read'], x['username']))
    return JsonResponse({
        'announcement_id': announcement.id,
        'announcement_title': announcement.title,
        'target_audience': announcement.get_target_audience_display(),
        'total_target': len(stats),
        'total_read': announcement.read_count(),
        'read_percentage': announcement.read_percentage(),
        'stats': stats
    })
# ========== ADMIN DASHBOARD VIEWS ==========

@login_required
def admin_dashboard(request):
    """Main admin dashboard"""
    from .models import Resource, UserProfile, PrintJob, SchoolSetting
    from django.contrib.auth.models import User
    
    if request.user.profile.role != "admin":
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    total_resources = Resource.objects.count()
    total_users = User.objects.count()
    total_print_jobs = PrintJob.objects.count()
    pending_teachers = UserProfile.objects.filter(role="teacher", is_approved=False).count()
    school = SchoolSetting.objects.first()
    
    return render(request, "digitallibrary/admin_dashboard.html", {
        "total_resources": total_resources,
        "total_users": total_users,
        "total_print_jobs": total_print_jobs,
        "pending_teachers": pending_teachers,
        "school": school,
    })


@login_required
def dashboard_statistics(request):
    """Dashboard statistics view"""
    from .models import Resource, PrintJob, ActivityLog, SchoolSetting
    from django.db.models import Count
    
    if request.user.profile.role != "admin":
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    resources_by_grade = Resource.objects.values("grade").annotate(count=Count("id")).order_by("grade")
    resources_by_year = Resource.objects.values("year").annotate(count=Count("id")).order_by("-year")
    resources_by_subject = Resource.objects.values("subject__name").annotate(count=Count("id")).order_by("subject__name")
    print_jobs_by_status = PrintJob.objects.values("status").annotate(count=Count("id"))
    recent_activity = ActivityLog.objects.all()[:50]
    school = SchoolSetting.objects.first()
    
    return render(request, "digitallibrary/dashboard_statistics.html", {
        "resources_by_grade": resources_by_grade,
        "resources_by_year": resources_by_year,
        "resources_by_subject": resources_by_subject,
        "print_jobs_by_status": print_jobs_by_status,
        "recent_activity": recent_activity,
        "school": school,
    })


@login_required
def manage_users(request):
    """Manage system users"""
    from .models import UserProfile, SchoolSetting
    from django.contrib.auth.models import User
    
    if request.user.profile.role != "admin":
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    users = UserProfile.objects.select_related("user").all()
    school = SchoolSetting.objects.first()
    
    return render(request, "digitallibrary/manage_users.html", {
        "users": users,
        "school": school
    })


@login_required
def change_user_role(request, user_id):
    """Change user role"""
    from .models import UserProfile
    
    if request.user.profile.role != "admin":
        messages.error(request, "Access Denied.")
        return redirect("digitallibrary:home")
    
    profile = get_object_or_404(UserProfile, user_id=user_id)
    if request.method == "POST":
        new_role = request.POST.get("role")
        if new_role in dict(UserProfile.ROLE_CHOICES).keys():
            profile.role = new_role
            profile.save()
            messages.success(request, f"User role updated to {new_role}")
    return redirect("digitallibrary:manage_users")


# ========== NOTIFICATION VIEWS ==========

@login_required
def notification_list(request):
    """List user notifications"""
    from .models import Notification, SchoolSetting
    from django.core.paginator import Paginator
    from django.utils import timezone
    
    notifications = Notification.objects.filter(recipient=request.user, is_archived=False).order_by("-created_at")
    unread = notifications.filter(is_read=False)
    unread.update(is_read=True, read_at=timezone.now())
    
    paginator = Paginator(notifications, 20)
    page = request.GET.get('page', 1)
    notifications = paginator.get_page(page)
    school = SchoolSetting.objects.first()
    
    return render(request, 'digitallibrary/notifications.html', {
        'notifications': notifications,
        'school': school
    })


def api_notifications(request):
    """API endpoint for notifications - Safe version"""
    from django.db import connection
    from django.http import JsonResponse
    from django.utils import timezone
    from datetime import timedelta
    from django.db import ProgrammingError
    
    # For public schema, return empty
    if connection.schema_name == 'public':
        return JsonResponse({
            'unread_count': 0,
            'notifications': []
        })
    
    if not request.user.is_authenticated:
        return JsonResponse({'unread_count': 0, 'notifications': []})
    
    try:
        from digitallibrary.models import Notification
        
        notifications = Notification.objects.filter(
            recipient=request.user, 
            is_archived=False
        ).order_by("-created_at")[:20]
        
        def get_time_ago(created_at):
            from django.utils.timesince import timesince
            now = timezone.now()
            if created_at.date() == now.date():
                return f"{timesince(created_at)} ago"
            elif created_at.date() == now.date() - timedelta(days=1):
                return "Yesterday"
            else:
                return created_at.strftime("%b %d, %Y")
        
        # Try to get unread count safely
        try:
            unread_count = Notification.get_unread_count(request.user)
        except:
            unread_count = 0
        
        data = {
            'unread_count': unread_count,
            'notifications': [{
                'id': n.id,
                'title': n.title,
                'message': n.message,
                'link': n.link or '#',
                'type': n.notification_type,
                'is_read': n.is_read,
                'time_ago': get_time_ago(n.created_at),
                'created_at': n.created_at.isoformat()
            } for n in notifications]
        }
        return JsonResponse(data)
        
    except ProgrammingError as e:
        # Table doesn't exist yet - return empty data
        print(f"Notifications table not ready: {e}")
        return JsonResponse({
            'unread_count': 0,
            'notifications': []
        })
    except Exception as e:
        # Any other error - return empty data
        print(f"Error in api_notifications: {e}")
        return JsonResponse({
            'unread_count': 0,
            'notifications': []
        })

def api_mark_notification_read(request, pk):
    """Mark notification as read - Safe version"""
    from django.db import connection
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from django.db import ProgrammingError
    
    if connection.schema_name == 'public':
        return JsonResponse({'success': True})
    
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    try:
        from digitallibrary.models import Notification
        notification = get_object_or_404(Notification, pk=pk, recipient=request.user)
        notification.mark_as_read()
        return JsonResponse({'success': True})
    except ProgrammingError as e:
        # Table doesn't exist
        print(f"Notification table not ready: {e}")
        return JsonResponse({'success': True})  # Pretend it worked
    except Exception as e:
        print(f"Error marking notification read: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def api_mark_all_read(request):
    """Mark all notifications as read - Safe version"""
    from django.db import connection
    from django.http import JsonResponse
    from django.utils import timezone
    from django.db import ProgrammingError
    
    if connection.schema_name == 'public':
        return JsonResponse({'success': True})
    
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    try:
        from digitallibrary.models import Notification
        Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True, read_at=timezone.now())
        return JsonResponse({'success': True})
    except ProgrammingError as e:
        # Table doesn't exist
        print(f"Notification table not ready: {e}")
        return JsonResponse({'success': True})  # Pretend it worked
    except Exception as e:
        print(f"Error marking all read: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def api_archive_notification(request, pk):
    """Archive a notification - Safe version"""
    from django.db import connection
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from django.db import ProgrammingError
    
    if connection.schema_name == 'public':
        return JsonResponse({'success': True})
    
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    try:
        from digitallibrary.models import Notification
        notification = get_object_or_404(Notification, pk=pk, recipient=request.user)
        notification.is_archived = True
        notification.save()
        return JsonResponse({'success': True})
    except ProgrammingError as e:
        # Table doesn't exist
        print(f"Notification table not ready: {e}")
        return JsonResponse({'success': True})  # Pretend it worked
    except Exception as e:
        print(f"Error archiving notification: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# ========== SUBJECT AND CATEGORY MANAGEMENT ==========

@login_required
def get_subjects(request):
    """Get all subjects (AJAX)"""
    from .models import Subject
    
    if request.user.profile.role not in ["admin", "principal"]:
        return JsonResponse({"error": "Unauthorized"}, status=403)
    
    subjects = Subject.objects.all().values("id", "name").order_by("name")
    return JsonResponse(list(subjects), safe=False)


@login_required
@require_POST
def add_subject(request):
    """Add a new subject via AJAX"""
    import json
    from .models import Subject
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        subject_name = request.POST.get('name', '').strip()
        
        if not subject_name:
            return JsonResponse({'status': 'error', 'error': 'Subject name is required'}, status=400)
        
        existing = Subject.objects.filter(name__iexact=subject_name).first()
        if existing:
            return JsonResponse({
                'status': 'success',
                'id': existing.id,
                'name': existing.name,
                'message': 'Subject already exists'
            })
        
        try:
            subject = Subject.objects.create(name=subject_name, is_active=True)
            return JsonResponse({
                'status': 'success',
                'id': subject.id,
                'name': subject.name
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'error': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'error': 'Invalid request'}, status=400)


@login_required
@require_POST
def delete_subject(request, pk):
    """Delete a subject"""
    from .models import Subject
    
    if request.user.profile.role != "admin":
        return JsonResponse({"error": "Unauthorized"}, status=403)
    
    try:
        subject = get_object_or_404(Subject, pk=pk)
        if subject.resources.exists():
            return JsonResponse({
                "error": "Cannot delete subject that is in use", 
                "resources_count": subject.resources.count()
            }, status=400)
        subject.delete()
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def get_categories(request):
    """Get all categories (AJAX)"""
    from .models import Category
    
    if request.user.profile.role not in ["admin", "principal"]:
        return JsonResponse({"error": "Unauthorized"}, status=403)
    
    categories = Category.objects.all().values("id", "name").order_by("name")
    return JsonResponse(list(categories), safe=False)


# ========== INCREMENT VIEW COUNT ==========

@login_required
def increment_resource_view(request, pk):
    """Increment resource view count via AJAX"""
    from .models import Resource
    
    resource = get_object_or_404(Resource, pk=pk)
    resource.increment_views()
    return JsonResponse({"success": True, "views": resource.views})
# ========== PUBLIC METRICS API ==========

from rest_framework.decorators import api_view
from rest_framework.response import Response

def test_metrics(request):
    """Test metrics API endpoint - Safe version"""
    from django.http import JsonResponse
    from django.db import connection
    
    try:
        with connection.cursor() as cursor:
            # Get schools count
            cursor.execute("SELECT COUNT(*) FROM tenants_school WHERE is_active = true")
            schools = cursor.fetchone()[0] or 0
            
            # Get teachers count
            cursor.execute("SELECT COUNT(*) FROM auth_user WHERE is_staff = true AND is_superuser = false AND is_active = true")
            teachers = cursor.fetchone()[0] or 0
            
            # Get students count
            cursor.execute("SELECT COUNT(*) FROM auth_user WHERE is_staff = false AND is_active = true")
            students = cursor.fetchone()[0] or 0
            
            return JsonResponse({
                'success': True,
                'schools': schools,
                'teachers': teachers,
                'students': students,
                'message': 'Metrics retrieved successfully'
            })
    except Exception as e:
        print(f"Error in test_metrics: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e),
            'schools': 0,
            'teachers': 0,
            'students': 0
        }, status=500)


# ========== PAPER LIBRARY VIEWS ==========

@login_required
def paper_library(request):
    """Browse papers organized by PaperSet"""
    from .models import PaperSet, Subject, SchoolSetting
    from .forms import PaperSetFilterForm
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    form = PaperSetFilterForm(request.GET)
    papers = PaperSet.objects.filter(is_active=True).select_related('subject').prefetch_related('resources')
    
    if request.GET.get('grade'):
        papers = papers.filter(grade=request.GET['grade'])
    if request.GET.get('subject'):
        papers = papers.filter(subject_id=request.GET['subject'])
    if request.GET.get('year'):
        papers = papers.filter(year=request.GET['year'])
    if request.GET.get('paper_type'):
        papers = papers.filter(paper_type=request.GET['paper_type'])
    
    q = request.GET.get('q')
    if q:
        papers = papers.filter(
            Q(title__icontains=q) | Q(grade__icontains=q) | Q(subject__name__icontains=q)
        )
    
    paginator = Paginator(papers, 20)
    page = request.GET.get('page', 1)
    papers_page = paginator.get_page(page)
    school = SchoolSetting.objects.first()
    
    context = {
        'papers': papers_page,
        'filter_form': form,
        'total_papers': papers.count(),
        'total_resources': PaperResource.objects.count(),
        'school': school,
    }
    return render(request, 'digitallibrary/paper_library.html', context)


@login_required
def paper_detail(request, pk):
    """View a single paper set with all its resources"""
    from .models import PaperSet, SchoolSetting
    
    paper = get_object_or_404(PaperSet.objects.prefetch_related('resources'), pk=pk)
    paper.view_count += 1
    paper.save(update_fields=['view_count'])
    
    resources_by_kind = {kind: None for kind, _ in PaperResource.KIND_CHOICES}
    for resource in paper.resources.all():
        resources_by_kind[resource.kind] = resource
    
    school = SchoolSetting.objects.first()
    
    context = {
        'paper': paper,
        'resources_by_kind': resources_by_kind,
        'school': school,
    }
    return render(request, 'digitallibrary/paper_detail.html', context)


@login_required
def download_paper_resource(request, resource_id):
    """Download a paper resource with tracking"""
    from .models import PaperResource, ActivityLog, SchoolSetting
    import os
    from django.http import HttpResponse
    
    resource = get_object_or_404(PaperResource, pk=resource_id)
    paper_set = resource.paper_set
    paper_set.download_count += 1
    paper_set.save(update_fields=['download_count'])
    
    ActivityLog.objects.create(
        user=request.user,
        action='download',
        description=f"Downloaded {resource.get_kind_display()} for {paper_set.title}"
    )
    
    if resource.file:
        response = HttpResponse(resource.file, content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(resource.file.name)}"'
        return response
    
    messages.error(request, "File not found")
    return redirect('digitallibrary:paper_detail', pk=paper_set.id)


@login_required
def upload_paper_resource(request):
    """Upload a new paper resource (teacher/admin only)"""
    from .forms import PaperResourceForm
    from .models import SchoolSetting
    
    if not can_upload(request.user):
        messages.error(request, "Access Denied. Only teachers and administrators can upload.")
        return redirect('digitallibrary:paper_library')
    
    school = SchoolSetting.objects.first()
    
    if request.method == 'POST':
        form = PaperResourceForm(request.POST, request.FILES)
        if form.is_valid():
            resource = form.save()
            messages.success(request, f"Resource '{resource.title}' uploaded successfully!")
            return redirect('digitallibrary:paper_detail', pk=resource.paper_set.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        paper_set_id = request.GET.get('paper_set')
        initial = {'paper_set': paper_set_id} if paper_set_id else {}
        form = PaperResourceForm(initial=initial)
    
    context = {
        'form': form,
        'title': 'Upload Paper Resource',
        'school': school,
    }
    return render(request, 'digitallibrary/upload_paper_resource.html', context)


@staff_member_required
def create_paper_set(request):
    """Create a new paper set (admin only)"""
    from .models import Subject, PaperSet, SchoolSetting
    
    school = SchoolSetting.objects.first()
    
    if request.method == 'POST':
        grade = request.POST.get('grade')
        subject_id = request.POST.get('subject')
        year = request.POST.get('year')
        term = request.POST.get('term')
        paper_type = request.POST.get('paper_type')
        exam_type = request.POST.get('exam_type')
        is_featured = request.POST.get('is_featured') == 'on'
        
        if not all([grade, year, paper_type]):
            messages.error(request, "Grade, Year, and Paper Type are required.")
        else:
            paper_set, created = PaperSet.objects.get_or_create(
                grade=grade,
                subject_id=subject_id if subject_id else None,
                year=year,
                term=term if term else None,
                paper_type=paper_type,
                defaults={'exam_type': exam_type, 'is_featured': is_featured, 'is_active': True}
            )
            if created:
                messages.success(request, f"Paper set '{paper_set.title}' created successfully!")
            else:
                messages.info(request, f"Paper set '{paper_set.title}' already exists.")
            return redirect('digitallibrary:paper_detail', pk=paper_set.id)
    
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    years = range(2000, 2027)
    
    context = {
        'subjects': subjects,
        'years': years,
        'paper_types': PaperSet.PAPER_TYPES,
        'exam_types': PaperSet.EXAM_TYPES,
        'school': school,
        'title': 'Create Paper Set',
    }
    return render(request, 'digitallibrary/create_paper_set.html', context)


# ========== TENANT SELECTOR ==========

def tenant_selector(request):
    """Page to select which school/tenant to work with"""
    from tenants.models import School
    from .models import SchoolSetting
    
    tenants = School.objects.all()
    school = SchoolSetting.objects.first()
    
    context = {
        'tenants': tenants,
        'school': school,
        'title': 'Select School',
    }
    return render(request, 'digitallibrary/tenant_selector.html', context)


# ========== CHECK RESULT MODEL ==========

def check_result_model(request):
    """Debug view to check what Result models are available"""
    from django.apps import apps
    
    result_models = []
    for model in apps.get_models():
        model_name = model.__name__.lower()
        if 'result' in model_name:
            result_models.append({
                'name': model.__name__,
                'fields': [f.name for f in model._meta.fields],
                'app': model._meta.app_label
            })
    
    context = {
        'result_models': result_models,
        'total_models': len(result_models)
    }
    
    return render(request, 'debug/models_check.html', context)


# ========== BULK DOWNLOAD ==========

@login_required
@require_http_methods(['POST'])
def bulk_download_student_packages(request):
    """Generate ZIP file with individual student reports and fee statements"""
    from .models import Student, Exam, SchoolSetting
    from datetime import datetime
    import io
    import zipfile
    
    class_id = request.POST.get('class_id', '')
    exam_id = request.POST.get('exam_id', '')
    term = request.POST.get('term', '')
    year = request.POST.get('year', datetime.now().year)
    include_fee_statement = request.POST.get('include_fee_statement', 'on')
    include_report_card = request.POST.get('include_report_card', 'on')
    
    students_qs = Student.objects.filter(is_active=True)
    if class_id:
        students_qs = students_qs.filter(current_class_id=class_id)
    
    exam = None
    if exam_id:
        try:
            exam = Exam.objects.get(id=exam_id)
        except Exam.DoesNotExist:
            pass
    
    school = SchoolSetting.objects.first()
    school_name = school.name if school else "School Name"
    school_logo = school.logo.path if school and school.logo else None
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for student in students_qs:
            student_folder = f"{student.admission_number}_{student.first_name}_{student.last_name}".replace(' ', '_')
            
            if include_report_card:
                report_card_pdf = generate_student_report_card(
                    student, exam, term, year, school_name, school_logo
                )
                zip_file.writestr(
                    f"{student_folder}/Report_Card_{student.admission_number}.pdf",
                    report_card_pdf
                )
            
            if include_fee_statement:
                fee_statement_pdf = generate_fee_statement(
                    student, term, year, school_name, school_logo
                )
                zip_file.writestr(
                    f"{student_folder}/Fee_Statement_{student.admission_number}.pdf",
                    fee_statement_pdf
                )
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    zip_buffer.seek(0)
    
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="student_packages_{timestamp}.zip"'
    
    return response
# ========== PUBLIC METRICS API ==========

@api_view(['GET'])
def public_metrics(request):
    """Public API endpoint for central dashboard metrics - REAL DATA from all tenant schools"""
    from tenants.models import School
    from django_tenants.utils import tenant_context
    from django.core.cache import cache
    from django.db.models import Sum, Q
    from django.utils import timezone
    from datetime import timedelta
    from .models import UserProfile, Student, Resource, PrintJob, ActivityLog, Subject
    
    cache_key = 'central_dashboard_metrics'
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return Response(cached_data)
    
    try:
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        schools = School.objects.all()
        total_schools = schools.count()
        schools_this_month = schools.filter(created_on__gte=month_start).count()
        
        total_teachers = 0
        teachers_this_month = 0
        total_resources = 0
        total_pdfs = 0
        total_views = 0
        downloads_today = 0
        views_today = 0
        total_print_jobs = 0
        prints_today = 0
        total_students = 0
        active_users_today = 0
        resources_by_subject = {}
        recent_uploads = []
        
        for school in schools:
            try:
                with tenant_context(school):
                    teacher_count = UserProfile.objects.filter(role='teacher', is_approved=True).count()
                    total_teachers += teacher_count
                    teachers_new = UserProfile.objects.filter(
                        role='teacher',
                        created_at__gte=month_start
                    ).count()
                    teachers_this_month += teachers_new
                    
                    resources = Resource.objects.all()
                    total_resources += resources.count()
                    total_pdfs += resources.filter(resource_type='PDF').count()
                    total_views += resources.aggregate(total=Sum('views'))['total'] or 0
                    
                    downloads_today += ActivityLog.objects.filter(
                        action='download',
                        timestamp__gte=today_start
                    ).count()
                    views_today += ActivityLog.objects.filter(
                        action='resource_view',
                        timestamp__gte=today_start
                    ).count()
                    
                    for subject in Subject.objects.filter(is_active=True):
                        count = resources.filter(subject=subject).count()
                        if count > 0:
                            subject_name = subject.name
                            resources_by_subject[subject_name] = resources_by_subject.get(subject_name, 0) + count
                    
                    recent = resources.order_by('-created_at')[:5]
                    for r in recent:
                        recent_uploads.append({
                            'title': r.title,
                            'school_name': school.name,
                            'created_at': r.created_at.isoformat()
                        })
                    
                    total_print_jobs += PrintJob.objects.count()
                    prints_today += PrintJob.objects.filter(
                        created_at__gte=today_start
                    ).count()
                    
                    total_students += Student.objects.filter(is_active=True).count()
                    
                    active_users_today += ActivityLog.objects.filter(
                        timestamp__gte=today_start
                    ).values('user').distinct().count()
                    
            except Exception as e:
                print(f"Error processing tenant {school.schema_name}: {e}")
                continue
        
        recent_uploads.sort(key=lambda x: x['created_at'], reverse=True)
        recent_uploads = recent_uploads[:10]
        
        resources_by_subject_list = [
            {'name': name, 'count': count, 'resource_count': count}
            for name, count in sorted(resources_by_subject.items(), key=lambda x: x[1], reverse=True)
        ]
        
        metrics = {
            'total_schools': total_schools,
            'schools_this_month': schools_this_month,
            'total_teachers': total_teachers,
            'teachers_this_month': teachers_this_month,
            'total_resources': total_resources,
            'total_pdfs': total_pdfs,
            'total_downloads': total_views,
            'total_views': total_views,
            'total_students': total_students,
            'total_print_jobs': total_print_jobs,
            'active_users_today': active_users_today,
            'downloads_today': downloads_today,
            'views_today': views_today,
            'prints_today': prints_today,
            'resources_by_subject': resources_by_subject_list,
            'recent_uploads': recent_uploads,
            'last_updated': now.isoformat(),
        }
        
        cache.set(cache_key, metrics, 300)
        return Response(metrics)
        
    except Exception as e:
        print(f"Error in public_metrics: {e}")
        import traceback
        traceback.print_exc()
        return Response({
            'error': str(e),
            'total_schools': 0,
            'schools_this_month': 0,
            'total_teachers': 0,
            'teachers_this_month': 0,
            'total_resources': 0,
            'total_pdfs': 0,
            'total_downloads': 0,
            'total_views': 0,
            'total_students': 0,
            'total_print_jobs': 0,
            'active_users_today': 0,
            'downloads_today': 0,
            'views_today': 0,
            'prints_today': 0,
            'resources_by_subject': [],
            'recent_uploads': [],
            'last_updated': timezone.now().isoformat(),
        }, status=200)


@api_view(['POST'])
def register_school_api(request):
    """API endpoint for school registration - sends email notification"""
    from django.conf import settings
    from django.core.mail import send_mail
    import re
    
    try:
        data = request.data
        
        school_name = data.get('school_name')
        admin_name = data.get('admin_name')
        email = data.get('email')
        phone = data.get('phone')
        location = data.get('location')
        teacher_count = data.get('teacher_count')
        student_count = data.get('student_count')
        
        if not all([school_name, admin_name, email, phone, location]):
            return Response({
                'success': False,
                'error': 'Please fill in all required fields: school_name, admin_name, email, phone, location'
            }, status=400)
        
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email):
            return Response({
                'success': False,
                'error': 'Please enter a valid email address'
            }, status=400)
        
        try:
            email_body = f"""
            New School Registration Request
            
            School Name: {school_name}
            Admin Name: {admin_name}
            Email: {email}
            Phone: {phone}
            Location: {location}
            Teachers: {teacher_count if teacher_count else 'Not specified'}
            Students: {student_count if student_count else 'Not specified'}
            
            Please follow up with this school to complete onboarding.
            
            ---
            This is an automated message from the School Library System.
            """
            
            send_mail(
                subject=f'New School Registration: {school_name}',
                message=email_body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=settings.ADMIN_EMAILS,
                fail_silently=False,
            )
            print(f"Email sent to {settings.ADMIN_EMAILS}")
        except Exception as e:
            print(f"Email error: {e}")
        
        return Response({
            'success': True,
            'message': 'School registration request received! We will contact you soon.',
            'data': {
                'school_name': school_name,
                'admin_name': admin_name,
                'email': email
            }
        })
        
    except Exception as e:
        print(f"Registration error: {e}")
        import traceback
        traceback.print_exc()
        return Response({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=500)


@api_view(['GET'])
def central_stats(request):
    """Simple stats for the central dashboard"""
    from tenants.models import School
    from django.core.cache import cache
    from django.utils import timezone
    
    cache_key = 'central_simple_stats'
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return Response(cached_data)
    
    try:
        schools = School.objects.all()
        total_schools = schools.count()
        
        stats = {
            'total_schools': total_schools,
            'status': 'active',
            'timestamp': timezone.now().isoformat()
        }
        
        cache.set(cache_key, stats, 300)
        return Response(stats)
        
    except Exception as e:
        return Response({
            'total_schools': 0,
            'status': 'error',
            'error': str(e),
            'timestamp': timezone.now().isoformat()
        }, status=200)


@api_view(['GET'])
def get_schools_list(request):
    """Get list of all registered schools (tenants)"""
    from tenants.models import School
    
    try:
        schools = School.objects.all().values('id', 'name', 'schema_name', 'created_on')
        return Response(list(schools))
    except Exception as e:
        return Response({
            'error': str(e),
            'schools': []
        }, status=200)


@api_view(['GET'])
def get_school_stats(request, school_id):
    """Get statistics for a specific school (tenant)"""
    from tenants.models import School
    from django_tenants.utils import tenant_context
    from django.db.models import Sum
    from .models import Resource, UserProfile, Student, PrintJob, Announcement, FeePayment, Exam, StudentResult
    
    try:
        school = School.objects.get(id=school_id)
        
        with tenant_context(school):
            stats = {
                'school_id': school.id,
                'school_name': school.name,
                'schema_name': school.schema_name,
                'total_resources': Resource.objects.count(),
                'total_teachers': UserProfile.objects.filter(role='teacher', is_approved=True).count(),
                'total_students': Student.objects.filter(is_active=True).count(),
                'total_print_jobs': PrintJob.objects.count(),
                'total_views': Resource.objects.aggregate(total=Sum('views'))['total'] or 0,
                'total_announcements': Announcement.objects.count(),
                'total_fee_payments': FeePayment.objects.count(),
                'total_exams': Exam.objects.count(),
                'total_results': StudentResult.objects.count(),
            }
        
        return Response(stats)
    except School.DoesNotExist:
        return Response({'error': 'School not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)
# ========== FEES MANAGEMENT VIEWS ==========

from decimal import Decimal

@login_required
def fees_dashboard(request):
    """Main fees dashboard with statistics - Accessible by Admin, Principal, and Bursar"""
    from .models import Student, FeeStructure, FeePayment, FeeBalance, Class, SchoolSetting
    from decimal import Decimal
    from django.db.models import Sum
    
    user_role = request.user.profile.role
    if user_role not in ['admin', 'principal', 'bursar']:
        messages.error(request, f"Access Denied. {user_role.capitalize()}s cannot access the fees dashboard.")
        return redirect('digitallibrary:home')
    
    current_year = request.GET.get('year', str(timezone.now().year))
    current_term = request.GET.get('term', '1')
    
    try:
        current_term = int(current_term)
    except ValueError:
        current_term = 1
    
    students = Student.objects.filter(is_active=True).select_related('current_class')
    total_students = students.count()
    
    fee_structures = FeeStructure.objects.filter(
        academic_year=current_year,
        term=current_term
    ).select_related('student_class')
    
    class_fee_map = {}
    for fs in fee_structures:
        if fs.student_class:
            class_fee_map[fs.student_class.id] = Decimal(str(fs.total_fees))
    
    total_expected_all = Decimal('0')
    students_with_fee_structure = 0
    students_without_fee_structure = 0
    
    for student in students:
        if student.current_class:
            class_id = student.current_class.id
            if class_id in class_fee_map:
                total_expected_all += class_fee_map[class_id]
                students_with_fee_structure += 1
            else:
                students_without_fee_structure += 1
        else:
            students_without_fee_structure += 1
    
    total_paid_all = FeePayment.objects.filter(
        academic_year=current_year,
        term=current_term
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    if not isinstance(total_paid_all, Decimal):
        total_paid_all = Decimal(str(total_paid_all))
    
    total_balance_all = total_expected_all - total_paid_all
    
    balances = FeeBalance.objects.filter(
        academic_year=current_year,
        term=current_term
    ).select_related('student')
    
    paid_count = balances.filter(status='PAID').count()
    partial_count = balances.filter(status='PARTIAL').count()
    defaulting_count = balances.filter(status='DEFAULTING').count()
    overpaid_count = balances.filter(status='OVERPAID').count()
    
    if balances.count() == 0 and total_students > 0:
        defaulting_count = total_students
        paid_count = 0
        partial_count = 0
    
    if total_expected_all > 0:
        collection_percentage = float(total_paid_all / total_expected_all * 100)
    else:
        collection_percentage = 0
    
    recent_payments = FeePayment.objects.filter(
        academic_year=current_year,
        term=current_term
    ).order_by('-payment_date')[:10]
    
    defaulters = []
    for balance in balances.filter(balance__gt=0).exclude(status='OVERPAID'):
        defaulters.append({
            'student': balance.student,
            'balance': balance.balance,
            'total_expected': balance.total_expected,
            'total_paid': balance.total_paid,
            'status': balance.status
        })
    
    students_with_balance = balances.values('student').distinct().count()
    if students_with_balance < total_students:
        for student in students:
            if not balances.filter(student=student).exists():
                expected = Decimal('0')
                if student.current_class and student.current_class.id in class_fee_map:
                    expected = class_fee_map[student.current_class.id]
                if expected > 0:
                    defaulters.append({
                        'student': student,
                        'balance': expected,
                        'total_expected': expected,
                        'total_paid': Decimal('0'),
                        'status': 'DEFAULTING'
                    })
    
    seen = set()
    unique_defaulters = []
    for d in defaulters:
        if d['student'].id not in seen:
            seen.add(d['student'].id)
            unique_defaulters.append(d)
    unique_defaulters.sort(key=lambda x: x['balance'], reverse=True)
    defaulters = unique_defaulters[:20]
    
    class_breakdown = []
    for class_obj in Class.objects.all().order_by('name'):
        student_count = students.filter(current_class=class_obj).count()
        if student_count > 0:
            if class_obj.id in class_fee_map:
                fee_amount = class_fee_map[class_obj.id]
                total_expected_for_class = student_count * fee_amount
                total_paid_for_class = FeePayment.objects.filter(
                    academic_year=current_year,
                    term=current_term,
                    student__current_class=class_obj
                ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
                
                if not isinstance(total_paid_for_class, Decimal):
                    total_paid_for_class = Decimal(str(total_paid_for_class))
                
                if total_expected_for_class > 0:
                    collection_pct = float(total_paid_for_class / total_expected_for_class * 100)
                else:
                    collection_pct = 0
                
                class_breakdown.append({
                    'name': class_obj.name,
                    'students': student_count,
                    'fee_per_student': float(fee_amount),
                    'total_expected': float(total_expected_for_class),
                    'total_paid': float(total_paid_for_class),
                    'balance': float(total_expected_for_class - total_paid_for_class),
                    'collection_percentage': round(collection_pct, 1)
                })
            else:
                class_breakdown.append({
                    'name': class_obj.name,
                    'students': student_count,
                    'fee_per_student': 0,
                    'total_expected': 0,
                    'total_paid': 0,
                    'balance': 0,
                    'collection_percentage': 0,
                    'no_fee_structure': True
                })
    
    available_years = FeeStructure.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
    if not available_years:
        available_years = [current_year]
    
    school = SchoolSetting.objects.first()
    
    context = {
        'current_year': current_year,
        'current_term': current_term,
        'total_expected': float(total_expected_all),
        'total_paid': float(total_paid_all),
        'total_balance': float(total_balance_all),
        'collection_percentage': round(collection_percentage, 1),
        'paid_count': paid_count,
        'partial_count': partial_count,
        'defaulting_count': defaulting_count,
        'overpaid_count': overpaid_count,
        'recent_payments': recent_payments,
        'defaulters': defaulters,
        'available_years': available_years,
        'class_breakdown': class_breakdown,
        'total_students': total_students,
        'school': school,
        'school_name': school.name if school else 'School Name',
        'school_logo': school.logo.url if school and school.logo else None,
    }
    return render(request, 'fees/dashboard.html', context)


@tenant_app_view
def fee_structure_list(request):
    """List all fee structures"""
    from .models import FeeStructure, SchoolSetting
    
    fee_structures = FeeStructure.objects.all().select_related('student_class').order_by('-academic_year', 'student_class__name')
    
    year = request.GET.get('year')
    if year:
        fee_structures = fee_structures.filter(academic_year=year)
    
    term = request.GET.get('term')
    if term:
        try:
            fee_structures = fee_structures.filter(term=int(term))
        except ValueError:
            pass
    
    for fs in fee_structures:
        if fs.pk:
            current_total = fs.calculate_total()
            if fs.total_fees != current_total:
                fs.total_fees = current_total
                fs.save(update_fields=['total_fees'])
    
    available_years = FeeStructure.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
    school = SchoolSetting.objects.first()
    
    context = {
        'fee_structures': fee_structures,
        'current_year': request.GET.get('year', str(timezone.now().year)),
        'current_term': request.GET.get('term', ''),
        'available_years': available_years,
        'school': school,
    }
    return render(request, 'fees/fee_structure_list.html', context)


@login_required
def get_subject_students(request, exam_id):
    """AJAX endpoint to get students for a specific subject under an exam"""
    from .models import Exam, Subject, Student
    
    exam = get_object_or_404(Exam, id=exam_id)
    subject_id = request.GET.get('subject_id')
    
    if not subject_id:
        return JsonResponse({'success': False, 'error': 'No subject selected'})
    
    subject = get_object_or_404(Subject, id=subject_id)
    
    if exam.student_class:
        students = exam.student_class.students.all()
    else:
        students = Student.objects.all()
    
    existing_results = StudentResult.objects.filter(
        exam=exam, 
        subject=subject,
        student__in=students
    ).select_related('student')
    
    results_dict = {result.student_id: result for result in existing_results}
    
    students_data = []
    for student in students:
        result = results_dict.get(student.id)
        students_data.append({
            'id': student.id,
            'admission_number': student.admission_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'score': result.score if result else None,
            'grade': result.grade if result else None
        })
    
    return JsonResponse({
        'success': True,
        'subject': {
            'id': subject.id,
            'name': subject.name
        },
        'students': students_data,
        'max_score': exam.max_score,
        'existing_count': len(existing_results),
        'total_count': students.count()
    })


@tenant_app_view
def student_list(request):
    """List all students"""
    from .models import Student, Class, SchoolSetting
    from .forms import StudentSearchForm
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    students = Student.objects.filter(is_active=True).select_related('current_class')
    
    search_form = StudentSearchForm(request.GET)
    if search_form.is_valid():
        query = search_form.cleaned_data.get('query')
        class_filter = search_form.cleaned_data.get('class_filter')
        gender_filter = search_form.cleaned_data.get('gender_filter')
        status_filter = search_form.cleaned_data.get('status_filter')
        
        if query:
            students = students.filter(
                Q(admission_number__icontains=query) |
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(parent_phone__icontains=query) |
                Q(upi_number__icontains=query)
            )
        
        if class_filter:
            students = students.filter(current_class=class_filter)
        
        if gender_filter:
            students = students.filter(gender=gender_filter)
        
        if status_filter == 'active':
            students = students.filter(is_active=True)
        elif status_filter == 'inactive':
            students = students.filter(is_active=False)
    
    paginator = Paginator(students, 20)
    page_number = request.GET.get('page')
    students_page = paginator.get_page(page_number)
    
    classes = Class.objects.all().order_by('name')
    students_with_phone = Student.objects.filter(is_active=True, parent_phone__isnull=False).exclude(parent_phone='').count()
    
    male_count = Student.objects.filter(is_active=True, gender='M').count()
    female_count = Student.objects.filter(is_active=True, gender='F').count()
    other_count = Student.objects.filter(is_active=True, gender='O').count()
    total_students = Student.objects.filter(is_active=True).count()
    active_students = students.count()
    school = SchoolSetting.objects.first()
    
    context = {
        'students': students_page,
        'search_form': search_form,
        'classes': classes,
        'students_with_phone': students_with_phone,
        'male_count': male_count,
        'female_count': female_count,
        'other_count': other_count,
        'total_students': total_students,
        'active_students': active_students,
        'school': school,
    }
    return render(request, 'fees/student_list.html', context)


def student_detail(request, pk):
    """View student details with fee information"""
    from .models import Student, FeeBalance, FeePayment, FeeStructure, Class, SchoolSetting
    from django.db.models import Sum
    
    student = get_object_or_404(Student, pk=pk)
    current_year = request.GET.get('year', str(timezone.now().year))
    current_term = int(request.GET.get('term', '1'))
    
    fee_balances = FeeBalance.objects.filter(
        student=student,
        academic_year=current_year
    ).order_by('term')
    
    payments = FeePayment.objects.filter(student=student).order_by('-payment_date')
    total_paid = payments.aggregate(total=Sum('amount'))['total'] or 0
    
    class_id = student.current_class.id if student.current_class else None
    
    total_expected = FeeStructure.objects.filter(
        academic_year=current_year,
        term=current_term,
        student_class_id=class_id
    ).aggregate(total=Sum('total_fees'))['total'] or 0
    
    current_balance = total_expected - total_paid
    classes = Class.objects.all().order_by('name')
    
    available_years = FeePayment.objects.filter(student=student).values_list('academic_year', flat=True).distinct().order_by('-academic_year')
    if not available_years:
        available_years = [current_year]
    school = SchoolSetting.objects.first()
    
    context = {
        'student': student,
        'fee_balances': fee_balances,
        'payments': payments,
        'total_paid': total_paid,
        'total_expected': total_expected,
        'current_balance': current_balance,
        'current_year': current_year,
        'current_term': current_term,
        'classes': classes,
        'available_years': available_years,
        'school': school,
    }
    return render(request, 'fees/student_detail.html', context)
# ========== FEES MANAGEMENT VIEWS (continued) ==========

@tenant_app_view
def defaulter_list(request):
    """List all students with outstanding balances"""
    from .models import FeeBalance, Class, SchoolSetting
    from django.db import models
    
    current_year = request.GET.get('year', str(timezone.now().year))
    current_term = request.GET.get('term', '1')
    
    balances = FeeBalance.objects.filter(
        academic_year=current_year,
        term=current_term,
        balance__gt=0
    ).exclude(status='OVERPAID').select_related('student')
    
    class_filter = request.GET.get('class')
    if class_filter:
        balances = balances.filter(student__current_class_id=class_filter)
    
    balances = balances.order_by('-balance')
    total_due = balances.aggregate(total=models.Sum('balance'))['total'] or 0
    school = SchoolSetting.objects.first()
    
    context = {
        'balances': balances,
        'total_due': total_due,
        'current_year': current_year,
        'current_term': current_term,
        'school': school,
    }
    return render(request, 'fees/defaulter_list.html', context)


@tenant_app_view
def collection_report(request):
    """Collection report by date range"""
    from .models import FeePayment, SchoolSetting
    from django.db import models
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    payments = FeePayment.objects.all()
    
    if start_date:
        payments = payments.filter(payment_date__gte=start_date)
    if end_date:
        payments = payments.filter(payment_date__lte=end_date)
    
    payments = payments.order_by('-payment_date')
    total_collected = payments.aggregate(total=models.Sum('amount'))['total'] or 0
    
    by_method = {}
    for method, label in FeePayment.PAYMENT_METHODS:
        total = payments.filter(payment_method=method).aggregate(total=models.Sum('amount'))['total'] or 0
        by_method[label] = total
    
    school = SchoolSetting.objects.first()
    
    context = {
        'payments': payments,
        'total_collected': total_collected,
        'by_method': by_method,
        'start_date': start_date,
        'end_date': end_date,
        'school': school,
    }
    return render(request, 'fees/collection_report.html', context)


def fee_structure_create(request):
    """Create new fee structure with dynamic components"""
    from .models import FeeStructure, Class as ClassModel, FeeComponent, SchoolSetting
    from .forms import FeeStructureForm
    
    if request.method == 'POST':
        form = FeeStructureForm(request.POST)
        if form.is_valid():
            fee_structure = form.save()
            
            for key, value in request.POST.items():
                if key.startswith('new_component_name_'):
                    index = key.replace('new_component_name_', '')
                    name = value.strip()
                    amount = request.POST.get(f'new_component_amount_{index}', '')
                    is_optional = request.POST.get(f'new_component_optional_{index}') == 'on'
                    description = request.POST.get(f'new_component_description_{index}', '')
                    
                    if name and amount:
                        try:
                            FeeComponent.objects.create(
                                fee_structure=fee_structure,
                                name=name,
                                amount=float(amount),
                                is_optional=is_optional,
                                description=description
                            )
                        except ValueError:
                            pass
            
            total = fee_structure.calculate_total()
            fee_structure.total_fees = total
            fee_structure.save(update_fields=['total_fees'])
            
            messages.success(request, f'Fee structure created successfully! Total: KES {total:,.2f}')
            return redirect('digitallibrary:fee_structure_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = FeeStructureForm()
    
    try:
        classes = ClassModel.objects.filter(is_active=True).order_by('name')
    except:
        classes = []
    
    school = SchoolSetting.objects.first()
    
    context = {
        'form': form,
        'title': 'Create Fee Structure',
        'is_edit': False,
        'classes': classes,
        'school': school,
    }
    return render(request, 'fees/fee_structure_form.html', context)


def fee_structure_edit(request, pk):
    """Edit fee structure with dynamic components"""
    from .models import FeeStructure, Class as ClassModel, FeeComponent, SchoolSetting
    from .forms import FeeStructureForm
    
    fee_structure = get_object_or_404(FeeStructure, pk=pk)
    fee_components = fee_structure.custom_fees.all()
    
    if request.method == 'POST':
        form = FeeStructureForm(request.POST, instance=fee_structure)
        if form.is_valid():
            fee_structure = form.save()
            
            kept_component_ids = []
            
            for key, value in request.POST.items():
                if key.startswith('component_id_'):
                    component_id = int(value)
                    kept_component_ids.append(component_id)
                    
                    name = request.POST.get(f'component_name_{component_id}', '')
                    amount = request.POST.get(f'component_amount_{component_id}', '')
                    is_optional = request.POST.get(f'component_optional_{component_id}') == 'on'
                    description = request.POST.get(f'component_description_{component_id}', '')
                    
                    if name and amount:
                        try:
                            FeeComponent.objects.update_or_create(
                                id=component_id,
                                defaults={
                                    'fee_structure': fee_structure,
                                    'name': name,
                                    'amount': float(amount),
                                    'is_optional': is_optional,
                                    'description': description
                                }
                            )
                        except ValueError:
                            FeeComponent.objects.filter(id=component_id).delete()
                    else:
                        FeeComponent.objects.filter(id=component_id).delete()
            
            fee_structure.custom_fees.exclude(id__in=kept_component_ids).delete()
            
            for key, value in request.POST.items():
                if key.startswith('new_component_name_'):
                    index = key.replace('new_component_name_', '')
                    name = value.strip()
                    amount = request.POST.get(f'new_component_amount_{index}', '')
                    is_optional = request.POST.get(f'new_component_optional_{index}') == 'on'
                    description = request.POST.get(f'new_component_description_{index}', '')
                    
                    if name and amount:
                        try:
                            FeeComponent.objects.create(
                                fee_structure=fee_structure,
                                name=name,
                                amount=float(amount),
                                is_optional=is_optional,
                                description=description
                            )
                        except ValueError:
                            pass
            
            total = fee_structure.calculate_total()
            fee_structure.total_fees = total
            fee_structure.save(update_fields=['total_fees'])
            
            messages.success(request, f'Fee structure updated successfully! Total: KES {total:,.2f}')
            return redirect('digitallibrary:fee_structure_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = FeeStructureForm(instance=fee_structure)
    
    components_data = []
    for component in fee_components:
        components_data.append({
            'id': component.id,
            'name': component.name,
            'amount': str(component.amount),
            'is_optional': component.is_optional,
            'description': component.description or '',
        })
    
    try:
        classes = ClassModel.objects.filter(is_active=True).order_by('name')
    except:
        classes = []
    
    school = SchoolSetting.objects.first()
    
    context = {
        'form': form,
        'title': f'Edit Fee Structure - {fee_structure.student_class.name if fee_structure.student_class else "N/A"}',
        'fee_structure': fee_structure,
        'fee_components': components_data,
        'is_edit': True,
        'classes': classes,
        'school': school,
    }
    return render(request, 'fees/fee_structure_form.html', context)


def fee_structure_delete(request, pk):
    """Delete a fee structure"""
    from .models import FeeStructure, SchoolSetting
    
    fee_structure = get_object_or_404(FeeStructure, pk=pk)
    school = SchoolSetting.objects.first()
    
    if request.method == 'POST':
        class_name = fee_structure.student_class.name if fee_structure.student_class else 'N/A'
        term = fee_structure.term
        year = fee_structure.academic_year
        fee_structure.delete()
        messages.success(request, f'Fee structure for {class_name} - Term {term} {year} deleted successfully!')
        return redirect('digitallibrary:fee_structure_list')
    
    context = {
        'fee_structure': fee_structure,
        'school': school,
    }
    return render(request, 'fees/fee_structure_confirm_delete.html', context)


def fee_structure_delete_component(request, pk):
    """Delete a fee component via AJAX"""
    from .models import FeeComponent
    
    if request.method == 'POST':
        try:
            component = get_object_or_404(FeeComponent, pk=pk)
            component.delete()
            return JsonResponse({'success': True, 'message': 'Component deleted successfully'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@tenant_app_view
def payment_record(request):
    """Record a new payment"""
    from .models import Student, FeePayment, SchoolSetting
    from .forms import FeePaymentForm
    
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        if student_id:
            post_data = request.POST.copy()
            post_data['student'] = student_id
            form = FeePaymentForm(post_data, request.FILES)
        else:
            form = FeePaymentForm(request.POST, request.FILES)
            
        if form.is_valid():
            payment = form.save(commit=False)
            payment.recorded_by = request.user
            
            if not payment.receipt_number:
                payment.receipt_number = generate_receipt_number()
            
            payment.save()
            update_fee_balance_after_payment(payment)
            
            messages.success(
                request, 
                f'Payment of KES {payment.amount:,.2f} recorded for {payment.student.get_full_name()}. '
                f'Receipt: {payment.receipt_number}'
            )
            return redirect('digitallibrary:student_detail', pk=payment.student.pk)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = FeePaymentForm()
    
    recent_payments = FeePayment.objects.all().order_by('-payment_date')[:10]
    all_students = Student.objects.filter(is_active=True).select_related('current_class').order_by('first_name', 'last_name')
    school = SchoolSetting.objects.first()
    
    context = {
        'form': form,
        'recent_payments': recent_payments,
        'all_students': all_students,
        'title': 'Record Payment',
        'school': school,
    }
    return render(request, 'fees/payment_record.html', context)


def payment_receipt(request, pk):
    """View and print receipt for a payment"""
    from .models import FeePayment, SchoolSetting
    
    payment = get_object_or_404(FeePayment, pk=pk)
    school = SchoolSetting.objects.first()
    
    if request.user.profile.role not in ['admin', 'principal'] and payment.recorded_by != request.user:
        messages.error(request, "Access Denied.")
        return redirect('digitallibrary:home')
    
    context = {
        'payment': payment,
        'school': school,
        'title': 'Payment Receipt',
    }
    return render(request, 'fees/payment_receipt.html', context)


def export_defaulters_csv(request):
    """Export defaulters list to CSV"""
    import csv
    from .models import FeeBalance, SchoolSetting
    from django.http import HttpResponse
    from django.db import models
    
    current_year = request.GET.get('year', str(timezone.now().year))
    current_term = request.GET.get('term', '1')
    
    balances = FeeBalance.objects.filter(
        academic_year=current_year,
        term=current_term,
        balance__gt=0
    ).exclude(status='OVERPAID').select_related('student')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="defaulters_{current_year}_term{current_term}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Admission No', 'Student Name', 'Class', 'Parent Name', 'Parent Phone', 'Balance'])
    
    for balance in balances:
        writer.writerow([
            balance.student.admission_number,
            balance.student.get_full_name(),
            balance.student.current_class.name if balance.student.current_class else 'N/A',
            balance.student.parent_name,
            balance.student.parent_phone,
            f"KES {balance.balance:,.2f}"
        ])
    
    return response


def export_fees_csv(request):
    """Export fee data to CSV"""
    import csv
    from .models import FeeBalance, SchoolSetting
    from django.http import HttpResponse
    
    current_year = request.GET.get('year', str(timezone.now().year))
    current_term = request.GET.get('term', '1')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="fee_report_{current_year}_term{current_term}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Admission Number', 'Student Name', 'Class', 'Total Expected', 'Total Paid', 'Balance', 'Status'])
    
    balances = FeeBalance.objects.filter(
        academic_year=current_year,
        term=current_term
    ).select_related('student')
    
    for balance in balances:
        writer.writerow([
            balance.student.admission_number,
            balance.student.get_full_name(),
            balance.student.current_class.name if balance.student.current_class else 'N/A',
            f"KES {balance.total_expected:,.2f}",
            f"KES {balance.total_paid:,.2f}",
            f"KES {balance.balance:,.2f}",
            balance.status
        ])
    
    return response
# ========== STUDENT BULK UPLOAD VIEW ==========

import pandas as pd
from django.core.validators import ValidationError

def student_bulk_upload(request):
    """Bulk upload students via Excel/CSV"""
    from .forms import BulkStudentUploadForm
    from .models import Student, Class
    from django.core.validators import ValidationError
    
    if request.method == 'POST':
        form = BulkStudentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            # Read file based on extension
            ext = excel_file.name.split('.')[-1].lower()
            try:
                if ext == 'csv':
                    df = pd.read_csv(excel_file)
                else:
                    df = pd.read_excel(excel_file)
            except Exception as e:
                messages.error(request, f'Error reading file: {str(e)}')
                return redirect('digitallibrary:student_bulk_upload')
            
            # Normalize columns
            df.columns = df.columns.str.strip().str.lower()
            
            # Check for required columns
            required_fields = ['first name', 'last name', 'admission number']
            missing_fields = [f for f in required_fields if f not in df.columns]
            if missing_fields:
                messages.error(request, f'Missing required columns: {", ".join(missing_fields)}')
                return redirect('digitallibrary:student_bulk_upload')
            
            success_count = 0
            error_count = 0
            errors = []
            new_classes_created = set()
            
            for index, row in df.iterrows():
                try:
                    # Skip empty rows
                    if pd.isna(row.get('first name', '')) and pd.isna(row.get('last name', '')):
                        continue
                    
                    # Get or create class
                    class_obj = None
                    class_name = row.get('class name', '')
                    if pd.notna(class_name) and str(class_name).strip():
                        class_name = str(class_name).strip()
                        class_obj, created = Class.objects.get_or_create(
                            name__iexact=class_name,
                            defaults={'name': class_name}
                        )
                        if created:
                            new_classes_created.add(class_name)
                    
                    # Get gender value
                    gender_map = {'MALE': 'M', 'M': 'M', 'FEMALE': 'F', 'F': 'F', 
                                  'OTHER': 'O', 'O': 'O'}
                    gender_raw = str(row.get('gender', 'N')).upper().strip()
                    gender = gender_map.get(gender_raw, 'N')
                    
                    # Get admission year
                    admission_year = 2026
                    year_value = row.get('admission year', '')
                    if pd.notna(year_value):
                        try:
                            year_str = str(year_value).strip()
                            admission_year = int(float(year_str)) if year_str else 2026
                        except (ValueError, TypeError):
                            admission_year = 2026
                    
                    # Check if student already exists
                    admission_number = str(row.get('admission number', '')).strip()
                    if not admission_number:
                        errors.append(f'Row {index + 2}: Admission number is required')
                        error_count += 1
                        continue
                        
                    if Student.objects.filter(admission_number=admission_number).exists():
                        errors.append(f'Row {index + 2}: Student with admission number {admission_number} already exists')
                        error_count += 1
                        continue
                    
                    # Get first and last name
                    first_name = str(row.get('first name', '')).strip()
                    last_name = str(row.get('last name', '')).strip()
                    
                    if not first_name or not last_name:
                        errors.append(f'Row {index + 2}: First name and last name are required')
                        error_count += 1
                        continue
                    
                    # Create student
                    student = Student(
                        first_name=first_name,
                        last_name=last_name,
                        admission_number=admission_number,
                        upi_number=str(row.get('upi number', '')).strip() if pd.notna(row.get('upi number', '')) else '',
                        middle_name=str(row.get('middle name', '')).strip() if pd.notna(row.get('middle name', '')) else '',
                        gender=gender,
                        admission_year=admission_year,
                        current_class=class_obj,
                        parent_name=str(row.get('parent name', '')).strip() if pd.notna(row.get('parent name', '')) else '',
                        parent_email=str(row.get('parent email', '')).strip() if pd.notna(row.get('parent email', '')) else '',
                        parent_phone=str(row.get('parent phone', '')).strip() if pd.notna(row.get('parent phone', '')) else '',
                        parent_alternative_phone=str(row.get('alternative phone', '')).strip() if pd.notna(row.get('alternative phone', '')) else '',
                        physical_address=str(row.get('physical address', '')).strip() if pd.notna(row.get('physical address', '')) else '',
                        is_active=True
                    )
                    
                    # Validate and save
                    try:
                        student.full_clean()
                        student.save()
                        success_count += 1
                    except ValidationError as e:
                        error_msg = ', '.join(e.messages)
                        errors.append(f'Row {index + 2}: {error_msg}')
                        error_count += 1
                        
                except Exception as e:
                    error_text = str(e)
                    errors.append(f'Row {index + 2}: {error_text}')
                    error_count += 1
            
            # Summary message
            summary = f'Successfully imported {success_count} students. Failed: {error_count}'
            if new_classes_created:
                summary += f' | Created classes: {", ".join(new_classes_created)}'
            
            if success_count > 0:
                messages.success(request, summary)
            if errors:
                error_preview = errors[:5]
                for error in error_preview:
                    messages.warning(request, error)
                if len(errors) > 5:
                    messages.info(request, f'And {len(errors) - 5} more errors...')
            
            return redirect('digitallibrary:student_list')
    else:
        form = BulkStudentUploadForm()
    
    return render(request, 'digitallibrary/student_bulk_upload.html', {
        'form': form,
        'title': 'Bulk Upload Students'
    })


# ========== STUDENT EDIT VIEW (if missing) ==========

@login_required
@user_passes_test(lambda u: u.is_staff or u.role == 'admin')
def student_edit(request, pk):
    """Edit an existing student"""
    from .forms import StudentForm
    from .models import Class, Subject
    
    student = get_object_or_404(Student, pk=pk)
    
    print("\n" + "=" * 80)
    print("STUDENT EDIT VIEW - START")
    print(f"Editing student: {student.first_name} {student.last_name} (ID: {student.id})")
    print("=" * 80)
    
    if request.method == 'POST':
        print("\n📝 REQUEST METHOD: POST")
        
        elective_subjects = request.POST.getlist('elective_subjects')
        print(f"\n📚 ELECTIVE SUBJECTS FROM POST: {elective_subjects}")
        
        pathway_value = request.POST.get('pathway', '')
        print(f"   Pathway from POST: {pathway_value}")
        
        form = StudentForm(request.POST, request.FILES, instance=student)
        
        if form.is_valid():
            student = form.save(commit=False)
            
            if pathway_value:
                student.pathway = pathway_value
            
            # Handle class assignment
            new_class_name = form.cleaned_data.get('new_class')
            current_class_id = form.cleaned_data.get('current_class')
            
            if new_class_name:
                class_obj, created = Class.objects.get_or_create(
                    name=new_class_name.title()
                )
                student.current_class = class_obj
            elif current_class_id:
                try:
                    if isinstance(current_class_id, Class):
                        student.current_class = current_class_id
                    else:
                        student.current_class = Class.objects.get(id=current_class_id)
                except (Class.DoesNotExist, ValueError, TypeError):
                    messages.error(request, 'Selected class does not exist.')
                    return render(request, 'digitallibrary/student_form.html', {'form': form, 'student': student})
            
            student.save()
            
            # Handle subjects
            student.subjects.clear()
            
            for subject_name in elective_subjects:
                if subject_name and subject_name.strip():
                    subject_obj, _ = Subject.objects.get_or_create(
                        name=subject_name.strip(),
                        defaults={'is_active': True}
                    )
                    student.subjects.add(subject_obj)
            
            # Add compulsory subjects
            if student.pathway:
                compulsory_mapping = {
                    'arts_sports': ['English', 'Kiswahili/KSL', 'Core Mathematics', 'Community Service Learning (CSL)'],
                    'social_sciences': ['English', 'Kiswahili/KSL', 'Core Mathematics', 'Community Service Learning (CSL)'],
                    'stem': ['English', 'Kiswahili/KSL', 'Core Mathematics', 'Community Service Learning (CSL)'],
                }
                
                for subject_name in compulsory_mapping.get(student.pathway, []):
                    subject_obj, _ = Subject.objects.get_or_create(
                        name=subject_name,
                        defaults={'is_compulsory': True, 'category': 'compulsory'}
                    )
                    student.subjects.add(subject_obj)
            
            messages.success(request, f'Student {student.first_name} {student.last_name} updated successfully!')
            return redirect('digitallibrary:student_detail', pk=student.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StudentForm(instance=student)
        current_subjects = list(student.subjects.values_list('name', flat=True))
    
    classes = Class.objects.all().order_by('name')
    pathway_value = student.pathway if student.pathway else ''
    
    context = {
        'form': form,
        'classes': classes,
        'student': student,
        'current_subjects': current_subjects if 'current_subjects' in locals() else [],
        'pathway_value': pathway_value,
        'title': 'Edit Student',
        'action': 'Edit',
    }
    
    return render(request, 'digitallibrary/student_form.html', context)
# ========== STUDENT CREATE VIEW ==========

@login_required
@user_passes_test(lambda u: u.is_staff or u.role == 'admin')
def student_create(request):
    """Create a new student with class assignment and subjects"""
    from .forms import StudentForm
    from .models import Class, Subject
    
    print("\n" + "=" * 80)
    print("STUDENT CREATE VIEW - START")
    print("=" * 80)
    
    if request.method == 'POST':
        print("\n📝 REQUEST METHOD: POST")
        print("-" * 40)
        
        # Debug: Print all POST data
        print("\n📤 POST DATA RECEIVED:")
        for key, value in request.POST.items():
            if key == 'csrfmiddlewaretoken':
                continue
            print(f"   {key}: {value}")
        
        elective_subjects = request.POST.getlist('elective_subjects')
        print(f"\n📚 ELECTIVE SUBJECTS: {elective_subjects}")
        
        subjects_hidden = request.POST.get('elective_subjects_hidden', '')
        if subjects_hidden:
            print(f"   📚 ELECTIVE SUBJECTS FROM HIDDEN: {subjects_hidden}")
        
        form = StudentForm(request.POST, request.FILES)
        
        print("\n✅ FORM VALIDATION:")
        if form.is_valid():
            print("   ✓ Form is valid")
            
            print("\n📊 CLEANED DATA:")
            for field, value in form.cleaned_data.items():
                if field not in ['csrfmiddlewaretoken']:
                    print(f"   {field}: {value}")
            
            student = form.save(commit=False)
            print(f"\n👨‍🎓 Student object created:")
            print(f"   First Name: {student.first_name}")
            print(f"   Last Name: {student.last_name}")
            print(f"   Admission: {student.admission_number}")
            
            # Get pathway from form data
            pathway_value = request.POST.get('pathway', '')
            if pathway_value:
                student.pathway = pathway_value
                print(f"   Pathway set to: {student.pathway}")
            else:
                print(f"   ⚠ No pathway selected")
            
            # Handle class assignment
            new_class_name = form.cleaned_data.get('new_class')
            current_class_id = form.cleaned_data.get('current_class')
            print(f"\n🏫 CLASS ASSIGNMENT:")
            print(f"   New class name: {new_class_name}")
            print(f"   Current class ID: {current_class_id}")
            
            if new_class_name:
                print(f"   ✓ Creating new class: {new_class_name}")
                class_obj, created = Class.objects.get_or_create(
                    name=new_class_name.title()
                )
                student.current_class = class_obj
                if created:
                    messages.info(request, f'New class "{new_class_name}" has been created.')
                    print(f"   ✓ New class created: {class_obj.name}")
                else:
                    print(f"   ⚠ Class already exists: {class_obj.name}")
            elif current_class_id:
                print(f"   ✓ Using existing class ID: {current_class_id}")
                try:
                    if isinstance(current_class_id, Class):
                        student.current_class = current_class_id
                        print(f"   ✓ Class is already an object: {current_class_id.name}")
                    else:
                        student.current_class = Class.objects.get(id=current_class_id)
                        print(f"   ✓ Class retrieved: {student.current_class.name}")
                except (Class.DoesNotExist, ValueError, TypeError) as e:
                    print(f"   ✗ Error: {e}")
                    messages.error(request, 'Selected class does not exist.')
                    return render(request, 'digitallibrary/student_form.html', {'form': form})
            else:
                print("   ⚠ No class selected")
            
            # Save the student first
            student.save()
            print(f"\n💾 Student saved with ID: {student.id}")
            
            # ========== HANDLE ELECTIVE SUBJECTS ==========
            print("\n📚 HANDLING ELECTIVE SUBJECTS:")
            
            elective_subjects = request.POST.getlist('elective_subjects')
            print(f"   Method 1 - getlist('elective_subjects'): {elective_subjects}")
            
            subjects_hidden = request.POST.get('elective_subjects_hidden', '')
            if subjects_hidden:
                hidden_subjects = [s.strip() for s in subjects_hidden.split(',') if s.strip()]
                print(f"   Method 2 - from hidden input: {hidden_subjects}")
                if hidden_subjects and not elective_subjects:
                    elective_subjects = hidden_subjects
            
            single_subject = request.POST.get('elective_subjects', '')
            if single_subject and not elective_subjects:
                print(f"   Method 3 - single value: {single_subject}")
                elective_subjects = [single_subject]
            
            # Clear any existing subjects first
            student.subjects.clear()
            print(f"   ✓ Cleared existing subjects")
            
            # Add elective subjects
            if elective_subjects and len(elective_subjects) > 0:
                print(f"   ✓ Found {len(elective_subjects)} elective subjects to add:")
                added_count = 0
                for subject_name in elective_subjects:
                    if subject_name and subject_name.strip():
                        subject_obj, created = Subject.objects.get_or_create(
                            name=subject_name.strip(),
                            defaults={'is_active': True}
                        )
                        student.subjects.add(subject_obj)
                        added_count += 1
                        print(f"      + Added elective subject: {subject_name} (created: {created})")
                
                if added_count > 0:
                    messages.info(request, f'{added_count} elective subjects selected.')
                else:
                    print("   ⚠ No valid elective subjects found after filtering")
            else:
                print("   ⚠ No elective subjects found in POST (empty list)")
            
            # ========== ADD COMPULSORY SUBJECTS ==========
            print(f"\n📖 COMPULSORY SUBJECTS:")
            print(f"   Student pathway: {student.pathway}")
            
            if student.pathway:
                compulsory_subjects = {
                    'arts_sports': [
                        'English', 'Kiswahili/KSL', 'Core Mathematics', 
                        'Community Service Learning (CSL)'
                    ],
                    'social_sciences': [
                        'English', 'Kiswahili/KSL', 'Core Mathematics', 
                        'Community Service Learning (CSL)'
                    ],
                    'stem': [
                        'English', 'Kiswahili/KSL', 'Core Mathematics', 
                        'Community Service Learning (CSL)'
                    ],
                }
                
                pathway_subjects = compulsory_subjects.get(student.pathway, [])
                print(f"   Compulsory subjects for '{student.pathway}': {pathway_subjects}")
                
                for subject_name in pathway_subjects:
                    subject_obj, _ = Subject.objects.get_or_create(
                        name=subject_name,
                        defaults={
                            'is_compulsory': True,
                            'category': 'compulsory'
                        }
                    )
                    student.subjects.add(subject_obj)
                    print(f"      + Added compulsory subject: {subject_name}")
            else:
                print("   ⚠ No pathway selected - skipping compulsory subjects")
            
            # ========== FINAL SUMMARY ==========
            final_subject_count = student.subjects.count()
            print(f"\n✅ FINAL SUMMARY:")
            print(f"   Student ID: {student.id}")
            print(f"   Student Name: {student.first_name} {student.last_name}")
            print(f"   Class: {student.current_class.name if student.current_class else 'None'}")
            print(f"   Pathway: {student.pathway or 'None'}")
            print(f"   Total Subjects: {final_subject_count}")
            
            if final_subject_count > 0:
                print("\n   Subjects enrolled:")
                for idx, subj in enumerate(student.subjects.all(), 1):
                    print(f"      {idx}. {subj.name}")
            else:
                print("\n   ⚠ WARNING: No subjects were added to the student!")
            
            print("\n" + "=" * 80)
            print("✅ STUDENT CREATION COMPLETE")
            print("=" * 80 + "\n")
            
            messages.success(request, f'Student {student.first_name} {student.last_name} created successfully!')
            return redirect('digitallibrary:student_detail', pk=student.pk)
        else:
            print("\n❌ FORM IS INVALID:")
            print(f"   Form errors: {form.errors}")
            for field, errors in form.errors.items():
                print(f"   {field}: {', '.join(errors)}")
            messages.error(request, f'Please correct the errors below: {", ".join([f for field, errors in form.errors.items() for f in errors])}')
    else:
        print("\n📝 REQUEST METHOD: GET - Showing empty form")
        form = StudentForm()
    
    # Get all classes for the dropdown
    classes = Class.objects.all().order_by('name')
    
    context = {
        'form': form,
        'classes': classes,
        'title': 'Create Student',
        'action': 'Create',
    }
    
    return render(request, 'digitallibrary/student_form.html', context)
# ========== PRINT FEE STRUCTURE VIEW ==========

@fees_access
def print_fee_structure(request, fee_structure_id):
    """Print fee structure details"""
    from .models import FeeStructure, FeeComponent, SchoolSetting
    
    fee_structure = get_object_or_404(FeeStructure, id=fee_structure_id)
    school = SchoolSetting.objects.first()
    
    # Get fee components
    fee_components = FeeComponent.objects.filter(fee_structure=fee_structure)
    
    context = {
        'fee_structure': fee_structure,
        'fee_components': fee_components,
        'school': school,
        'school_name': school.name if school else 'School Name',
        'school_logo': school.logo.url if school and school.logo else None,
        'school_motto': school.motto if school else 'Excellence in Education',
        'title': 'Print Fee Structure',
        'is_print_view': True,
    }
    return render(request, 'fees/fee_structure_print.html', context)


# ========== FEE UPDATE PAGE ==========

def fee_update_page(request):
    """Page to update student fees"""
    from .models import Student, FeeStructure, Payment
    
    students = Student.objects.select_related('current_class').all()
    
    for student in students:
        # Calculate total fees for student
        fee_items = FeeStructure.objects.filter(
            student_class=student.current_class
        )
        student.total_fees = sum(fee.amount for fee in fee_items)
        
        # Calculate total paid
        payments = Payment.objects.filter(student=student)
        student.total_paid = payments.aggregate(total=models.Sum('amount'))['total'] or 0
        
        student.balance = student.total_fees - student.total_paid
    
    return render(request, 'digitallibrary/fee_update_form.html', {
        'students': students
    })


def update_student_fees(request):
    """Process fee updates"""
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        action = request.POST.get('action')
        
        if action == 'record_payment':
            # Redirect to payment recording page with selected students
            ids = ','.join(student_ids)
            return redirect(f"{reverse('digitallibrary:payment_record')}?students={ids}")
    
    return redirect('digitallibrary:fee_update_page')


# ========== GET STUDENTS BY CLASS ==========

def get_students_by_class(request, class_id):
    """API to get students for a specific class"""
    from .models import Student
    
    students = Student.objects.filter(
        current_class_id=class_id,
        is_active=True
    ).values('id', 'admission_number', 'first_name', 'last_name')
    
    return JsonResponse({
        'success': True,
        'students': list(students)
    })


def get_all_students(request):
    """API to get all active students"""
    from .models import Student
    
    students = Student.objects.filter(
        is_active=True
    ).values('id', 'admission_number', 'first_name', 'last_name')
    
    return JsonResponse({
        'success': True,
        'students': list(students)
    })


# ========== SUBMIT FEEDBACK ==========

from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
import json

@csrf_exempt
@require_http_methods(["POST"])
def submit_feedback(request):
    """Submit feedback with full school info"""
    from .models import Feedback
    
    # Get client IP - helper function
    def get_client_ip(req):
        x_forwarded_for = req.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = req.META.get('REMOTE_ADDR')
        return ip
    
    # Send notification - helper function
    def send_feedback_notification(feedback):
        from django.conf import settings
        from django.core.mail import send_mail
        try:
            subject = f"[Feedback] {feedback.school_name or 'Unknown School'} - {feedback.subject}"
            message = f"""
New Feedback Received

SCHOOL INFORMATION
School: {feedback.school_name or 'Unknown'}
Location: {feedback.school_location or 'Not specified'}
School ID: {feedback.school_id or 'N/A'}

USER INFORMATION
Name: {feedback.user_name or 'Anonymous'}
Role: {feedback.user_role or 'User'}
Email: {feedback.user_email or 'Not provided'}

FEEDBACK DETAILS
Type: {feedback.get_feedback_type_display()}
Priority: {feedback.get_priority_display()}
Rating: {feedback.rating or 'No rating'}/5
Subject: {feedback.subject}

Message:
{feedback.message}

Time: {feedback.created_at.strftime('%Y-%m-%d %H:%M:%S')}
"""
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [settings.ADMIN_EMAIL],
                fail_silently=False,
            )
        except Exception as e:
            print(f"❌ Failed to send email: {e}")
    
    try:
        # Parse JSON data - read body ONCE
        data = json.loads(request.body)
        
        # Get school from tenant or request
        school = None
        if hasattr(request, 'tenant'):
            school = request.tenant
        
        # Get user info
        user = request.user if request.user.is_authenticated else None
        user_role = None
        if user and hasattr(user, 'profile'):
            user_role = user.profile.role
        
        # Create feedback with all available info
        feedback = Feedback.objects.create(
            user=user,
            user_role=data.get('user_role') or user_role,
            user_email=data.get('user_email') or (user.email if user else None),
            user_name=data.get('user_name') or (user.get_full_name() if user else None),
            school_id=data.get('school_id') or (getattr(school, 'school_id', None) if school else None),
            school_name=data.get('school_name') or (school.name if school else None),
            school_location=data.get('school_location') or (getattr(school, 'location', None) if school else None),
            school_email=data.get('school_email') or (getattr(school, 'contact_email', None) if school else None),
            school_phone=data.get('school_phone') or (getattr(school, 'contact_phone', None) if school else None),
            school_domain=data.get('school_domain') or (getattr(school, 'domain', None) if school else None),
            school_subdomain=data.get('school_subdomain') or request.headers.get('X-Subdomain'),
            feedback_type=data.get('feedback_type', 'general'),
            priority=data.get('priority', 'medium'),
            subject=data.get('subject', '')[:200],
            message=data.get('message', ''),
            rating=int(data.get('rating', 0)) if data.get('rating') else None,
            page_url=data.get('page_url', ''),
            browser_info=request.headers.get('User-Agent', '')[:500],
            ip_address=get_client_ip(request),
        )
        
        print(f"✅ Feedback saved - School: {feedback.school_name}, ID: {feedback.school_id}")
        
        # Send notification
        send_feedback_notification(feedback)
        
        return JsonResponse({
            'status': 'success',
            'message': 'Feedback submitted successfully',
            'feedback_id': feedback.id
        })
        
    except json.JSONDecodeError as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Invalid JSON: {str(e)}'
        }, status=400)
    except Exception as e:
        print(f"❌ Error in submit_feedback: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


# ========== HELPERS ==========

def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def send_feedback_notification(feedback):
    """Send email notification for new feedback"""
    from django.conf import settings
    from django.core.mail import send_mail
    
    try:
        subject = f"[Feedback] {feedback.school_name or 'Unknown School'} - {feedback.subject}"
        
        message = f"""
        New Feedback Received
        
        SCHOOL INFORMATION
        School: {feedback.school_name or 'Unknown'}
        Location: {feedback.school_location or 'Not specified'}
        School ID: {feedback.school_id or 'N/A'}
        
        USER INFORMATION
        Name: {feedback.user_name or 'Anonymous'}
        Role: {feedback.user_role or 'User'}
        Email: {feedback.user_email or 'Not provided'}
        
        FEEDBACK DETAILS
        Type: {feedback.get_feedback_type_display()}
        Priority: {feedback.get_priority_display()}
        Rating: {feedback.rating or 'No rating'}/5
        Subject: {feedback.subject}
        
        Message:
        {feedback.message}
        
        Time: {feedback.created_at.strftime('%Y-%m-%d %H:%M:%S')}
        """
        
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [settings.ADMIN_EMAIL],
            fail_silently=False,
        )
        print(f"✅ Feedback email sent for {feedback.school_name}")
    except Exception as e:
        print(f"❌ Failed to send email: {e}")
# ========== PERFORMANCE DASHBOARD VIEWS ==========

@tenant_app_view
def performance_dashboard(request):
    """Performance dashboard with actual data"""
    from .models import Exam, Student, Subject, Class, SchoolSetting, StudentResult
    from django.db.models import Avg, Count
    
    # Try to import the correct result model
    Result = StudentResult
    
    # Get filter parameters
    current_year = request.GET.get('year', '')
    current_term = request.GET.get('term', '')
    selected_class = request.GET.get('class', '')
    selected_subject = request.GET.get('subject', '')
    
    # Base queryset for exams
    exams_qs = Exam.objects.all()
    if current_year:
        exams_qs = exams_qs.filter(academic_year=current_year)
    if current_term:
        exams_qs = exams_qs.filter(term=current_term)
    if selected_class:
        exams_qs = exams_qs.filter(student_class_id=selected_class)
    
    exams = exams_qs
    
    # Get all students
    students_qs = Student.objects.filter(is_active=True)
    if selected_class:
        students_qs = students_qs.filter(current_class_id=selected_class)
    total_students = students_qs.count()
    
    # Initialize empty values
    avg_score = 0
    pass_rate = 0
    grade_distribution = {}
    top_students = []
    subject_performance = []
    total_results = 0
    
    # If Result model exists, get performance data
    if Result:
        results_qs = Result.objects.all()
        if current_year:
            results_qs = results_qs.filter(exam__academic_year=current_year)
        if current_term:
            results_qs = results_qs.filter(exam__term=current_term)
        if selected_class:
            results_qs = results_qs.filter(student__current_class_id=selected_class)
        if selected_subject:
            results_qs = results_qs.filter(subject_id=selected_subject)
        
        total_results = results_qs.count()
        
        # Calculate average score
        avg_score_data = results_qs.aggregate(avg=Avg('score'))
        avg_score = avg_score_data['avg'] or 0
        
        # Calculate pass rate (score >= 50%)
        passed_results = results_qs.filter(score__gte=50).count()
        pass_rate = (passed_results / total_results * 100) if total_results > 0 else 0
        
        # Grade distribution
        grade_distribution = {
            'A (80-100)': results_qs.filter(score__gte=80).count(),
            'B (70-79)': results_qs.filter(score__gte=70, score__lt=80).count(),
            'C (60-69)': results_qs.filter(score__gte=60, score__lt=70).count(),
            'D (50-59)': results_qs.filter(score__gte=50, score__lt=60).count(),
            'E (0-49)': results_qs.filter(score__lt=50).count(),
        }
        
        # Top performing students
        top_students_data = results_qs.values('student').annotate(
            avg=Avg('score')
        ).order_by('-avg')[:10]
        
        for ts in top_students_data:
            student = Student.objects.filter(id=ts['student']).first()
            if student:
                avg = ts['avg']
                if avg >= 80:
                    grade = 'A'
                elif avg >= 75:
                    grade = 'A-'
                elif avg >= 70:
                    grade = 'B+'
                elif avg >= 65:
                    grade = 'B'
                elif avg >= 60:
                    grade = 'B-'
                elif avg >= 55:
                    grade = 'C+'
                elif avg >= 50:
                    grade = 'C'
                elif avg >= 45:
                    grade = 'C-'
                elif avg >= 40:
                    grade = 'D+'
                else:
                    grade = 'E'
                top_students.append({
                    'student': student,
                    'average': avg,
                    'grade': grade,
                })
        
        # Subject performance
        subjects = Subject.objects.all()
        for subject in subjects:
            subject_results = results_qs.filter(subject=subject)
            if subject_results.exists():
                avg = subject_results.aggregate(avg=Avg('score'))['avg'] or 0
                if avg >= 80:
                    grade = 'A'
                elif avg >= 75:
                    grade = 'A-'
                elif avg >= 70:
                    grade = 'B+'
                elif avg >= 65:
                    grade = 'B'
                elif avg >= 60:
                    grade = 'B-'
                elif avg >= 55:
                    grade = 'C+'
                elif avg >= 50:
                    grade = 'C'
                elif avg >= 45:
                    grade = 'C-'
                elif avg >= 40:
                    grade = 'D+'
                else:
                    grade = 'E'
                subject_performance.append({
                    'name': subject.name,
                    'students': subject_results.values('student').distinct().count(),
                    'average': avg,
                    'grade': grade,
                })
    
    # Available years for filter
    available_years = Exam.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
    
    # Classes for filter
    classes = Class.objects.all().order_by('name')
    
    # Subjects for filter
    subjects = Subject.objects.all().order_by('name')
    
    school = SchoolSetting.objects.first()
    
    context = {
        'total_students': total_students,
        'total_exams': exams.count(),
        'avg_score': avg_score,
        'pass_rate': pass_rate,
        'grade_distribution': grade_distribution,
        'top_students': top_students,
        'subject_performance': subject_performance,
        'available_years': available_years,
        'current_year': current_year,
        'current_term': current_term,
        'selected_class': selected_class,
        'selected_subject': selected_subject,
        'classes': classes,
        'subjects': subjects,
        'total_results': total_results,
        'school': school,
        'school_name': school.name if school else 'Performance Dashboard',
        'school_logo': school.logo.url if school and school.logo else None,
        'school_motto': school.motto if school else '',
    }
    
    return render(request, 'performance/dashboard.html', context)


def exam_performance_detail(request, exam_id):
    """View detailed performance for a specific exam"""
    from .models import Exam, Subject, Student, StudentResult
    from django.db.models import Avg, Sum
    
    exam = Exam.objects.get(id=exam_id)
    
    if exam.student_class:
        students = exam.student_class.students.filter(is_active=True)
    else:
        students = Student.objects.filter(is_active=True)
    
    total_students = students.count()
    subjects = Subject.objects.all()
    total_subjects = subjects.count()
    results = StudentResult.objects.filter(exam=exam)
    
    class_average = results.aggregate(avg=Avg('score'))['avg'] or 0
    
    total_results = results.values('student').distinct().count()
    passed_results = results.filter(score__gte=50).values('student').distinct().count()
    pass_rate = (passed_results / total_results * 100) if total_results > 0 else 0
    
    top_student_data = results.values('student').annotate(total=Sum('score')).order_by('-total').first()
    top_student = None
    if top_student_data:
        top_student = Student.objects.filter(id=top_student_data['student']).first()
    
    subjects_performance = []
    for subject in subjects:
        subject_results = results.filter(subject=subject)
        if subject_results.exists():
            avg = subject_results.aggregate(avg=Avg('score'))['avg'] or 0
            highest = subject_results.aggregate(max=Avg('score'))['max'] or 0
            lowest = subject_results.aggregate(min=Avg('score'))['min'] or 0
            passed = subject_results.filter(score__gte=50).count()
            
            if avg >= 80:
                grade = 'A'
            elif avg >= 70:
                grade = 'B'
            elif avg >= 60:
                grade = 'C'
            elif avg >= 50:
                grade = 'D'
            else:
                grade = 'E'
            
            subjects_performance.append({
                'id': subject.id,
                'name': subject.name,
                'average': avg,
                'highest': highest,
                'lowest': lowest,
                'passed': passed,
                'total_students': total_students,
                'grade': grade,
            })
    
    rankings = []
    for student in students:
        student_results = results.filter(student=student)
        if student_results.exists():
            subject_scores = []
            for subject in subjects:
                subject_result = student_results.filter(subject=subject).first()
                subject_scores.append(subject_result.score if subject_result else None)
            
            total = sum([r.score for r in student_results if r.score])
            average = total / student_results.count() if student_results.count() > 0 else 0
            
            if average >= 80:
                grade = 'A'
            elif average >= 75:
                grade = 'A-'
            elif average >= 70:
                grade = 'B+'
            elif average >= 65:
                grade = 'B'
            elif average >= 60:
                grade = 'B-'
            elif average >= 55:
                grade = 'C+'
            elif average >= 50:
                grade = 'C'
            elif average >= 45:
                grade = 'C-'
            elif average >= 40:
                grade = 'D+'
            else:
                grade = 'E'
            
            rankings.append({
                'student': student,
                'subject_scores': subject_scores,
                'total': total,
                'average': average,
                'grade': grade,
            })
    
    rankings.sort(key=lambda x: x['average'], reverse=True)
    
    context = {
        'exam': exam,
        'total_students': total_students,
        'total_subjects': total_subjects,
        'class_average': class_average,
        'pass_rate': pass_rate,
        'top_student': top_student,
        'subjects_performance': subjects_performance,
        'subjects_list': subjects,
        'rankings': rankings,
    }
    
    return render(request, 'performance/exam_performance_detail.html', context)


def system_dashboard(request):
    """Executive dashboard with filtering"""
    from .models import Exam, Class, Subject, Student, StudentResult
    from django.db.models import Avg
    
    current_year = request.GET.get('year', '')
    current_term = request.GET.get('term', '')
    selected_class = request.GET.get('class', '')
    selected_subject = request.GET.get('subject', '')
    
    exams_qs = Exam.objects.all().order_by('-academic_year', '-created_at')
    
    if current_year:
        exams_qs = exams_qs.filter(academic_year=current_year)
    if current_term:
        exams_qs = exams_qs.filter(term=current_term)
    if selected_class:
        exams_qs = exams_qs.filter(student_class_id=selected_class)
    
    available_years = Exam.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
    
    recent_exams = []
    for exam in exams_qs:
        if exam.student_class:
            total_students = exam.student_class.students.filter(is_active=True).count()
        else:
            total_students = Student.objects.filter(is_active=True).count()
        
        results_qs = StudentResult.objects.filter(exam=exam)
        
        if selected_subject:
            results_qs = results_qs.filter(subject_id=selected_subject)
            results_count = results_qs.values('student').distinct().count()
        else:
            results_count = results_qs.values('student').distinct().count()
        
        completion_rate = (results_count / total_students * 100) if total_students > 0 else 0
        
        recent_exams.append({
            'id': exam.id,
            'name': exam.name,
            'academic_year': exam.academic_year,
            'term': exam.term,
            'class_name': exam.student_class.name if exam.student_class else 'All Classes',
            'completion_rate': completion_rate,
            'total_students': total_students,
            'results_count': results_count,
        })
    
    results_qs = StudentResult.objects.all()
    if current_year:
        results_qs = results_qs.filter(exam__academic_year=current_year)
    if current_term:
        results_qs = results_qs.filter(exam__term=current_term)
    if selected_class:
        results_qs = results_qs.filter(student__current_class_id=selected_class)
    if selected_subject:
        results_qs = results_qs.filter(subject_id=selected_subject)
    
    avg_score = results_qs.aggregate(avg=Avg('score'))['avg'] or 0
    
    total_results = results_qs.count()
    passed_results = results_qs.filter(score__gte=50).count()
    pass_rate = (passed_results / total_results * 100) if total_results > 0 else 0
    
    students_qs = Student.objects.filter(is_active=True)
    if selected_class:
        students_qs = students_qs.filter(current_class_id=selected_class)
    total_students = students_qs.count()
    
    total_exams = exams_qs.count()
    classes = Class.objects.all().order_by('name')
    subjects = Subject.objects.all().order_by('name')
    
    context = {
        'total_students': total_students,
        'total_exams': total_exams,
        'avg_score': avg_score,
        'pass_rate': pass_rate,
        'recent_exams': recent_exams,
        'current_year': current_year,
        'current_term': current_term,
        'selected_class': selected_class,
        'selected_subject': selected_subject,
        'classes': classes,
        'subjects': subjects,
        'available_years': available_years,
    }
    
    return render(request, 'performance/system_dashboard.html', context)
# ========== ENTER RESULTS FORM VIEW ==========


@tenant_app_view
def enter_results_form(request):
    """
    Streamlined results entry page - select exam first, then subject, then enter scores.
    """
    from .models import Exam, Subject, Student, StudentResult, GradingSystem, SchoolSetting
    from django.db import connection
    from django_tenants.utils import get_tenant_model
    from django.shortcuts import redirect
    from django.contrib import messages
    import json

    # Get tenant
    School = get_tenant_model()
    from django_tenants.utils import get_tenant
    tenant = get_tenant(request)
    connection.set_tenant(tenant)

    print("\n" + "="*60)
    print("🔵 enter_results_form called")
    print(f"   Method: {request.method}")
    if request.method == 'POST':
        print(f"   POST data: {request.POST}")
        print(f"   POST keys: {list(request.POST.keys())}")
    print("="*60)

    # Get all exams and subjects
    exams = Exam.objects.all().order_by('-academic_year', '-created_at')
    subjects = Subject.objects.all().order_by('name')
    
    # Get all active grading systems
    all_grading_systems = GradingSystem.objects.filter(is_active=True).order_by('-is_default', 'name')
    
    # Get selected exam and subject
    selected_exam_id = request.GET.get('exam')
    selected_subject_id = request.GET.get('subject')
    
    # Handle POST request - SAVE RESULTS
    if request.method == 'POST':
        exam_id = request.POST.get('exam_id')
        subject_id = request.POST.get('subject_id')
        
        print(f"   exam_id from POST: {exam_id}")
        print(f"   subject_id from POST: {subject_id}")
        
        if exam_id and subject_id:
            try:
                exam = Exam.objects.get(id=exam_id)
                subject = Subject.objects.get(id=subject_id)
                
                saved_count = 0
                
                # Loop through all POST data to find scores
                for key, value in request.POST.items():
                    if key.startswith('score_') and value:
                        student_id = key.replace('score_', '')
                        print(f"   Processing: Student {student_id}, Score {value}")
                        
                        try:
                            student_id = int(student_id)
                            score = float(value)
                            
                            # Validate score range
                            if score < 0 or (exam.max_score and score > exam.max_score):
                                print(f"   ⚠️ Score {score} out of range")
                                continue
                            
                            # Get CBE grade
                            with connection.cursor() as cursor:
                                cursor.execute("""
                                    SELECT id, points 
                                    FROM digitallibrary_kneccbegrade 
                                    WHERE min_score <= %s AND max_score >= %s
                                    LIMIT 1
                                """, [score, score])
                                grade_row = cursor.fetchone()
                                
                                if grade_row:
                                    grade_id, points = grade_row
                                    
                                    # Save using direct SQL
                                    cursor.execute("""
                                        INSERT INTO digitallibrary_studentresult 
                                        (student_id, exam_id, subject_id, score, grade_id, points, entered_by_id, entered_at, updated_at)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                                        ON CONFLICT (student_id, exam_id, subject_id) 
                                        DO UPDATE SET 
                                            score = EXCLUDED.score,
                                            grade_id = EXCLUDED.grade_id,
                                            points = EXCLUDED.points,
                                            updated_at = NOW()
                                    """, [student_id, exam.id, subject.id, score, grade_id, points, request.user.id])
                                    
                                    saved_count += 1
                                    print(f"   ✅ Saved!")
                                else:
                                    print(f"   ❌ No grade found")
                                    
                        except Exception as e:
                            print(f"   ❌ Error: {e}")
                
                if saved_count > 0:
                    messages.success(request, f'✅ Successfully saved {saved_count} results!')
                else:
                    messages.warning(request, '⚠️ No results were saved.')
                    
            except Exception as e:
                messages.error(request, f'Error: {e}')
                print(f"   ❌ Exception: {e}")
        else:
            messages.error(request, 'Missing exam or subject')
        
        # Redirect back
        return redirect(f'{request.path}?exam={exam_id}&subject={subject_id}')
    
    # ============================================================
    # GET REQUEST - LOAD THE FORM
    # ============================================================
    selected_exam = None
    selected_subject = None
    students = []
    existing_results = {}
    
    if selected_exam_id:
        try:
            selected_exam = Exam.objects.get(id=selected_exam_id)
            print(f"\n📋 Loading form for exam: {selected_exam.name}")
            
            # Get students
            if selected_exam.student_class:
                students = list(selected_exam.student_class.students.filter(is_active=True))
            else:
                students = list(Student.objects.filter(is_active=True))
            
            students.sort(key=lambda x: (x.first_name, x.last_name))
            
            if selected_subject_id:
                try:
                    selected_subject = Subject.objects.get(id=selected_subject_id)
                    
                    # Get existing results
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            SELECT s.student_id, s.score, g.grade, s.points
                            FROM digitallibrary_studentresult s
                            LEFT JOIN digitallibrary_kneccbegrade g ON s.grade_id = g.id
                            WHERE s.exam_id = %s AND s.subject_id = %s
                        """, [selected_exam.id, selected_subject.id])
                        
                        for row in cursor.fetchall():
                            existing_results[row[0]] = {
                                'score': row[1],
                                'grade': row[2],
                                'points': row[3]
                            }
                    print(f"   Found {len(existing_results)} existing results")
                    
                except Subject.DoesNotExist:
                    pass
                    
        except Exam.DoesNotExist:
            print(f"Exam not found: {selected_exam_id}")
            selected_exam = None
    
    # Build results dict
    results_dict = {}
    for student in students:
        results_dict[student.id] = existing_results.get(student.id)
    
    context = {
        'exams': exams,
        'subjects': subjects,
        'selected_exam': selected_exam,
        'selected_subject': selected_subject,
        'students': students,
        'existing_results': results_dict,
        'all_grading_systems': all_grading_systems,
        'active_grading_system': None,
        'school': SchoolSetting.objects.first(),
    }
    
    return render(request, 'performance/enter_results_form.html', context)    
    
def bulk_select(request):
    """Step 1: Select exam and subject for bulk entry"""
    from .models import Exam, Subject, Student
    
    exams = Exam.objects.all().order_by('-academic_year', '-created_at')
    subjects = Subject.objects.all().order_by('name')
    
    selected_exam_id = None
    selected_subject_id = None
    selected_exam = None
    total_students = 0
    
    if request.method == 'POST':
        exam_id = request.POST.get('exam')
        subject_id = request.POST.get('subject')
        
        if exam_id and subject_id:
            try:
                exam = Exam.objects.get(id=exam_id)
                subject = Subject.objects.get(id=subject_id)
                return redirect('digitallibrary:bulk_results_entry', exam_id=exam.id, subject_id=subject.id)
            except (Exam.DoesNotExist, Subject.DoesNotExist):
                messages.error(request, 'Invalid selection')
        
        selected_exam_id = exam_id
        selected_subject_id = subject_id
        if selected_exam_id:
            try:
                selected_exam = Exam.objects.get(id=selected_exam_id)
            except Exam.DoesNotExist:
                pass
    else:
        exam_id = request.GET.get('exam')
        if exam_id:
            try:
                selected_exam = Exam.objects.get(id=exam_id)
                selected_exam_id = exam_id
            except Exam.DoesNotExist:
                pass
    
    if selected_exam:
        if selected_exam.student_class:
            total_students = selected_exam.student_class.students.filter(is_active=True).count()
        else:
            total_students = Student.objects.filter(is_active=True).count()
    
    context = {
        'exams': exams,
        'subjects': subjects,
        'selected_exam_id': selected_exam_id,
        'selected_subject_id': selected_subject_id,
        'selected_exam': selected_exam,
        'total_students': total_students,
    }
    
    return render(request, 'digitallibrary/bulk_select.html', context)


def bulk_results_entry(request, exam_id, subject_id):
    """Step 2: Enter results for all students in a table"""
    from .models import Exam, Subject, Student, StudentResult
    
    exam = Exam.objects.get(id=exam_id)
    subject = Subject.objects.get(id=subject_id)
    
    if exam.student_class:
        students = exam.student_class.students.filter(is_active=True)
    else:
        students = Student.objects.filter(is_active=True)
    
    students = students.order_by('first_name', 'last_name')
    
    existing_results = {}
    results = StudentResult.objects.filter(exam=exam, subject=subject, student__in=students)
    existing_results = {r.student_id: r for r in results}
    
    if request.method == 'POST':
        saved_count = 0
        for key, value in request.POST.items():
            if key.startswith('score_') and value:
                student_id = key.replace('score_', '')
                try:
                    score = float(value)
                    student = Student.objects.get(id=student_id)
                    
                    result, created = StudentResult.objects.update_or_create(
                        exam=exam,
                        subject=subject,
                        student=student,
                        defaults={'score': score}
                    )
                    saved_count += 1
                except (ValueError, Student.DoesNotExist):
                    continue
        
        messages.success(request, f'Successfully saved {saved_count} results for {subject.name}')
        return redirect('digitallibrary:bulk_results_entry', exam_id=exam.id, subject_id=subject.id)
    
    context = {
        'exam': exam,
        'subject': subject,
        'students': students,
        'existing_results': existing_results,
        'completion_percentage': int((len(existing_results) / len(students)) * 100) if students else 0,
        'pending_count': len(students) - len(existing_results) if students else 0,
    }
    
    return render(request, 'digitallibrary/bulk_results_entry.html', context)


def exam_results_entry(request, exam_id):
    """Enter results for an exam - by subject, filtered by registered student subjects"""
    from .models import Exam, Subject, Student, StudentResult
    
    exam = get_object_or_404(Exam, pk=exam_id)
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    
    selected_subject_id = request.GET.get('subject')
    selected_subject = None
    students = []
    existing_results = {}
    
    if selected_subject_id:
        try:
            selected_subject = Subject.objects.get(pk=selected_subject_id, is_active=True)
            
            students_qs = exam.get_students_for_exam()
            students = students_qs.filter(
                subjects=selected_subject,
                is_active=True
            ).distinct().order_by('admission_number')
            
            existing_results_qs = StudentResult.objects.filter(
                exam=exam,
                subject=selected_subject,
                student__in=students
            ).select_related('student')
            
            existing_results = {result.student_id: result for result in existing_results_qs}
            
        except Subject.DoesNotExist:
            messages.error(request, "Selected subject does not exist.")
    
    if request.method == 'POST':
        subject_id = request.POST.get('subject_id')
        
        if subject_id:
            selected_subject = get_object_or_404(Subject, pk=subject_id, is_active=True)
            
            students = exam.get_students_for_exam().filter(
                subjects=selected_subject,
                is_active=True
            ).distinct()
            
            saved_count = 0
            
            for student in students:
                score_key = f'score_{student.id}'
                if score_key in request.POST:
                    score = request.POST.get(score_key)
                    
                    if score and score.strip():
                        try:
                            score_value = float(score)
                            if 0 <= score_value <= (exam.max_score or 100):
                                StudentResult.objects.update_or_create(
                                    student=student,
                                    exam=exam,
                                    subject=selected_subject,
                                    defaults={'score': score_value, 'entered_by': request.user}
                                )
                                saved_count += 1
                        except ValueError:
                            pass
            
            if saved_count > 0:
                messages.success(request, f'Results for {exam.name} - {selected_subject.name} saved successfully!')
            else:
                messages.warning(request, 'No results were saved.')
            
            return redirect(f'{request.path}?subject={subject_id}')
    
    context = {
        'exam': exam,
        'subjects': subjects,
        'selected_subject': selected_subject,
        'students': students,
        'existing_results': existing_results,
        'title': f'Enter Results - {exam.name}',
        'school': SchoolSetting.objects.first(),
    }
    
    return render(request, 'performance/exam_results_entry.html', context)
# ========== EXPORT EXAM PERFORMANCE VIEW ==========

def export_exam_performance(request, exam_id):
    """Export exam performance to CSV"""
    import csv
    from django.http import HttpResponse
    from .models import Exam, Subject, Student, StudentResult
    
    exam = Exam.objects.get(id=exam_id)
    
    if exam.student_class:
        students = exam.student_class.students.filter(is_active=True)
    else:
        students = Student.objects.filter(is_active=True)
    
    subjects = Subject.objects.all()
    results = StudentResult.objects.filter(exam=exam)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{exam.name}_performance.csv"'
    
    writer = csv.writer(response)
    
    header = ['Rank', 'Admission Number', 'Student Name']
    for subject in subjects:
        header.append(subject.name)
    header.extend(['Total Score', 'Average Score', 'Grade', 'Status'])
    writer.writerow(header)
    
    rankings = []
    for student in students:
        student_results = results.filter(student=student)
        if student_results.exists():
            scores = []
            for subject in subjects:
                subject_result = student_results.filter(subject=subject).first()
                scores.append(subject_result.score if subject_result else '')
            
            total = sum([r.score for r in student_results])
            average = total / student_results.count()
            
            if average >= 80:
                grade = 'A'
            elif average >= 70:
                grade = 'B'
            elif average >= 60:
                grade = 'C'
            elif average >= 50:
                grade = 'D'
            else:
                grade = 'E'
            status = 'Pass' if average >= 50 else 'Fail'
            
            rankings.append({
                'student': student,
                'scores': scores,
                'total': total,
                'average': average,
                'grade': grade,
                'status': status,
            })
    
    rankings.sort(key=lambda x: x['average'], reverse=True)
    
    for idx, ranking in enumerate(rankings, 1):
        row = [idx, ranking['student'].admission_number, f"{ranking['student'].first_name} {ranking['student'].last_name}"]
        row.extend(ranking['scores'])
        row.extend([ranking['total'], f"{ranking['average']:.1f}", ranking['grade'], ranking['status']])
        writer.writerow(row)
    
    return response


def export_performance_report(request):
    """Export performance report to CSV"""
    import csv
    from django.http import HttpResponse
    from .models import StudentResult, Student
    from django.db.models import Avg
    
    academic_year = request.GET.get('year', '')
    term = request.GET.get('term', '')
    selected_class = request.GET.get('class', '')
    selected_exam = request.GET.get('exam', '')
    
    results_qs = StudentResult.objects.all()
    
    if academic_year:
        results_qs = results_qs.filter(exam__academic_year=academic_year)
    if term:
        results_qs = results_qs.filter(exam__term=term)
    if selected_class:
        results_qs = results_qs.filter(student__current_class_id=selected_class)
    if selected_exam:
        results_qs = results_qs.filter(exam_id=selected_exam)
    
    top_students = results_qs.values('student').annotate(
        avg=Avg('score')
    ).order_by('-avg')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="performance_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Rank', 'Admission Number', 'Student Name', 'Class', 'Average Score', 'Grade'])
    
    for idx, ts in enumerate(top_students, 1):
        student = Student.objects.filter(id=ts['student']).first()
        if student:
            avg = ts['avg']
            if avg >= 80:
                grade = 'A'
            elif avg >= 70:
                grade = 'B'
            elif avg >= 60:
                grade = 'C'
            elif avg >= 50:
                grade = 'D'
            else:
                grade = 'E'
            
            writer.writerow([
                idx,
                student.admission_number,
                f"{student.first_name} {student.last_name}",
                student.current_class.name if student.current_class else 'N/A',
                f"{avg:.1f}",
                grade
            ])
    
    return response


def export_ranking_csv(request, exam_id):
    """Export exam rankings to CSV"""
    import csv
    from django.http import HttpResponse
    from .models import Exam, ExamResultSummary
    
    exam = Exam.objects.get(id=exam_id)
    rankings = ExamResultSummary.objects.filter(exam=exam).select_related('student').order_by('rank')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{exam.name}_rankings.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Rank', 'Admission Number', 'Student Name', 'Total Score', 'Average Score', 'Grade'])
    
    for ranking in rankings:
        writer.writerow([
            ranking.rank,
            ranking.student.admission_number,
            f"{ranking.student.first_name} {ranking.student.last_name}",
            ranking.total_score,
            ranking.average_score,
            ranking.overall_grade,
        ])
    
    return response


def subject_exam_performance_detail(request, subject_id, exam_id):
    """View performance for a specific subject in a specific exam"""
    from .models import Subject, Exam, Student, StudentResult
    from django.db.models import Avg
    
    subject = Subject.objects.get(id=subject_id)
    exam = Exam.objects.get(id=exam_id)
    
    results = StudentResult.objects.filter(subject=subject, exam=exam).select_related('student')
    
    total_students = results.count()
    avg_score = results.aggregate(avg=Avg('score'))['avg'] or 0
    highest = results.aggregate(max=Avg('score'))['max'] or 0
    lowest = results.aggregate(min=Avg('score'))['min'] or 0
    passed = results.filter(score__gte=50).count()
    pass_rate = (passed / total_students * 100) if total_students > 0 else 0
    
    grade_distribution = {
        'A (80-100)': results.filter(score__gte=80).count(),
        'B (70-79)': results.filter(score__gte=70, score__lt=80).count(),
        'C (60-69)': results.filter(score__gte=60, score__lt=70).count(),
        'D (50-59)': results.filter(score__gte=50, score__lt=60).count(),
        'E (0-49)': results.filter(score__lt=50).count(),
    }
    
    context = {
        'subject': subject,
        'exam': exam,
        'results': results.order_by('-score'),
        'total_students': total_students,
        'avg_score': avg_score,
        'highest': highest,
        'lowest': lowest,
        'passed': passed,
        'pass_rate': pass_rate,
        'grade_distribution': grade_distribution,
    }
    
    return render(request, 'performance/subject_exam_performance_detail.html', context)


def subject_exam_performance(request, subject_id, exam_id):
    """View performance for a specific subject in an exam"""
    from .models import Subject, Exam, Result, Student
    
    subject = Subject.objects.get(id=subject_id)
    exam = Exam.objects.get(id=exam_id)
    
    results = Result.objects.filter(exam=exam, subject=subject).select_related('student')
    
    total_students = Student.objects.filter(is_active=True).count()
    avg_score = results.aggregate(avg=Avg('score'))['avg'] or 0
    top_score = results.aggregate(max=Avg('score'))['max'] or 0
    lowest_score = results.aggregate(min=Avg('score'))['min'] or 0
    
    context = {
        'subject': subject,
        'exam': exam,
        'results': results,
        'total_students': total_students,
        'avg_score': avg_score,
        'top_score': top_score,
        'lowest_score': lowest_score,
    }
    
    return render(request, 'performance/subject_exam_performance.html', context)


def view_subject_results(request, exam_id, subject_id):
    """View all results for a subject in an exam (for class teacher)"""
    from .models import Exam, Subject, Result
    
    exam = Exam.objects.get(id=exam_id)
    subject = Subject.objects.get(id=subject_id)
    results = Result.objects.filter(exam=exam, subject=subject).select_related('student').order_by('-score')
    
    context = {
        'exam': exam,
        'subject': subject,
        'results': results,
    }
    
    return render(request, 'performance/view_subject_results.html', context)
# ========== PARENT PORTAL VIEWS ==========

def parent_login(request):
    """Parent login using phone number and OTP"""
    from .forms import ParentLoginForm
    from .models import Student, ParentOTP
    from django.utils import timezone
    from datetime import timedelta
    
    if request.method == "POST":
        form = ParentLoginForm(request.POST)
        
        if form.is_valid():
            phone = form.cleaned_data["phone"]
            
            # Check if any student has this phone number
            students = Student.objects.filter(
                Q(parent_phone=phone) | Q(parent_alternative_phone=phone),
                is_active=True
            )
            
            if not students.exists():
                messages.error(request, "No student is linked to this phone number.")
                return redirect("digitallibrary:parent_login")
            
            # Generate OTP
            otp_code = ParentOTP.generate_otp()
            expires_at = timezone.now() + timedelta(minutes=10)
            
            ParentOTP.objects.create(
                phone=phone,
                otp_code=otp_code,
                expires_at=expires_at,
            )
            
            # Store phone in session
            request.session["parent_phone_pending"] = phone
            
            # Send SMS with OTP
            message = f"Your ShuleHub Parent Portal verification code is: {otp_code}"
            
            if settings.MOCK_SMS_MODE:
                messages.info(request, f"TEST MODE: Your OTP is: {otp_code}")
            else:
                try:
                    from .sms_utils import send_sms
                    result = send_sms(phone, message)
                    if result['success']:
                        messages.success(request, f"Verification code sent to {phone}")
                    else:
                        messages.error(request, f"SMS delivery failed. Please use this code: {otp_code}")
                except Exception as e:
                    messages.error(request, f"SMS service error. Please use this code: {otp_code}")
            
            return redirect("digitallibrary:verify_parent_otp")
    else:
        form = ParentLoginForm()
    
    return render(request, "parent_portal/parent_login.html", {
        "form": form,
        "title": "Parent Login",
    })


def verify_parent_otp(request):
    """Verify OTP for parent login"""
    from .forms import ParentOTPForm
    from .models import ParentOTP, Student
    
    phone = request.session.get("parent_phone_pending")
    
    if not phone:
        messages.error(request, "Please enter your phone number first.")
        return redirect("digitallibrary:parent_login")
    
    if request.method == "POST":
        form = ParentOTPForm(request.POST)
        
        if form.is_valid():
            otp_code = form.cleaned_data["otp_code"]
            
            # Check if OTP exists and is valid
            otp = ParentOTP.objects.filter(
                phone=phone,
                otp_code=otp_code,
                is_used=False,
            ).order_by("-created_at").first()
            
            if not otp:
                messages.error(request, "Invalid OTP code. Please try again.")
                return redirect("digitallibrary:verify_parent_otp")
            
            if otp.is_expired():
                messages.error(request, "OTP has expired. Please request a new one.")
                otp.delete()
                request.session.pop("parent_phone_pending", None)
                return redirect("digitallibrary:parent_login")
            
            # Mark OTP as used
            otp.is_used = True
            otp.save(update_fields=["is_used"])
            
            # Set parent phone in session and clear pending
            request.session["parent_phone"] = phone
            request.session.pop("parent_phone_pending", None)
            
            messages.success(request, "Login successful! Welcome to the Parent Portal.")
            return redirect("digitallibrary:parent_dashboard")
        else:
            messages.error(request, "Please enter a valid 6-digit OTP code.")
    else:
        form = ParentOTPForm()
    
    return render(request, "parent_portal/verify_parent_otp.html", {
        "form": form,
        "phone": phone,
        "title": "Verify OTP",
    })


def parent_logout(request):
    """Parent logout"""
    request.session.pop("parent_phone", None)
    request.session.pop("parent_phone_pending", None)
    messages.success(request, "You have been logged out.")
    return redirect("digitallibrary:parent_login")


def parent_dashboard(request):
    """Parent dashboard showing linked students with REAL-TIME fee information"""
    from decimal import Decimal
    from .models import Student, SchoolSetting, Term, FeePayment, HistoricalArrears, StudentResult, FeeStructure, FeeBalance
    from django.db.models import Sum, Q
    
    phone = request.session.get("parent_phone")
    
    if not phone:
        messages.error(request, "Please login first.")
        return redirect("digitallibrary:parent_login")
    
    # Get students linked to this parent phone number
    students = Student.objects.filter(
        Q(parent_phone=phone) | Q(parent_alternative_phone=phone),
        is_active=True
    ).select_related("current_class")
    
    # Get current active term
    current_term = Term.objects.filter(is_active=True).first()
    
    # If no active term, get the latest term
    if not current_term:
        current_term = Term.objects.order_by('-academic_year', '-term_number').first()
    
    # Calculate fee summary for each student - REAL TIME
    students_data = []
    total_fees_paid = Decimal('0.00')
    total_results = 0
    parent_name = None
    
    for student in students:
        # Get parent name from first student
        if not parent_name:
            parent_name = student.parent_name or "Parent"
        
        if current_term:
            academic_year = current_term.academic_year
            term_number = current_term.term_number
            
            # METHOD 1: Try to get from FeeBalance model first
            fee_balance = FeeBalance.objects.filter(
                student=student,
                academic_year=academic_year,
                term=term_number
            ).first()
            
            if fee_balance and fee_balance.total_expected > 0:
                # Use existing fee balance record
                total_expected = fee_balance.total_expected
                total_paid = fee_balance.total_paid
                current_balance = fee_balance.balance
            else:
                # METHOD 2: Calculate from scratch in REAL TIME
                # Calculate total expected from fee structures
                fee_structures = FeeStructure.objects.filter(
                    student_class=student.current_class,
                    academic_year=academic_year,
                    term=term_number
                )
                
                total_expected = Decimal('0.00')
                for fs in fee_structures:
                    total_expected += fs.total_fees
                
                # Calculate total paid from fee payments
                total_paid = FeePayment.objects.filter(
                    student=student,
                    academic_year=academic_year,
                    term=term_number
                ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                
                current_balance = total_expected - total_paid
                
                # Create or update FeeBalance for future use
                FeeBalance.objects.update_or_create(
                    student=student,
                    academic_year=academic_year,
                    term=term_number,
                    defaults={
                        'total_expected': total_expected,
                        'total_paid': total_paid,
                        'balance': current_balance,
                        'status': 'PAID' if current_balance == 0 else 'PARTIAL' if total_paid > 0 else 'DEFAULTING'
                    }
                )
            
            # Get historical arrears (unsettled only)
            historical_arrears = HistoricalArrears.objects.filter(
                student=student,
                is_settled=False
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            
            # Total outstanding including arrears
            total_outstanding = current_balance + historical_arrears
            
            total_fees_paid += total_paid
            
            # Debug output
            print(f"DEBUG: {student.first_name} {student.last_name}")
            print(f"  Expected: {total_expected}")
            print(f"  Paid: {total_paid}")
            print(f"  Balance: {current_balance}")
            print(f"  Arrears: {historical_arrears}")
            print(f"  Outstanding: {total_outstanding}")
        else:
            total_expected = Decimal('0.00')
            total_paid = Decimal('0.00')
            current_balance = Decimal('0.00')
            historical_arrears = Decimal('0.00')
            total_outstanding = Decimal('0.00')
        
        # Get results count
        results_count = StudentResult.objects.filter(
            student=student
        ).values('exam').distinct().count()
        total_results += results_count
        
        # Get subject count
        subject_count = student.subjects.count() or 8
        
        # Determine performance level
        performance = "Good"
        from .models import PerformanceSummary
        latest_performance = PerformanceSummary.objects.filter(
            student=student
        ).order_by('-academic_year', '-term').first()
        if latest_performance:
            if latest_performance.average_score >= 80:
                performance = "Excellent"
            elif latest_performance.average_score >= 70:
                performance = "Very Good"
            elif latest_performance.average_score >= 60:
                performance = "Good"
            elif latest_performance.average_score >= 50:
                performance = "Average"
            else:
                performance = "Needs Improvement"
        
        students_data.append({
            'student': student,
            'total_expected': total_expected,
            'total_paid': total_paid,
            'current_balance': current_balance,
            'historical_arrears': historical_arrears,
            'total_outstanding': total_outstanding,
            'results_count': results_count,
            'subject_count': subject_count,
            'performance': performance,
        })
    
    school = SchoolSetting.objects.first()
    
    # Get notifications/announcements
    from .models import Announcement
    announcements = Announcement.objects.filter(
        Q(target_audience='all') | Q(target_audience='parents'),
        Q(expires_at__isnull=True) | Q(expires_at__gt=timezone.now())
    ).order_by('-created_at')[:5]
    
    print(f"DEBUG: Final Total Fees Paid across all students: {total_fees_paid}")
    
    return render(request, "parent_portal/parent_dashboard.html", {
        "students_data": students_data,
        "students": students,
        "total_fees_paid": total_fees_paid,
        "total_results": total_results,
        "students_count": students.count(),
        "parent_name": parent_name or "Parent",
        "title": "Parent Dashboard",
        "school": school,
        "current_term": current_term,
        "announcements": announcements,
    })


def parent_fee_detail(request, student_id):
    """Parent view for detailed student fee information - REAL-TIME calculation"""
    from decimal import Decimal
    from .models import Student, Term, SchoolSetting, FeePayment, HistoricalArrears, FeeStructure, FeeBalance
    from django.db.models import Q, Sum
    
    phone = request.session.get("parent_phone")
    
    if not phone:
        messages.error(request, "Please login first.")
        return redirect("digitallibrary:parent_login")
    
    # Get student and verify it belongs to this parent
    try:
        student = Student.objects.get(
            Q(parent_phone=phone) | Q(parent_alternative_phone=phone),
            id=student_id,
            is_active=True
        )
    except Student.DoesNotExist:
        messages.error(request, "Student not found or not linked to your account.")
        return redirect("digitallibrary:parent_dashboard")
    
    # Get term from request or use current term
    academic_year = request.GET.get('academic_year')
    term_number = request.GET.get('term')
    
    if academic_year and term_number:
        term_number = int(term_number)
    else:
        current_term = Term.objects.filter(is_active=True).first()
        if current_term:
            academic_year = current_term.academic_year
            term_number = current_term.term_number
        else:
            academic_year = '2026'
            term_number = 1
    
    # Calculate fees for selected term - REAL TIME
    # Total expected from fee structures
    fee_structures = FeeStructure.objects.filter(
        student_class=student.current_class,
        academic_year=academic_year,
        term=term_number
    )
    total_expected = Decimal('0.00')
    for fs in fee_structures:
        total_expected += fs.total_fees
    
    # Total paid from fee payments
    total_paid = FeePayment.objects.filter(
        student=student,
        academic_year=academic_year,
        term=term_number
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    current_balance = total_expected - total_paid
    
    # Historical arrears
    historical_arrears = HistoricalArrears.objects.filter(
        student=student,
        is_settled=False
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    total_outstanding = current_balance + historical_arrears
    
    # Get payment history
    payments = FeePayment.objects.filter(
        student=student,
        academic_year=academic_year,
        term=term_number
    ).order_by('-payment_date')
    
    # Get fee breakdown by component
    fee_breakdown = []
    for fs in fee_structures:
        # Get individual components if any
        components = fs.custom_fees.all()
        if components.exists():
            for component in components:
                fee_breakdown.append({
                    'name': component.name,
                    'amount': component.amount
                })
        else:
            fee_breakdown.append({
                'name': f"Term {fs.term} Fees",
                'amount': fs.total_fees
            })
    
    # Get all available terms for filtering
    terms = Term.objects.all().order_by('-academic_year', '-term_number')
    
    school = SchoolSetting.objects.first()
    
    # Update or create FeeBalance for consistency
    FeeBalance.objects.update_or_create(
        student=student,
        academic_year=academic_year,
        term=term_number,
        defaults={
            'total_expected': total_expected,
            'total_paid': total_paid,
            'balance': current_balance,
            'status': 'PAID' if current_balance == 0 else 'PARTIAL' if total_paid > 0 else 'DEFAULTING'
        }
    )
    
    return render(request, "parent_portal/fee_detail.html", {
        "student": student,
        "total_expected": total_expected,
        "total_paid": total_paid,
        "current_balance": current_balance,
        "historical_arrears": historical_arrears,
        "total_outstanding": total_outstanding,
        "payments": payments,
        "fee_breakdown": fee_breakdown,
        "academic_year": academic_year,
        "term_number": term_number,
        "terms": terms,
        "school": school,
    })
def parent_student_detail(request, student_id):
    """View details for a specific student"""
    from .models import Student, FeeBalance, FeePayment, StudentResult, SchoolSetting
    
    phone = request.session.get("parent_phone")
    
    if not phone:
        messages.error(request, "Please login first.")
        return redirect("digitallibrary:parent_login")
    
    # Verify parent has access to this student
    students = Student.objects.filter(
        Q(parent_phone=phone) | Q(parent_alternative_phone=phone),
        is_active=True
    )
    student = get_object_or_404(students, id=student_id)
    
    fee_balances = FeeBalance.objects.filter(student=student).order_by("-academic_year", "-term")
    payments = FeePayment.objects.filter(student=student).order_by("-payment_date", "-created_at")[:10]
    results = StudentResult.objects.filter(student=student).select_related("exam", "subject").order_by(
        "-exam__academic_year", "-exam__term", "subject__name"
    )[:20]
    
    school = SchoolSetting.objects.first()
    
    return render(request, "parent_portal/parent_student_detail.html", {
        "student": student,
        "fee_balances": fee_balances,
        "payments": payments,
        "results": results,
        "title": student.get_full_name(),
        "school": school,
    })


def parent_fee_statement(request, student_id):
    """View fee statement for a student"""
    from .models import Student, FeeBalance, FeePayment, SchoolSetting
    
    phone = request.session.get("parent_phone")
    
    if not phone:
        messages.error(request, "Please login first.")
        return redirect("digitallibrary:parent_login")
    
    students = Student.objects.filter(
        Q(parent_phone=phone) | Q(parent_alternative_phone=phone),
        is_active=True
    )
    student = get_object_or_404(students, id=student_id)
    
    fee_balances = FeeBalance.objects.filter(student=student).order_by("-academic_year", "-term")
    payments = FeePayment.objects.filter(student=student).order_by("-payment_date", "-created_at")
    
    school = SchoolSetting.objects.first()
    
    return render(request, "parent_portal/parent_fee_statement.html", {
        "student": student,
        "fee_balances": fee_balances,
        "payments": payments,
        "title": "Fee Statement",
        "school": school,
    })


def parent_results(request, student_id):
    """View results for a student"""
    from .models import Student, StudentResult, SchoolSetting
    
    phone = request.session.get("parent_phone")
    
    if not phone:
        messages.error(request, "Please login first.")
        return redirect("digitallibrary:parent_login")
    
    students = Student.objects.filter(
        Q(parent_phone=phone) | Q(parent_alternative_phone=phone),
        is_active=True
    )
    student = get_object_or_404(students, id=student_id)
    
    results = StudentResult.objects.filter(student=student).select_related("exam", "subject").order_by(
        "-exam__academic_year", "-exam__term", "subject__name"
    )
    
    school = SchoolSetting.objects.first()
    
    return render(request, "parent_portal/parent_results.html", {
        "student": student,
        "results": results,
        "title": "Results",
        "school": school,
    })


def parent_pay_fees(request, student_id):
    """Pay fees for a student"""
    from .models import Student, SchoolSetting
    
    phone = request.session.get("parent_phone")
    
    if not phone:
        messages.error(request, "Please login first.")
        return redirect("digitallibrary:parent_login")
    
    students = Student.objects.filter(
        Q(parent_phone=phone) | Q(parent_alternative_phone=phone),
        is_active=True
    )
    student = get_object_or_404(students, id=student_id)
    
    messages.info(request, "M-PESA payment will be connected in the next phase.")
    
    school = SchoolSetting.objects.first()
    
    return render(request, "parent_portal/parent_pay_fees.html", {
        "student": student,
        "title": "Pay Fees",
        "school": school,
    })


def parent_view_grades(request):
    """Show all grades for parent's children"""
    from .models import Student, StudentResult, SchoolSetting
    
    phone = request.session.get("parent_phone")
    
    if not phone:
        messages.error(request, "Please login first.")
        return redirect("digitallibrary:parent_login")
    
    children = Student.objects.filter(
        Q(parent_phone=phone) | Q(parent_alternative_phone=phone),
        is_active=True
    )
    
    school = SchoolSetting.objects.first()
    
    return render(request, 'digitallibrary/parent_grades.html', {
        'children': children,
        'school': school,
    })


def parent_view_attendance(request):
    """Show attendance records for parent's children"""
    from .models import Student, SchoolSetting
    
    phone = request.session.get("parent_phone")
    
    if not phone:
        messages.error(request, "Please login first.")
        return redirect("digitallibrary:parent_login")
    
    children = Student.objects.filter(
        Q(parent_phone=phone) | Q(parent_alternative_phone=phone),
        is_active=True
    )
    
    school = SchoolSetting.objects.first()
    
    return render(request, 'digitallibrary/parent_attendance.html', {
        'children': children,
        'school': school,
    })


def parent_fee_balance(request):
    """Show fee balance for parent's children"""
    from .models import Student, FeeBalance, SchoolSetting
    
    phone = request.session.get("parent_phone")
    
    if not phone:
        messages.error(request, "Please login first.")
        return redirect("digitallibrary:parent_login")
    
    children = Student.objects.filter(
        Q(parent_phone=phone) | Q(parent_alternative_phone=phone),
        is_active=True
    )
    
    school = SchoolSetting.objects.first()
    
    return render(request, 'digitallibrary/parent_fee.html', {
        'children': children,
        'school': school,
    })
# ========== PARENT RESEND OTP VIEW ==========

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from .models import ParentOTP, Student
import random
from datetime import timedelta
from django.utils import timezone

@csrf_exempt
@require_POST
def parent_resend_otp(request):
    """Resend OTP to parent phone number"""
    import json
    
    try:
        data = json.loads(request.body)
        phone = data.get('phone', '')
        
        if not phone:
            return JsonResponse({'success': False, 'error': 'Phone number is required'})
        
        # Clean phone number
        phone = phone.strip().replace(' ', '')
        if phone.startswith('+254'):
            phone = '0' + phone[4:]
        elif phone.startswith('254'):
            phone = '0' + phone[3:]
        
        # Check if any student has this phone number
        students = Student.objects.filter(
            Q(parent_phone=phone) | Q(parent_alternative_phone=phone),
            is_active=True
        )
        
        if not students.exists():
            return JsonResponse({'success': False, 'error': 'No student found with this phone number'})
        
        # Generate new OTP
        otp_code = str(random.randint(100000, 999999))
        expires_at = timezone.now() + timedelta(minutes=10)
        
        # Delete old unused OTPs for this phone
        ParentOTP.objects.filter(phone=phone, is_used=False).delete()
        
        # Create new OTP
        ParentOTP.objects.create(
            phone=phone,
            otp_code=otp_code,
            expires_at=expires_at,
            is_used=False
        )
        
        # Send SMS with OTP
        message = f"Your ShuleHub Parent Portal verification code is: {otp_code}"
        
        if settings.MOCK_SMS_MODE:
            # Mock mode - return OTP for testing
            return JsonResponse({
                'success': True,
                'message': 'OTP sent successfully (TEST MODE)',
                'otp': otp_code
            })
        else:
            # Real SMS
            from .sms_utils import send_sms
            result = send_sms(phone, message)
            if result['success']:
                return JsonResponse({
                    'success': True,
                    'message': 'Verification code resent successfully'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': result.get('error', 'Failed to send SMS')
                })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid request'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# Also add the parent_session_required decorator if missing
def parent_session_required(view_func):
    """Decorator to ensure parent session exists"""
    def wrapper(request, *args, **kwargs):
        if not request.session.get("parent_phone"):
            messages.error(request, "Please login first.")
            return redirect("digitallibrary:parent_login")
        return view_func(request, *args, **kwargs)
    return wrapper
# digitallibrary/views_school.py

from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from .models import SchoolSetting

@staff_member_required
def school_settings_view(request):
    """Custom view to update school settings including logo"""
    setting = SchoolSetting.objects.first()
    if not setting:
        setting = SchoolSetting()
    
    if request.method == 'POST':
        # Update text fields
        setting.name = request.POST.get('name', '')
        setting.motto = request.POST.get('motto', '')
        setting.phone = request.POST.get('phone', '')
        setting.email = request.POST.get('email', '')
        setting.address = request.POST.get('address', '')
        setting.website = request.POST.get('website', '')
        
        # Handle logo upload
        if request.FILES.get('logo'):
            setting.logo = request.FILES['logo']
        
        setting.save()
        messages.success(request, 'School settings updated successfully!')
        return redirect('digitallibrary:school_settings')
    
    return render(request, 'digitallibrary/school_settings.html', {'setting': setting})
# =========================
# GRADING SYSTEM VIEWS
# =========================

@staff_member_required
def set_grading_preference(request, exam_id):
    """Set the grading system preference for this exam session"""
    from django.shortcuts import redirect
    from django.contrib import messages
    from .models import GradingSystem, TeacherGradingPreference
    
    print(f"\n{'='*60}")
    print(f"🔧 set_grading_preference called")
    print(f"   exam_id: {exam_id}")
    print(f"   Method: {request.method}")
    print(f"   POST params: {dict(request.POST)}")
    print(f"{'='*60}")
    
    if request.method == 'POST':
        grading_system_id = request.POST.get('grading_system_id')
        subject_id = request.GET.get('subject')
        
        print(f"   grading_system_id: '{grading_system_id}'")
        print(f"   subject_id: '{subject_id}'")
        
        # Get or create teacher preference
        preference, created = TeacherGradingPreference.objects.get_or_create(
            teacher=request.user,
            exam_id=exam_id
        )
        
        if grading_system_id == 'cbe':
            # Use CBE grading
            preference.use_cbe_pathways = True
            preference.use_custom_grading = False
            preference.custom_grading_system = None
            request.session['active_grading_system_id'] = 'cbe'
            messages.success(request, '✓ CBE Grading System Activated (EE1, EE2, ME1, ME2, AE2, AE1, BE2, BE1)')
            print(f"   Set session: active_grading_system_id = 'cbe'")
            
        elif grading_system_id == 'traditional':
            # Use traditional grading
            preference.use_cbe_pathways = False
            preference.use_custom_grading = False
            preference.custom_grading_system = None
            request.session['active_grading_system_id'] = None
            messages.success(request, '✓ Traditional Grading System (KCSE) Activated')
            print(f"   Set session: active_grading_system_id = None")
            
        elif grading_system_id:
            try:
                # Try to get custom grading system by ID
                grading_system = GradingSystem.objects.get(id=int(grading_system_id), is_active=True)
                preference.use_cbe_pathways = False
                preference.use_custom_grading = True
                preference.custom_grading_system = grading_system
                request.session['active_grading_system_id'] = grading_system.id
                messages.success(request, f'✓ {grading_system.name} Grading System Activated')
                print(f"   Set session: active_grading_system_id = {grading_system.id}")
            except (GradingSystem.DoesNotExist, ValueError) as e:
                messages.error(request, f'Selected grading system not found')
                print(f"   ERROR: {e}")
        else:
            messages.error(request, 'Please select a grading system')
            print(f"   ERROR: No grading_system_id provided")
        
        preference.save()
        
        # Redirect back to the results entry form
        if subject_id and subject_id != 'None':
            redirect_url = f'/app/enter-results-form/?exam={exam_id}&subject={subject_id}'
        else:
            redirect_url = f'/app/enter-results-form/?exam={exam_id}'
        
        print(f"   Redirecting to: {redirect_url}")
        print(f"{'='*60}\n")
        return redirect(redirect_url)
    
    # For GET requests, just redirect to the form
    subject_id = request.GET.get('subject')
    if subject_id and subject_id != 'None':
        return redirect(f'/app/enter-results-form/?exam={exam_id}&subject={subject_id}')
    return redirect(f'/app/enter-results-form/?exam={exam_id}')

@login_required
def add_grading_scales(request, system_id):
    """Add grading scales to a custom grading system"""
    
    grading_system = get_object_or_404(GradingSystem, id=system_id, created_by=request.user)
    
    if request.method == 'POST':
        # Process grading scales
        grades = request.POST.getlist('grade')
        min_scores = request.POST.getlist('min_score')
        max_scores = request.POST.getlist('max_score')
        points = request.POST.getlist('points')
        remarks = request.POST.getlist('remark')
        
        # Delete existing grades
        grading_system.grades.all().delete()
        
        # Create new grades
        for i in range(len(grades)):
            if grades[i] and min_scores[i] and max_scores[i]:
                GradeScale.objects.create(
                    grading_system=grading_system,
                    grade=grades[i],
                    min_score=min_scores[i],
                    max_score=max_scores[i],
                    points=points[i] if points[i] else 0,
                    remark=remarks[i] if remarks[i] else ''
                )
        
        messages.success(request, f'Grading scales added to {grading_system.name}')
        return redirect('digitallibrary:set_grading_preference')
    
    # Default grade suggestions
    default_grades = [
        ('A', 80, 100, 12, 'Excellent'),
        ('B', 70, 79, 9, 'Good'),
        ('C', 60, 69, 6, 'Average'),
        ('D', 50, 59, 3, 'Below Average'),
        ('E', 0, 49, 1, 'Fail'),
    ]
    
    context = {
        'grading_system': grading_system,
        'default_grades': default_grades,
    }
    return render(request, 'digitallibrary/add_grading_scales.html', context)


def get_grade_for_score(score, exam=None, subject=None, student=None):
    """Get grade based on exam, subject, and student's CBE pathway if applicable"""
    
    # Check if student is in CBE pathway
    if student and hasattr(student, 'pathway') and student.pathway:
        try:
            cbe_pathway = CBEGradingPathway.objects.filter(
                pathway_type=student.pathway,
                is_active=True
            ).first()
            if cbe_pathway and cbe_pathway.grading_system:
                grade_obj = cbe_pathway.grading_system.grades.filter(
                    min_score__lte=score,
                    max_score__gte=score
                ).first()
                if grade_obj:
                    return grade_obj
        except:
            pass
    
    # Check for custom teacher grading
    if exam and subject:
        preference = TeacherGradingPreference.objects.filter(
            exam=exam,
            subject=subject
        ).first()
        
        if preference and preference.use_custom_grading and preference.custom_grading_system:
            grade_obj = preference.custom_grading_system.grades.filter(
                min_score__lte=score,
                max_score__gte=score
            ).first()
            if grade_obj:
                return grade_obj
    
    # Default grading system
    return Grade.objects.filter(
        min_score__lte=score,
        max_score__gte=score
    ).first()
from django.http import JsonResponse
from .models import Student, Exam, StudentResult, TeacherGradingPreference, CBEGradingPathway, GradeScale

@login_required
def api_calculate_grade(request):
    """API endpoint to calculate grade based on student's pathway and selected grading system"""
    if request.method == 'POST':
        data = json.loads(request.body)
        student_id = data.get('student_id')
        score = data.get('score')
        exam_id = data.get('exam_id')
        
        try:
            student = Student.objects.get(id=student_id)
            exam = Exam.objects.get(id=exam_id)
            
            # Check for grading preference
            preference = TeacherGradingPreference.objects.filter(
                exam=exam,
                subject__isnull=True
            ).first()
            
            grading_system = None
            
            if preference and preference.use_cbe_pathways:
                # Use CBE grading
                if student.pathway:
                    pathway = CBEGradingPathway.objects.filter(
                        pathway_type=student.pathway,
                        is_active=True
                    ).first()
                    if pathway:
                        grading_system = pathway.grading_system
            elif preference and preference.use_custom_grading:
                grading_system = preference.custom_grading_system
            
            # Calculate grade
            if grading_system:
                grade_scale = grading_system.grades.filter(
                    min_score__lte=score,
                    max_score__gte=score
                ).first()
                if grade_scale:
                    return JsonResponse({
                        'grade': grade_scale.grade,
                        'points': grade_scale.points,
                        'remark': grade_scale.remark,
                        'system': grading_system.system_type
                    })
            
            # Fallback to traditional grading
            percentage = (score / exam.max_score) * 100 if exam.max_score else score
            if percentage >= 80:
                return JsonResponse({'grade': 'A', 'points': 12})
            elif percentage >= 75:
                return JsonResponse({'grade': 'A-', 'points': 11})
            elif percentage >= 70:
                return JsonResponse({'grade': 'B+', 'points': 10})
            elif percentage >= 65:
                return JsonResponse({'grade': 'B', 'points': 9})
            elif percentage >= 60:
                return JsonResponse({'grade': 'B-', 'points': 8})
            elif percentage >= 55:
                return JsonResponse({'grade': 'C+', 'points': 7})
            elif percentage >= 50:
                return JsonResponse({'grade': 'C', 'points': 6})
            elif percentage >= 45:
                return JsonResponse({'grade': 'C-', 'points': 5})
            elif percentage >= 40:
                return JsonResponse({'grade': 'D+', 'points': 4})
            elif percentage >= 35:
                return JsonResponse({'grade': 'D', 'points': 3})
            elif percentage >= 30:
                return JsonResponse({'grade': 'D-', 'points': 2})
            else:
                return JsonResponse({'grade': 'E', 'points': 1})
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def generate_receipt(request, payment_id):
    """Generate PDF receipt for a payment"""
    payment = get_object_or_404(FeePayment, id=payment_id)
    
    # Check permission (admin, bursar, or the student's parent)
    if not (request.user.is_staff or 
            request.user.profile.role in ['admin', 'principal', 'bursar'] or
            (request.user.profile.role == 'parent' and payment.student in request.user.profile.children.all())):
        messages.error(request, 'You do not have permission to view this receipt')
        return redirect('digitallibrary:fees_dashboard')
    
    # Get school settings
    school = SchoolSetting.objects.first()
    
    # Calculate balance after this payment
    total_paid = FeePayment.objects.filter(student=payment.student).aggregate(
        total=models.Sum('amount')
    )['total'] or Decimal('0.00')
    
    total_fees = FeeStructure.objects.filter(
        student_class=payment.student.current_class,
        academic_year=payment.academic_year,
        term=payment.term
    ).aggregate(total=models.Sum('total_fees'))['total'] or Decimal('0.00')
    
    balance = total_fees - total_paid
    
    context = {
        'payment': payment,
        'student': payment.student,
        'school': school,
        'total_paid': total_paid,
        'total_fees': total_fees,
        'balance': balance,
        'generated_date': timezone.now(),
        'generated_by': request.user,
    }
    
    # Render HTML template
    template = get_template('digitallibrary/receipt_template.html')
    html = template.render(context, request)
    
    # Create PDF
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        # Mark receipt as generated
        payment.receipt_generated = True
        payment.save()
        
        # Create HTTP response with PDF
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="receipt_{payment.receipt_number}.pdf"'
        return response
    
    return HttpResponse('Error generating PDF', status=500)


@login_required
def download_receipt(request, payment_id):
    """Download PDF receipt"""
    payment = get_object_or_404(FeePayment, id=payment_id)
    
    # Check permission
    if not (request.user.is_staff or 
            request.user.profile.role in ['admin', 'principal', 'bursar'] or
            (request.user.profile.role == 'parent' and payment.student in request.user.profile.children.all())):
        messages.error(request, 'You do not have permission to download this receipt')
        return redirect('digitallibrary:fees_dashboard')
    
    school = SchoolSetting.objects.first()
    
    total_paid = FeePayment.objects.filter(student=payment.student).aggregate(
        total=models.Sum('amount')
    )['total'] or Decimal('0.00')
    
    total_fees = FeeStructure.objects.filter(
        student_class=payment.student.current_class,
        academic_year=payment.academic_year,
        term=payment.term
    ).aggregate(total=models.Sum('total_fees'))['total'] or Decimal('0.00')
    
    balance = total_fees - total_paid
    
    context = {
        'payment': payment,
        'student': payment.student,
        'school': school,
        'total_paid': total_paid,
        'total_fees': total_fees,
        'balance': balance,
        'generated_date': timezone.now(),
        'generated_by': request.user,
    }
    
    template = get_template('digitallibrary/receipt_template.html')
    html = template.render(context, request)
    
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="receipt_{payment.receipt_number}.pdf"'
        return response
    
    return HttpResponse('Error generating PDF', status=500)
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from decimal import Decimal
from .models import Student, HistoricalArrears, Class, FeeBalance

@login_required
def add_historical_arrears(request):
    """Add historical arrears for a student"""
    
    # Initialize context
    context = {
        'student': None,
        'current_balance': 0,
        'classes': Class.objects.all().order_by('name'),
        'admission_no': '',
        'searched': False,
    }
    
    # Handle GET request - search for student
    if request.method == 'GET' and 'admission_no' in request.GET:
        admission_no = request.GET.get('admission_no')
        context['admission_no'] = admission_no
        context['searched'] = True
        
        if admission_no:
            try:
                # Search for student by admission number
                student = Student.objects.get(admission_number=admission_no)
                context['student'] = student
                
                # Get current balance for the student (for latest term)
                latest_term = Term.objects.filter(is_active=True).first()
                if latest_term:
                    current_balance = student.get_fee_balance(latest_term.academic_year, latest_term.term_number)
                    context['current_balance'] = current_balance
                else:
                    context['current_balance'] = 0
                    
                messages.info(request, f'Student found: {student.first_name} {student.last_name}')
                
            except Student.DoesNotExist:
                messages.error(request, f'No student found with admission number: {admission_no}')
                context['student'] = None
    
    # Handle POST request - save historical arrears
    elif request.method == 'POST':
        student_id = request.POST.get('student_id')
        amount = request.POST.get('amount')
        original_academic_year = request.POST.get('original_academic_year')
        original_class_id = request.POST.get('original_class_id')
        original_term = request.POST.get('original_term')
        notes = request.POST.get('notes', '')
        
        # Validate required fields
        if not all([student_id, amount, original_academic_year, original_class_id, original_term]):
            messages.error(request, 'Please fill in all required fields')
            return redirect('digitallibrary:add_historical_arrears')
        
        try:
            student = Student.objects.get(id=student_id)
            amount = Decimal(str(amount))
            original_class = Class.objects.get(id=original_class_id)
            original_term = int(original_term)
            
            # Create historical arrears record
            historical_arrear = HistoricalArrears.objects.create(
                student=student,
                amount=amount,
                original_class=original_class,
                original_academic_year=original_academic_year,
                original_term=original_term,
                notes=notes,
                added_by=request.user,
                is_settled=False
            )
            
            messages.success(request, f'Successfully added KES {amount:,.2f} historical arrears for {student.first_name} {student.last_name}')
            return redirect('digitallibrary:student_fee_detail', student_id=student.id)
            
        except Student.DoesNotExist:
            messages.error(request, 'Student not found')
        except Class.DoesNotExist:
            messages.error(request, 'Selected class not found')
        except Exception as e:
            messages.error(request, f'Error adding arrears: {str(e)}')
    
    return render(request, 'digitallibrary/fees/add_historical_arrears.html', context)

@login_required
def student_fee_detail(request, student_id):
    """Display comprehensive fee details for a student"""
    student = get_object_or_404(Student, id=student_id)
    
    # Get current academic year and term from request
    current_year = request.GET.get('academic_year')
    current_term = request.GET.get('term')
    
    # If not specified, get the latest active term
    if not current_year or not current_term:
        latest_term = Term.objects.filter(is_active=True).first()
        if latest_term:
            current_year = latest_term.academic_year
            current_term = latest_term.term_number
        else:
            # Fallback to most recent term
            latest_term = Term.objects.order_by('-academic_year', '-term_number').first()
            if latest_term:
                current_year = latest_term.academic_year
                current_term = latest_term.term_number
            else:
                # If no terms exist, set defaults
                current_year = '2026'
                current_term = 1
    
    # Ensure term is integer
    try:
        current_term = int(current_term)
    except (TypeError, ValueError):
        current_term = 1
    
    # Calculate fee summary for selected term
    total_expected = student.get_total_fees_expected(current_year, current_term)
    total_paid = student.get_total_fees_paid(current_year, current_term)
    current_balance = total_expected - total_paid
    
    # Convert to Decimal for consistency
    if isinstance(current_balance, float):
        current_balance = Decimal(str(current_balance))
    
    # Get or create fee balance object
    fee_balance = None
    try:
        fee_balance = FeeBalance.objects.get(
            student=student,
            academic_year=current_year,
            term=current_term
        )
        # Update balance if needed
        if fee_balance.balance != current_balance:
            fee_balance.balance = current_balance
            fee_balance.total_expected = total_expected
            fee_balance.total_paid = total_paid
            fee_balance.save()
    except FeeBalance.DoesNotExist:
        # Create it with proper values
        fee_balance = FeeBalance.objects.create(
            student=student,
            academic_year=current_year,
            term=current_term,
            total_expected=total_expected,
            total_paid=total_paid,
            balance=current_balance,
            status='OVERPAID' if current_balance < 0 else 'PARTIAL' if total_paid > 0 else 'DEFAULTING'
        )
    
    # Get historical arrears (unsettled only)
    historical_arrears = HistoricalArrears.objects.filter(
        student=student, 
        is_settled=False
    )
    total_historical_arrears = historical_arrears.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    # Ensure total_historical_arrears is Decimal
    if not isinstance(total_historical_arrears, Decimal):
        total_historical_arrears = Decimal(str(total_historical_arrears))
    
    # Total including historical arrears (both are now Decimal)
    total_outstanding = current_balance + total_historical_arrears
    
    # Get payment history for selected term/year
    payments = student.get_payment_history(current_year, current_term)
    
    # Get all available terms for filtering
    terms = Term.objects.all().order_by('-academic_year', '-term_number')
    
    # Get all fee structures for this student's class
    fee_structures = FeeStructure.objects.filter(
        student_class=student.current_class
    ).order_by('-academic_year', '-term')
    
    # Debug print (remove in production)
    print(f"Student: {student.first_name} {student.last_name}")
    print(f"Year: {current_year}, Term: {current_term}")
    print(f"Expected: {total_expected}, Paid: {total_paid}, Balance: {current_balance}")
    print(f"Historical Arrears Count: {historical_arrears.count()}, Total: {total_historical_arrears}")
    print(f"Total Outstanding: {total_outstanding}")
    
    context = {
        'student': student,
        'total_expected': total_expected,
        'total_paid': total_paid,
        'current_balance': current_balance,
        'total_historical_arrears': total_historical_arrears,
        'total_outstanding': total_outstanding,
        'payments': payments,
        'historical_arrears': historical_arrears,
        'fee_balance': fee_balance,
        'current_year': current_year,
        'current_term': current_term,
        'terms': terms,
        'fee_structures': fee_structures,
    }
    
    return render(request, 'digitallibrary/student_fee_detail.html', context)
def calculate_grade(percentage, grading_system='both'):
    """
    Calculate grade based on either Traditional or CBC system.
    
    Args:
        percentage: The score percentage (0-100)
        grading_system: 'traditional', 'cbc', or 'both' (returns both formats)
    
    Returns:
        Dictionary with grades from both systems or specified system
    """
    
    # TRADITIONAL 8-4-4 GRADING SYSTEM
    if percentage >= 80:
        traditional = ('A', 'Excellent', 12, 'PASS')
    elif percentage >= 75:
        traditional = ('A-', 'Very Good', 11, 'PASS')
    elif percentage >= 70:
        traditional = ('B+', 'Good', 10, 'PASS')
    elif percentage >= 65:
        traditional = ('B', 'Above Average', 9, 'PASS')
    elif percentage >= 60:
        traditional = ('B-', 'Average', 8, 'PASS')
    elif percentage >= 55:
        traditional = ('C+', 'Satisfactory', 7, 'PASS')
    elif percentage >= 50:
        traditional = ('C', 'Acceptable', 6, 'PASS')
    elif percentage >= 45:
        traditional = ('C-', 'Below Average', 5, 'PASS')
    elif percentage >= 40:
        traditional = ('D+', 'Weak', 4, 'PASS')
    elif percentage >= 35:
        traditional = ('D', 'Very Weak', 3, 'FAIL')
    elif percentage >= 30:
        traditional = ('D-', 'Poor', 2, 'FAIL')
    else:
        traditional = ('E', 'Very Poor', 1, 'FAIL')
    
    # CBC/COMPETENCY-BASED GRADING SYSTEM
    if percentage >= 90:
        cbc = ('EE1', 'Exceptional/Excellent', 8, 'PASS')
    elif percentage >= 75:
        cbc = ('EE2', 'Very Good', 7, 'PASS')
    elif percentage >= 58:
        cbc = ('ME1', 'Good', 6, 'PASS')
    elif percentage >= 41:
        cbc = ('ME2', 'Fair', 5, 'PASS')
    elif percentage >= 31:
        cbc = ('AE1', 'Needs Improvement', 4, 'FAIL')
    elif percentage >= 21:
        cbc = ('AE2', 'Below Average', 3, 'FAIL')
    elif percentage >= 11:
        cbc = ('BE1', 'Well Below Average', 2, 'FAIL')
    elif percentage >= 1:
        cbc = ('BE2', 'Minimal', 1, 'FAIL')
    else:
        cbc = ('BE2', 'Minimal', 0, 'FAIL')
    
    if grading_system == 'traditional':
        return {
            'grade_letter': traditional[0],
            'grade_description': traditional[1],
            'points': traditional[2],
            'status': traditional[3],
            'system': 'Traditional (8-4-4)'
        }
    elif grading_system == 'cbc':
        return {
            'grade_letter': cbc[0],
            'grade_description': cbc[1],
            'points': cbc[2],
            'status': cbc[3],
            'system': 'CBC/CBE'
        }
    else:  # 'both' - return both grading systems
        return {
            'traditional': {
                'grade_letter': traditional[0],
                'grade_description': traditional[1],
                'points': traditional[2],
                'status': traditional[3]
            },
            'cbc': {
                'grade_letter': cbc[0],
                'grade_description': cbc[1],
                'points': cbc[2],
                'status': cbc[3]
            }
        }


def get_school_grading_system(request):
    """Determine which grading system the school uses"""
    # You can store this in your School/Tenant model
    if hasattr(request.tenant, 'grading_system'):
        return request.tenant.grading_system  # 'traditional', 'cbc', or 'both'
    return 'traditional'  # Default to traditional
def generate_report_card(request, exam_id=None, student_id=None):
    """
    Generate a professional report card for a student's exam results.
    Usage: /app/report-card/?exam=1&student=5
    """
    from .models import Exam, Student, StudentResult, SchoolSetting
    from decimal import Decimal
    from django.utils import timezone
    from django.contrib import messages  # ← ADD THIS
    
    # Get exam and student from request
    exam_id = request.GET.get('exam') or exam_id
    student_id = request.GET.get('student') or student_id
    
    if not exam_id or not student_id:
        messages.error(request, "Exam and Student are required to generate report card")
        return redirect('digitallibrary:performance_dashboard')
    
    try:
        exam = Exam.objects.get(id=exam_id)
        student = Student.objects.get(id=student_id, is_active=True)
    except Exam.DoesNotExist:
        messages.error(request, "Exam not found")
        return redirect('digitallibrary:performance_dashboard')
    except Student.DoesNotExist:
        messages.error(request, "Student not found")
        return redirect('digitallibrary:performance_dashboard')
    
    # Get results for this student and exam
    results = StudentResult.objects.filter(
        student=student,
        exam=exam
    ).select_related('subject')
    
    if not results.exists():
        messages.warning(request, f"No results found for {student.get_full_name()} in {exam.name}")
        return redirect('digitallibrary:student_performance', student_id=student.id)
    
    subject_results = []
    strengths = []
    weaknesses = []
    total_points = 0
    
    # Get grading system from school
    grading_system = 'traditional'
    if hasattr(request.tenant, 'grading_system'):
        grading_system = request.tenant.grading_system
    
    for result in results:
        # Calculate percentage
        max_score = float(exam.max_score) if exam.max_score else 100.0
        percentage = (float(result.score) / max_score) * 100
        
        # Get grade using the calculate_grade function
        grades = calculate_grade(percentage, grading_system)
        
        if grading_system == 'both':
            grade_display = f"{grades['traditional']['grade_letter']} / {grades['cbc']['grade_letter']}"
            grade_desc = f"{grades['traditional']['grade_description']} / {grades['cbc']['grade_description']}"
            status = grades['traditional']['status']
            points = grades['traditional']['points']
        else:
            grade_display = grades['grade_letter']
            grade_desc = grades['grade_description']
            status = grades['status']
            points = grades['points']
        
        total_points += points
        
        # Get teacher comment
        teacher_comment = get_teacher_comment(result.subject.name, percentage)
        
        subject_data = {
            'subject': result.subject,
            'score': float(result.score),
            'max_score': max_score,
            'percentage': round(percentage, 1),
            'grade_letter': grade_display,
            'grade_description': grade_desc,
            'status': status,
            'points': points,
            'teacher_comment': teacher_comment,
        }
        
        subject_results.append(subject_data)
        
        # Identify strengths and weaknesses
        if percentage >= 70:
            strengths.append(subject_data)
        elif percentage < 50:
            weaknesses.append(subject_data)
    
    # Calculate overall statistics
    mean_percentage = sum(r['percentage'] for r in subject_results) / len(subject_results)
    overall_grade_info = calculate_grade(mean_percentage, grading_system)
    
    if grading_system == 'both':
        overall_grade_display = f"{overall_grade_info['traditional']['grade_letter']} / {overall_grade_info['cbc']['grade_letter']}"
    else:
        overall_grade_display = overall_grade_info['grade_letter']
    
    # Generate overall remarks
    if mean_percentage >= 80:
        teacher_remarks = f"Excellent performance! {student.first_name} has shown exceptional understanding across all subjects. Keep up the great work!"
    elif mean_percentage >= 70:
        teacher_remarks = f"Very good performance. {student.first_name} is doing well. With a little more effort, can achieve even better results."
    elif mean_percentage >= 60:
        teacher_remarks = f"Good performance. {student.first_name} has a solid understanding. Focus on improving in the weaker areas."
    elif mean_percentage >= 50:
        if weaknesses:
            weak_subjects = ', '.join([w['subject'].name for w in weaknesses[:2]])
            teacher_remarks = f"Satisfactory performance. {student.first_name} needs to put more effort into {weak_subjects}."
        else:
            teacher_remarks = f"Satisfactory performance. {student.first_name} is doing well but can improve further."
    elif mean_percentage >= 40:
        teacher_remarks = f"Below average performance. {student.first_name} needs significant improvement. Please attend remedial classes."
    else:
        teacher_remarks = f"Critical attention needed. {student.first_name} is struggling. Parent-teacher meeting is strongly recommended."
    
    # Get school info
    school = SchoolSetting.objects.first()
    
    context = {
        'student': student,
        'exam': exam,
        'subject_results': subject_results,
        'strengths': strengths,
        'weaknesses': weaknesses,
        'overall_stats': {
            'mean_score': round(mean_percentage, 1),
            'total_points': total_points,
            'total_subjects': len(subject_results),
            'overall_grade': overall_grade_display,
            'grading_system': grading_system.upper()
        },
        'teacher_remarks': teacher_remarks,
        'school': school,
        'school_name': school.name if school else 'ShuleHub',
        'report_date': timezone.now().strftime("%B %d, %Y"),
        'report_id': f"RPT-{exam.id}-{student.id}-{timezone.now().strftime('%Y%m%d')}",
    }
    
    return render(request, 'digitallibrary/professional_report_card.html', context)


def get_teacher_comment(subject_name, percentage):
    """Generate meaningful comment based on percentage"""
    if percentage >= 80:
        return f"🏆 EXCELLENT! {subject_name} is a strong subject. Keep up the great work!"
    elif percentage >= 70:
        return f"👍 VERY GOOD in {subject_name}. Aim for an A next time!"
    elif percentage >= 60:
        return f"📚 GOOD effort in {subject_name}. With more practice, you can score higher."
    elif percentage >= 50:
        return f"📖 SATISFACTORY in {subject_name}. Review your weak areas."
    elif percentage >= 40:
        return f"⚠️ FAIR performance in {subject_name}. Please consult the teacher."
    else:
        return f"❌ NEEDS IMPROVEMENT in {subject_name}. Extra classes recommended."  # ← ADD THIS        return f"❌ NEEDS IMPROVEMENT in {subject_name}. Extra classes recommended."
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.management import call_command
from django_tenants.utils import schema_context
from django.db import connection
from django.db.models import Count, Q
from django.http import JsonResponse
from django.utils import timezone

from tenants.models import School, Domain
# digitallibrary/views.py

from django.contrib.admin.views.decorators import staff_member_required
from .models import GradingSystem, GradeScale

@staff_member_required
def grading_system_list(request):
    """List all grading systems"""
    systems = GradingSystem.objects.all()
    return render(request, 'digitallibrary/grading/systems.html', {'systems': systems})

@staff_member_required
def grading_system_create(request):
    """Create a new grading system"""
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        
        system = GradingSystem.objects.create(
            name=name,
            description=description,
            created_by=request.user,
            is_default=not GradingSystem.objects.filter(is_default=True).exists()
        )
        
        messages.success(request, f'Grading system "{name}" created successfully!')
        return redirect('digitallibrary:grading_system_edit', system.id)
    
    return render(request, 'digitallibrary/grading/system_form.html')

@staff_member_required
def grading_system_edit(request, pk):
    """Edit grading system and its grade scales with subject assignment"""
    system = get_object_or_404(GradingSystem, id=pk)
    
    if request.method == 'POST':
        # Update basic info
        system.name = request.POST.get('name')
        system.description = request.POST.get('description')
        system.system_type = request.POST.get('system_type', system.system_type)
        system.passing_score = request.POST.get('passing_score', system.passing_score)
        system.is_active = request.POST.get('is_active') == 'on'
        
        # Handle default status
        if request.POST.get('set_default'):
            GradingSystem.objects.filter(school=system.school).update(is_default=False)
            system.is_default = True
        
        # Handle subject assignment
        subject_id = request.POST.get('subject')
        if subject_id:
            system.subject_id = subject_id
        else:
            system.subject = None
        
        system.is_subject_specific = request.POST.get('is_subject_specific') == 'on'
        system.save()
        
        # Handle applicable subjects (many-to-many)
        applicable_subjects = request.POST.getlist('applicable_subjects')
        if applicable_subjects:
            system.applicable_subjects.set(applicable_subjects)
        else:
            system.applicable_subjects.clear()
        
        # Handle grade scales
        grade_ids = request.POST.getlist('grade_id')
        grades = request.POST.getlist('grade[]')
        min_scores = request.POST.getlist('min_score[]')
        max_scores = request.POST.getlist('max_score[]')
        points = request.POST.getlist('points[]')
        remarks = request.POST.getlist('remark[]')
        
        # Update existing and create new grade scales
        existing_ids = []
        for i, grade in enumerate(grades):
            if grade and min_scores[i] and max_scores[i]:
                grade_scale, created = GradeScale.objects.update_or_create(
                    id=grade_ids[i] if i < len(grade_ids) and grade_ids[i] else None,
                    defaults={
                        'grading_system': system,
                        'grade': grade,
                        'min_score': min_scores[i],
                        'max_score': max_scores[i],
                        'points': points[i] if i < len(points) else 0,
                        'remark': remarks[i] if i < len(remarks) else '',
                        'is_active': True
                    }
                )
                existing_ids.append(grade_scale.id)
        
        # Delete removed grades
        GradeScale.objects.filter(grading_system=system).exclude(id__in=existing_ids).delete()
        
        messages.success(request, f'Grading system "{system.name}" updated successfully!')
        return redirect('digitallibrary:grading_system_edit', system.id)
    
    # Get all subjects for the dropdown
    from .models import Subject
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    
    context = {
        'system': system,
        'subjects': subjects,
    }
    return render(request, 'digitallibrary/grading/system_form.html', context)

@staff_member_required
def grading_system_delete(request, pk):
    """Delete a grading system"""
    system = get_object_or_404(GradingSystem, id=pk)
    
    if system.is_default:
        messages.error(request, 'Cannot delete the default grading system.')
    else:
        system.delete()
        messages.success(request, 'Grading system deleted successfully.')
    
    return redirect('digitallibrary:grading_system_list')
# digitallibrary/views.py

from .models import SMSLog, UserProfile

@staff_member_required
def sms_to_staff(request):
    """Send SMS to teachers and support staff"""
    if request.method == 'POST':
        recipient_type = request.POST.get('recipient_type')  # 'all', 'teachers', 'staff', 'specific'
        message = request.POST.get('message')
        selected_users = request.POST.getlist('users')
        
        # Get recipients based on type
        if recipient_type == 'all':
            users = User.objects.filter(profile__role__in=['teacher', 'bursar', 'secretary', 'admin'])
        elif recipient_type == 'teachers':
            users = User.objects.filter(profile__role='teacher')
        elif recipient_type == 'staff':
            users = User.objects.filter(profile__role__in=['bursar', 'secretary', 'admin'])
        elif recipient_type == 'specific' and selected_users:
            users = User.objects.filter(id__in=selected_users)
        else:
            users = []
        
        # Send SMS to each user
        from .sms_utils import send_sms
        
        sent_count = 0
        failed_count = 0
        
        for user in users:
            phone = user.profile.phone_number if hasattr(user, 'profile') else None
            if phone:
                success = send_sms(phone, message)
                SMSLog.objects.create(
                    recipient=phone,
                    recipient_name=user.get_full_name() or user.username,
                    message=message,
                    category='general',
                    status='sent' if success else 'failed',
                    sent_by=request.user
                )
                if success:
                    sent_count += 1
                else:
                    failed_count += 1
        
        messages.success(request, f'SMS sent to {sent_count} staff members. Failed: {failed_count}')
        return redirect('digitallibrary:sms_to_staff')
    
    # GET request - show form
    users = User.objects.filter(profile__role__in=['teacher', 'bursar', 'secretary', 'admin'])
    
    return render(request, 'digitallibrary/sms/sms_to_staff.html', {
        'users': users,
        'roles': ['teacher', 'bursar', 'secretary', 'admin']
    })
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import GradingPreferenceForm
from .models import TeacherGradingPreference, GradingSystem, CBEGradingPathway

@login_required
def teacher_grading_preference(request):
    """View for teachers to set their grading system preference"""
    
    # Get existing preference for this teacher (global or specific)
    preference = TeacherGradingPreference.objects.filter(
        teacher=request.user,
        is_global=True
    ).first()
    
    if request.method == 'POST':
        form = GradingPreferenceForm(request.POST, instance=preference)
        if form.is_valid():
            pref = form.save(commit=False)
            pref.teacher = request.user
            pref.is_global = True
            pref.save()
            messages.success(request, 'Your grading preference has been saved!')
            return redirect('digitallibrary:teacher_dashboard')
    else:
        form = GradingPreferenceForm(instance=preference)
    
    context = {
        'form': form,
        'grading_systems': GradingSystem.objects.filter(is_active=True),
        'cbe_pathways': CBEGradingPathway.objects.filter(is_active=True),
    }
    
    return render(request, 'digitallibrary/grading/teacher_preference.html', context)


@login_required
def exam_grading_preference(request, exam_id, subject_id=None):
    """Set grading preference for a specific exam/subject"""
    
    exam = get_object_or_404(Exam, id=exam_id)
    subject = None
    if subject_id:
        subject = get_object_or_404(Subject, id=subject_id)
    
    # Get existing preference
    preference = TeacherGradingPreference.objects.filter(
        teacher=request.user,
        exam=exam,
        subject=subject
    ).first()
    
    if request.method == 'POST':
        form = GradingPreferenceForm(request.POST, instance=preference)
        if form.is_valid():
            pref = form.save(commit=False)
            pref.teacher = request.user
            pref.exam = exam
            pref.subject = subject
            pref.is_global = False
            pref.save()
            messages.success(request, f'Grading preference saved for {exam.name}')
            return redirect('digitallibrary:enter_results_form')
    else:
        form = GradingPreferenceForm(instance=preference)
    
    context = {
        'form': form,
        'exam': exam,
        'subject': subject,
        'grading_systems': GradingSystem.objects.filter(is_active=True),
        'cbe_pathways': CBEGradingPathway.objects.filter(is_active=True),
    }
    
    return render(request, 'digitallibrary/grading/exam_preference.html', context)
# digitallibrary/views.py

from django.contrib.admin.views.decorators import staff_member_required
from .models import SubjectGradingConfig, Subject, GradingSystem, Term

@staff_member_required
def subject_grading_list(request, subject_id=None):
    """List grading configurations for subjects"""
    if subject_id:
        subject = get_object_or_404(Subject, id=subject_id)
        configs = SubjectGradingConfig.objects.filter(subject=subject)
        template = 'digitallibrary/grading/subject_configs.html'
        context = {
            'subject': subject,
            'configs': configs,
            'title': f'Grading Configurations - {subject.name}'
        }
    else:
        configs = SubjectGradingConfig.objects.all().select_related('subject', 'grading_system')
        template = 'digitallibrary/grading/all_subject_configs.html'
        context = {
            'configs': configs,
            'title': 'Subject Grading Configurations'
        }
    
    return render(request, template, context)

@staff_member_required
def subject_grading_create(request, subject_id):
    """Create a grading configuration for a subject"""
    subject = get_object_or_404(Subject, id=subject_id)
    grading_systems = GradingSystem.objects.filter(is_active=True)
    
    if request.method == 'POST':
        grading_system_id = request.POST.get('grading_system')
        academic_year = request.POST.get('academic_year')
        term = request.POST.get('term')
        max_score = request.POST.get('max_score')
        passing_score = request.POST.get('passing_score')
        exam_weight = request.POST.get('exam_weight')
        coursework_weight = request.POST.get('coursework_weight')
        
        grading_system = get_object_or_404(GradingSystem, id=grading_system_id)
        
        config, created = SubjectGradingConfig.objects.get_or_create(
            subject=subject,
            academic_year=academic_year,
            term=term if term else None,
            defaults={
                'grading_system': grading_system,
                'max_score': max_score or 100,
                'passing_score': passing_score or 50,
                'exam_weight': exam_weight or 70,
                'coursework_weight': coursework_weight or 30,
            }
        )
        
        if not created:
            config.grading_system = grading_system
            config.max_score = max_score or 100
            config.passing_score = passing_score or 50
            config.exam_weight = exam_weight or 70
            config.coursework_weight = coursework_weight or 30
            config.save()
            messages.success(request, f'Updated grading configuration for {subject.name}')
        else:
            messages.success(request, f'Created grading configuration for {subject.name}')
        
        return redirect('digitallibrary:subject_grading_list', subject_id=subject.id)
    
    context = {
        'subject': subject,
        'grading_systems': grading_systems,
        'academic_years': get_academic_years(),
        'terms': [(1, 'Term 1'), (2, 'Term 2'), (3, 'Term 3')],
    }
    return render(request, 'digitallibrary/grading/subject_grading_form.html', context)

@staff_member_required
def subject_grading_edit(request, config_id):
    """Edit a subject grading configuration"""
    config = get_object_or_404(SubjectGradingConfig, id=config_id)
    grading_systems = GradingSystem.objects.filter(is_active=True)
    
    if request.method == 'POST':
        config.grading_system_id = request.POST.get('grading_system')
        config.max_score = request.POST.get('max_score')
        config.passing_score = request.POST.get('passing_score')
        config.exam_weight = request.POST.get('exam_weight')
        config.coursework_weight = request.POST.get('coursework_weight')
        config.is_active = request.POST.get('is_active') == 'on'
        config.save()
        
        messages.success(request, f'Updated grading configuration for {config.subject.name}')
        return redirect('digitallibrary:subject_grading_list', subject_id=config.subject.id)
    
    context = {
        'config': config,
        'grading_systems': grading_systems,
        'academic_years': get_academic_years(),
        'terms': [(1, 'Term 1'), (2, 'Term 2'), (3, 'Term 3')],
    }
    return render(request, 'digitallibrary/grading/subject_grading_form.html', context)

def get_academic_years():
    """Helper to get academic year choices"""
    current_year = timezone.now().year
    return [(str(year), str(year)) for year in range(current_year - 5, current_year + 6)]
from django.http import HttpResponse
from django.template import loader
import os
from django.conf import settings

import os
from django.conf import settings
from django.shortcuts import render
from django.http import HttpResponse
from django_tenants.utils import schema_context
from tenants.models import School  # Your School model
# Import from your actual models
from digitallibrary.models import Student, Resource
# Teacher doesn't exist as a separate model - use UserProfile instead
from digitallibrary.models import UserProfile
# If you need a Teacher reference, you can do:
Teacher = UserProfile  # Alias for compatibility

def landing_page(request):
    """Landing page with real database statistics"""
    from django.db import connection
    from django.http import HttpResponse
    
    # Get real data using raw SQL
    schools_count = 0
    teachers_count = 0
    students_count = 0
    resources_count = 0
    views_count = 0
    
    try:
        with connection.cursor() as cursor:
            # Count schools
            cursor.execute("SELECT COUNT(*) FROM tenants_school WHERE is_active = true")
            result = cursor.fetchone()
            schools_count = result[0] if result else 0
            
            # Count teachers (staff users who are not superusers)
            cursor.execute("SELECT COUNT(*) FROM auth_user WHERE is_staff = true AND is_superuser = false AND is_active = true")
            result = cursor.fetchone()
            teachers_count = result[0] if result else 0
            
            # Count students (non-staff active users)
            cursor.execute("SELECT COUNT(*) FROM auth_user WHERE is_staff = false AND is_active = true")
            result = cursor.fetchone()
            students_count = result[0] if result else 0
            
            # Count resources
            try:
                cursor.execute("SELECT COUNT(*) FROM digitallibrary_resource WHERE is_approved = true")
                result = cursor.fetchone()
                resources_count = result[0] if result else 0
            except:
                resources_count = 0
            
            # Count views
            try:
                cursor.execute("SELECT COUNT(*) FROM digitallibrary_resourceview")
                result = cursor.fetchone()
                views_count = result[0] if result else 0
            except:
                views_count = 0
            
            print(f"Real data - Schools: {schools_count}, Teachers: {teachers_count}, Students: {students_count}, Resources: {resources_count}, Views: {views_count}")
            
    except Exception as e:
        print(f"Error getting counts: {e}")
    
    # Beautiful HTML template with real data
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ShuleHub | Digital School Management System</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <style>
        * {{ font-family: 'Inter', sans-serif; }}
        body {{ background: linear-gradient(135deg, #0f0c29 0%, #302b63 50%, #24243e 100%); }}
        .stat-card {{ transition: all 0.3s ease; }}
        .stat-card:hover {{ transform: translateY(-10px); }}
    </style>
</head>
<body class="text-white">
    <!-- Navigation -->
    <nav class="bg-black/30 backdrop-blur-md fixed w-full z-50">
        <div class="container mx-auto px-6 py-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center space-x-2">
                    <div class="w-10 h-10 bg-gradient-to-br from-green-500 to-emerald-600 rounded-xl flex items-center justify-center">
                        <svg class="w-6 h-6 text-white" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                            <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M12 6.253v13m0-13C10.832 5.477 9.246 5 7.5 5S4.168 5.477 3 6.253v13C4.168 18.477 5.754 18 7.5 18s3.332.477 4.5 1.253m0-13C13.168 5.477 14.754 5 16.5 5c1.747 0 3.332.477 4.5 1.253v13C19.832 18.477 18.247 18 16.5 18c-1.746 0-3.332.477-4.5 1.253"></path>
                        </svg>
                    </div>
                    <span class="text-2xl font-bold text-transparent bg-clip-text bg-gradient-to-r from-green-400 to-emerald-500">ShuleHub</span>
                </div>
                <div class="hidden md:flex space-x-8">
                    <a href="#features" class="text-gray-300 hover:text-green-400 transition">Features</a>
                    <a href="#stats" class="text-gray-300 hover:text-green-400 transition">Impact</a>
                    <a href="#contact" class="text-gray-300 hover:text-green-400 transition">Contact</a>
                </div>
                <a href="/admin/" class="px-6 py-2 bg-gradient-to-r from-green-600 to-emerald-600 rounded-full text-sm font-semibold hover:shadow-lg hover:shadow-green-500/30 transition transform hover:scale-105">
                    Admin Login
                </a>
            </div>
        </div>
    </nav>

    <!-- Hero Section -->
    <section class="pt-32 pb-20 text-center relative">
        <div class="container mx-auto px-6">
            <div class="inline-block mb-6 animate-bounce">
                <div class="glass rounded-2xl px-8 py-4 bg-white/10 backdrop-blur">
                    <span class="text-green-400 font-semibold tracking-wider">🚀 NEXT-GEN EDUCATION PLATFORM</span>
                </div>
            </div>
            <h1 class="text-5xl md:text-7xl lg:text-8xl font-black mb-6">
                Transform Your<br>
                <span class="text-transparent bg-clip-text bg-gradient-to-r from-green-400 to-emerald-500">School Management</span>
            </h1>
            <p class="text-xl md:text-2xl text-gray-400 max-w-3xl mx-auto mb-12">
                The most powerful integrated digital school management system empowering education across Kenya
            </p>
            <div class="flex flex-col sm:flex-row gap-4 justify-center">
                <a href="#contact" class="px-8 py-4 bg-gradient-to-r from-green-600 to-emerald-600 rounded-xl font-bold text-lg hover:shadow-2xl hover:shadow-green-500/40 transition transform hover:scale-105">
                    Start Free Trial <i class="fas fa-arrow-right ml-2"></i>
                </a>
                <a href="#features" class="px-8 py-4 bg-white/10 backdrop-blur rounded-xl font-bold text-lg border border-green-500/30 hover:border-green-500 transition transform hover:scale-105">
                    Explore Features <i class="fas fa-play ml-2"></i>
                </a>
            </div>
        </div>
    </section>

    <!-- Stats Section with REAL DATA -->
    <section id="stats" class="py-20">
        <div class="container mx-auto px-6">
            <div class="text-center mb-16">
                <h2 class="text-4xl md:text-5xl font-bold mb-4 text-transparent bg-clip-text bg-gradient-to-r from-green-400 to-emerald-500">Making an Impact Across Kenya</h2>
                <p class="text-xl text-gray-400">Real-time statistics from schools using ShuleHub</p>
            </div>
            
            <div class="grid grid-cols-2 md:grid-cols-5 gap-6 max-w-6xl mx-auto">
                <div class="bg-white/10 backdrop-blur rounded-2xl p-6 text-center stat-card">
                    <div class="text-4xl mb-2">🏫</div>
                    <div class="text-3xl md:text-4xl font-bold text-green-400">{schools_count:,}</div>
                    <div class="text-sm text-gray-300 mt-2 font-semibold">Schools Onboarded</div>
                </div>
                <div class="bg-white/10 backdrop-blur rounded-2xl p-6 text-center stat-card">
                    <div class="text-4xl mb-2">👨‍🏫</div>
                    <div class="text-3xl md:text-4xl font-bold text-green-400">{teachers_count:,}</div>
                    <div class="text-sm text-gray-300 mt-2 font-semibold">Teachers Onboarded</div>
                </div>
                <div class="bg-white/10 backdrop-blur rounded-2xl p-6 text-center stat-card">
                    <div class="text-4xl mb-2">👨‍🎓</div>
                    <div class="text-3xl md:text-4xl font-bold text-green-400">{students_count:,}</div>
                    <div class="text-sm text-gray-300 mt-2 font-semibold">Active Students</div>
                </div>
                <div class="bg-white/10 backdrop-blur rounded-2xl p-6 text-center stat-card">
                    <div class="text-4xl mb-2">📚</div>
                    <div class="text-3xl md:text-4xl font-bold text-green-400">{resources_count:,}</div>
                    <div class="text-sm text-gray-300 mt-2 font-semibold">Learning Resources</div>
                </div>
                <div class="bg-white/10 backdrop-blur rounded-2xl p-6 text-center stat-card">
                    <div class="text-4xl mb-2">👁️</div>
                    <div class="text-3xl md:text-4xl font-bold text-green-400">{views_count:,}</div>
                    <div class="text-sm text-gray-300 mt-2 font-semibold">Resources Accessed</div>
                </div>
            </div>
        </div>
    </section>

    <!-- Features Section -->
    <section id="features" class="py-20 bg-black/20">
        <div class="container mx-auto px-6">
            <h2 class="text-3xl md:text-4xl font-bold text-center mb-12 text-transparent bg-clip-text bg-gradient-to-r from-green-400 to-emerald-500">Why Choose ShuleHub?</h2>
            <div class="grid md:grid-cols-3 gap-8 max-w-6xl mx-auto">
                <div class="bg-white/5 rounded-2xl p-6 text-center hover:transform hover:scale-105 transition">
                    <div class="text-5xl mb-3">📚</div>
                    <h3 class="text-xl font-bold text-green-400 mb-2">Complete Platform</h3>
                    <p class="text-gray-400">All-in-one school management solution for Kenyan schools</p>
                </div>
                <div class="bg-white/5 rounded-2xl p-6 text-center hover:transform hover:scale-105 transition">
                    <div class="text-5xl mb-3">📖</div>
                    <h3 class="text-xl font-bold text-green-400 mb-2">Digital Library</h3>
                    <p class="text-gray-400">Access textbooks, past papers, and educational resources</p>
                </div>
                <div class="bg-white/5 rounded-2xl p-6 text-center hover:transform hover:scale-105 transition">
                    <div class="text-5xl mb-3">📊</div>
                    <h3 class="text-xl font-bold text-green-400 mb-2">Performance Tracking</h3>
                    <p class="text-gray-400">Track and analyze student academic performance</p>
                </div>
                <div class="bg-white/5 rounded-2xl p-6 text-center hover:transform hover:scale-105 transition">
                    <div class="text-5xl mb-3">📱</div>
                    <h3 class="text-xl font-bold text-green-400 mb-2">SMS Alerts</h3>
                    <p class="text-gray-400">Keep parents informed with automated notifications</p>
                </div>
                <div class="bg-white/5 rounded-2xl p-6 text-center hover:transform hover:scale-105 transition">
                    <div class="text-5xl mb-3">👨‍👩‍👧‍👦</div>
                    <h3 class="text-xl font-bold text-green-400 mb-2">Parent Portal</h3>
                    <p class="text-gray-400">Real-time access to results, fees, and attendance</p>
                </div>
                <div class="bg-white/5 rounded-2xl p-6 text-center hover:transform hover:scale-105 transition">
                    <div class="text-5xl mb-3">💰</div>
                    <h3 class="text-xl font-bold text-green-400 mb-2">Fee Management</h3>
                    <p class="text-gray-400">Track payments, manage balances, and generate receipts</p>
                </div>
            </div>
        </div>
    </section>

    <!-- CTA Banner -->
    <section class="py-20">
        <div class="container mx-auto px-6">
            <div class="bg-gradient-to-r from-green-600/20 to-emerald-600/20 rounded-3xl p-12 text-center max-w-5xl mx-auto backdrop-blur border border-green-500/30">
                <h2 class="text-3xl md:text-4xl font-bold mb-4">Ready to Transform <span class="text-transparent bg-clip-text bg-gradient-to-r from-green-400 to-emerald-500">Your School?</span></h2>
                <p class="text-xl text-gray-300 mb-8">Join {schools_count}+ schools already using ShuleHub to enhance learning outcomes</p>
                <div class="flex flex-wrap gap-4 justify-center">
                    <a href="mailto:kabasil81@gmail.com?subject=School%20Onboarding%20Request" class="px-8 py-4 bg-gradient-to-r from-green-600 to-emerald-600 rounded-xl font-bold hover:shadow-2xl hover:shadow-green-500/40 transition transform hover:scale-105">
                        Start Your Journey <i class="fas fa-arrow-right ml-2"></i>
                    </a>
                    <a href="tel:+254708941520" class="px-8 py-4 bg-white/10 rounded-xl font-bold border border-green-500/30 hover:border-green-500 transition transform hover:scale-105">
                        <i class="fas fa-phone-alt mr-2"></i> Schedule Demo
                    </a>
                </div>
            </div>
        </div>
    </section>

    <!-- Contact Section -->
    <section id="contact" class="py-20 bg-black/20">
        <div class="container mx-auto px-6">
            <div class="grid md:grid-cols-2 gap-12 max-w-5xl mx-auto">
                <div>
                    <h2 class="text-4xl font-bold mb-4 text-transparent bg-clip-text bg-gradient-to-r from-green-400 to-emerald-500">Get in Touch</h2>
                    <p class="text-gray-400 mb-8 text-lg">Have questions? We're here to help you transform your school's management.</p>
                    
                    <div class="space-y-6">
                        <div class="flex items-center space-x-4">
                            <div class="w-12 h-12 bg-green-500/20 rounded-xl flex items-center justify-center">
                                <svg class="w-6 h-6 text-green-400" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="1.5" d="M3 8l7.89 5.26a2 2 0 002.22 0L21 8M5 19h14a2 2 0 002-2V7a2 2 0 00-2-2H5a2 2 0 00-2 2v10a2 2 0 002 2z"></path>
                                </svg>
                            </div>
                            <div>
                                <p class="text-gray-400 text-sm">Email Us</p>
                                <a href="mailto:kabasil81@gmail.com" class="text-white font-semibold hover:text-green-400 transition">kabasil81@gmail.com</a>
                            </div>
                        </div>
                        
                        <div class="flex items-center space-x-4">
                            <div class="w-12 h-12 bg-green-500/20 rounded-xl flex items-center justify-center">
                                <svg class="w-6 h-6 text-green-400" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="1.5" d="M3 5a2 2 0 012-2h3.28a1 1 0 01.948.684l1.498 4.493a1 1 0 01-.502 1.21l-2.257 1.13a11.042 11.042 0 005.516 5.516l1.13-2.257a1 1 0 011.21-.502l4.493 1.498a1 1 0 01.684.949V19a2 2 0 01-2 2h-1C9.716 21 3 14.284 3 6V5z"></path>
                                </svg>
                            </div>
                            <div>
                                <p class="text-gray-400 text-sm">Call Us</p>
                                <a href="tel:+254708941520" class="text-white font-semibold hover:text-green-400 transition">+254 708 941 520</a>
                            </div>
                        </div>
                        
                        <div class="flex items-center space-x-4">
                            <div class="w-12 h-12 bg-green-500/20 rounded-xl flex items-center justify-center">
                                <svg class="w-6 h-6 text-green-400" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="1.5" d="M21 12a9 9 0 01-9 9m9-9a9 9 0 00-9-9m9 9H3m9 9a9 9 0 01-9-9m9 9c1.66 0 3-4 3-9s-1.34-9-3-9m0 18c-1.66 0-3-4-3-9s1.34-9 3-9"></path>
                                </svg>
                            </div>
                            <div>
                                <p class="text-gray-400 text-sm">School Subdomain</p>
                                <code class="text-green-400 text-sm">yourschool.shulehub.org</code>
                            </div>
                        </div>
                    </div>
                </div>
                
                <div class="bg-white/10 backdrop-blur rounded-2xl p-8">
                    <h3 class="text-2xl font-bold mb-4">Ready to get started?</h3>
                    <p class="text-gray-400 mb-6">Fill out this form and our team will reach out within 24 hours.</p>
                    
                    <form action="mailto:kabasil81@gmail.com" method="POST" enctype="text/plain" class="space-y-4">
                        <input type="text" placeholder="School Name" class="w-full px-4 py-3 bg-gray-900/50 rounded-xl border border-gray-700 focus:border-green-500 focus:outline-none transition text-white">
                        <input type="email" placeholder="Your Email" class="w-full px-4 py-3 bg-gray-900/50 rounded-xl border border-gray-700 focus:border-green-500 focus:outline-none transition text-white">
                        <input type="tel" placeholder="Phone Number" class="w-full px-4 py-3 bg-gray-900/50 rounded-xl border border-gray-700 focus:border-green-500 focus:outline-none transition text-white">
                        <textarea rows="4" placeholder="Message" class="w-full px-4 py-3 bg-gray-900/50 rounded-xl border border-gray-700 focus:border-green-500 focus:outline-none transition text-white"></textarea>
                        <button type="submit" class="w-full px-6 py-3 bg-gradient-to-r from-green-600 to-emerald-600 rounded-xl font-semibold hover:shadow-lg hover:shadow-green-500/30 transition transform hover:scale-105">
                            Send Message <i class="fas fa-paper-plane ml-2"></i>
                        </button>
                    </form>
                </div>
            </div>
        </div>
    </section>

    <!-- Footer -->
    <footer class="py-12 px-6 border-t border-gray-800">
        <div class="container mx-auto text-center">
            <div class="flex justify-center space-x-8 mb-6">
                <a href="#" class="text-gray-500 hover:text-green-400 transition"><i class="fab fa-facebook-f text-xl"></i></a>
                <a href="#" class="text-gray-500 hover:text-green-400 transition"><i class="fab fa-twitter text-xl"></i></a>
                <a href="#" class="text-gray-500 hover:text-green-400 transition"><i class="fab fa-linkedin-in text-xl"></i></a>
                <a href="#" class="text-gray-500 hover:text-green-400 transition"><i class="fab fa-instagram text-xl"></i></a>
            </div>
            <p class="text-gray-500">© 2026 ShuleHub. All rights reserved. Empowering Kenyan Education.</p>
        </div>
    </footer>

    <script src="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/js/all.min.js"></script>
</body>
</html>"""
    
    return HttpResponse(html)

# digitallibrary/views.py - Add these functions at the end of the file or near other student views

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods
from django.shortcuts import get_object_or_404, render
from .models import Student, StudentActionLog
import json
import logging

logger = logging.getLogger(__name__)


def is_admin_or_principal(user):
    """Check if user is admin or principal"""
    if user.is_superuser:
        return True
    if hasattr(user, 'profile') and user.profile.role in ['admin', 'principal']:
        return True
    return False


@login_required
@user_passes_test(is_admin_or_principal)
@require_http_methods(["POST"])
def soft_delete_student(request, student_id):
    """
    Soft delete a student - marks as inactive with reason
    """
    try:
        student = get_object_or_404(Student, id=student_id)
        
        # Parse request body
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            data = {}
        
        reason_type = data.get('reason_type', 'other')
        reason = data.get('reason', '')
        transfer_to = data.get('transfer_to', '')
        
        # Validate reason
        if reason_type == 'other' and not reason.strip():
            return JsonResponse({
                'success': False,
                'error': 'Please provide a reason for deactivation'
            })
        
        # Perform soft delete
        student.soft_delete(
            user=request.user,
            reason=reason or dict(Student.TRANSFER_REASON_CHOICES).get(reason_type, 'Deactivated'),
            reason_type=reason_type,
            transfer_to=transfer_to
        )
        
        # Log the action
        logger.info(f"Student {student.admission_number} ({student.get_full_name()}) deactivated by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': f'Student {student.get_full_name()} has been deactivated successfully',
            'student_id': student.id,
            'status': 'inactive'
        })
        
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student not found'}, status=404)
    except Exception as e:
        logger.error(f"Error deactivating student {student_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@user_passes_test(is_admin_or_principal)
@require_http_methods(["POST"])
def reactivate_student(request, student_id):
    """
    Reactivate a soft-deleted student
    """
    try:
        student = get_object_or_404(Student, id=student_id)
        
        # Parse request body
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            data = {}
        
        reason = data.get('reason', 'Reactivated by admin')
        
        # Check if student is already active
        if student.is_active:
            return JsonResponse({
                'success': False,
                'error': 'Student is already active'
            })
        
        # Perform reactivation
        student.reactivate(user=request.user, reason=reason)
        
        # Log the action
        logger.info(f"Student {student.admission_number} ({student.get_full_name()}) reactivated by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': f'Student {student.get_full_name()} has been reactivated successfully',
            'student_id': student.id,
            'status': 'active'
        })
        
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student not found'}, status=404)
    except Exception as e:
        logger.error(f"Error reactivating student {student_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@user_passes_test(is_admin_or_principal)
def student_action_log(request, student_id):
    """
    View to show student action history
    """
    student = get_object_or_404(Student, id=student_id)
    actions = student.action_logs.all()[:50]
    
    context = {
        'student': student,
        'actions': actions,
    }
    return render(request, 'fees/student_action_log.html', context)


@login_required
@user_passes_test(is_admin_or_principal)
@require_http_methods(["POST"])
def bulk_student_action(request):
    """
    Bulk action for multiple students (delete/reactivate)
    """
    try:
        data = json.loads(request.body)
        student_ids = data.get('student_ids', [])
        action = data.get('action', '')  # 'delete' or 'reactivate'
        reason = data.get('reason', '')
        
        if not student_ids:
            return JsonResponse({'success': False, 'error': 'No students selected'})
        
        results = {
            'successful': [],
            'failed': []
        }
        
        for student_id in student_ids:
            try:
                student = Student.objects.get(id=student_id)
                
                if action == 'delete':
                    if student.is_active:
                        student.soft_delete(user=request.user, reason=reason, reason_type='bulk')
                        results['successful'].append({
                            'id': student.id,
                            'name': student.get_full_name(),
                            'admission': student.admission_number
                        })
                    else:
                        results['failed'].append({
                            'id': student.id,
                            'name': student.get_full_name(),
                            'error': 'Already inactive'
                        })
                elif action == 'reactivate':
                    if not student.is_active:
                        student.reactivate(user=request.user, reason=reason)
                        results['successful'].append({
                            'id': student.id,
                            'name': student.get_full_name(),
                            'admission': student.admission_number
                        })
                    else:
                        results['failed'].append({
                            'id': student.id,
                            'name': student.get_full_name(),
                            'error': 'Already active'
                        })
            except Student.DoesNotExist:
                results['failed'].append({
                    'id': student_id,
                    'error': 'Student not found'
                })
        
        return JsonResponse({
            'success': True,
            'results': results,
            'total_successful': len(results['successful']),
            'total_failed': len(results['failed'])
        })
        
    except Exception as e:
        logger.error(f"Error in bulk action: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
from django.contrib import messages
from .models import SchoolSetting

def school_settings(request):
    """School settings page - only accessible by admins"""
    from .models import SchoolSetting
    
    # Check if user is admin
    if not request.user.is_authenticated:
        return redirect('digitallibrary:login')
    
    if not (request.user.is_superuser or 
            (hasattr(request.user, 'profile') and 
             request.user.profile.role in ['admin', 'principal'])):
        messages.error(request, 'You do not have permission to access school settings.')
        return redirect('digitallibrary:home')
    
    school_setting, created = SchoolSetting.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        school_setting.name = request.POST.get('name', school_setting.name)
        school_setting.motto = request.POST.get('motto', school_setting.motto)
        school_setting.address = request.POST.get('address', school_setting.address)
        school_setting.phone = request.POST.get('phone', school_setting.phone)
        school_setting.email = request.POST.get('email', school_setting.email)
        school_setting.website = request.POST.get('website', school_setting.website)
        
        if request.FILES.get('logo'):
            school_setting.logo = request.FILES['logo']
        
        school_setting.save()
        messages.success(request, 'School settings updated successfully!')
        return redirect('digitallibrary:school_settings')
    
    context = {
        'school_setting': school_setting,
    }
    return render(request, 'digitallibrary/school_settings.html', context)

@staff_member_required
def exam_results_entry(request, exam_id):
    """
    Enter results for an exam - by subject, filtered by registered student subjects
    """
    from .models import Exam, Subject, Student, StudentResult
    
    exam = get_object_or_404(Exam, pk=exam_id)
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    
    selected_subject_id = request.GET.get('subject')
    selected_subject = None
    students = []
    existing_results = {}
    
    if selected_subject_id:
        try:
            selected_subject = Subject.objects.get(pk=selected_subject_id, is_active=True)
            
            students_qs = exam.get_students_for_exam()
            students = students_qs.filter(
                subjects=selected_subject,
                is_active=True
            ).distinct().order_by('admission_number')
            
            existing_results_qs = StudentResult.objects.filter(
                exam=exam,
                subject=selected_subject,
                student__in=students
            ).select_related('student')
            
            existing_results = {result.student_id: result for result in existing_results_qs}
            
        except Subject.DoesNotExist:
            messages.error(request, "Selected subject does not exist.")
    
    if request.method == 'POST':
        subject_id = request.POST.get('subject_id')
        
        if subject_id:
            selected_subject = get_object_or_404(Subject, pk=subject_id, is_active=True)
            
            students = exam.get_students_for_exam().filter(
                subjects=selected_subject,
                is_active=True
            ).distinct()
            
            saved_count = 0
            
            for student in students:
                score_key = f'score_{student.id}'
                if score_key in request.POST:
                    score = request.POST.get(score_key)
                    
                    if score and score.strip():
                        try:
                            score_value = float(score)
                            if 0 <= score_value <= (exam.max_score or 100):
                                StudentResult.objects.update_or_create(
                                    student=student,
                                    exam=exam,
                                    subject=selected_subject,
                                    defaults={'score': score_value, 'entered_by': request.user}
                                )
                                saved_count += 1
                        except ValueError:
                            pass
            
            if saved_count > 0:
                messages.success(request, f'Results for {exam.name} - {selected_subject.name} saved successfully!')
            else:
                messages.warning(request, 'No results were saved.')
            
            return redirect(f'{request.path}?subject={subject_id}')
    
    context = {
        'exam': exam,
        'subjects': subjects,
        'selected_subject': selected_subject,
        'students': students,
        'existing_results': existing_results,
        'title': f'Enter Results - {exam.name}',
        'school': SchoolSetting.objects.first(),
    }
    
    return render(request, 'performance/exam_results_entry.html', context)


# digitallibrary/views.py

@staff_member_required
def bulk_enter_results(request):
    """
    Bulk Excel upload page - direct file upload and processing
    """
    from .models import Exam, Subject, SchoolSetting
    
    print("\n" + "="*60)
    print("🔵 bulk_enter_results view called")
    print(f"   Method: {request.method}")
    print("="*60)
    
    # Get exams and subjects for dropdowns
    exams = Exam.objects.filter(is_active=True).order_by('-academic_year', '-created_at')
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    
    print(f"📋 Exams found: {exams.count()}")
    print(f"📋 Subjects found: {subjects.count()}")
    
    if request.method == 'POST':
        exam_id = request.POST.get('exam')
        subject_id = request.POST.get('subject')
        grading_system = request.POST.get('grading_system', 'cbe')
        excel_file = request.FILES.get('excel_file')
        
        print(f"📝 POST data:")
        print(f"   exam_id: {exam_id}")
        print(f"   subject_id: {subject_id}")
        print(f"   grading_system: {grading_system}")
        print(f"   file: {excel_file.name if excel_file else 'None'}")
        
        if not exam_id or not subject_id or not excel_file:
            messages.error(request, 'Please select exam, subject and upload a file')
            return redirect('digitallibrary:bulk_enter_results')
        
        try:
            exam = Exam.objects.get(id=exam_id)
            subject = Subject.objects.get(id=subject_id)
            use_cbe = grading_system == 'cbe'
            
            # Process the Excel file
            import pandas as pd
            from django.db import connection
            
            # Read file
            ext = excel_file.name.split('.')[-1].lower()
            if ext == 'csv':
                df = pd.read_csv(excel_file)
            else:
                df = pd.read_excel(excel_file)
            
            # Normalize columns
            df.columns = df.columns.str.strip().str.lower()
            
            # Find admission and score columns
            admission_col = None
            score_col = None
            
            for col in df.columns:
                if 'admission' in col or 'adm' in col or 'reg' in col:
                    admission_col = col
                elif 'score' in col or 'mark' in col or 'result' in col:
                    score_col = col
            
            if admission_col is None or score_col is None:
                messages.error(request, 'Excel file must have "Admission Number" and "Score" columns')
                return redirect('digitallibrary:bulk_enter_results')
            
            results_processed = 0
            errors = []
            max_score = float(exam.max_score) if exam.max_score else 100.0
            
            with connection.cursor() as cursor:
                for index, row in df.iterrows():
                    admission_number = str(row[admission_col]).strip() if pd.notna(row[admission_col]) else None
                    score_value = row[score_col] if pd.notna(row[score_col]) else None
                    
                    if not admission_number or score_value is None:
                        continue
                    
                    try:
                        score = float(score_value)
                        
                        if score < 0 or score > max_score:
                            errors.append(f"Row {index + 2}: Score {score} out of range (0-{max_score})")
                            continue
                        
                        # Get student
                        from .models import Student
                        student = Student.objects.filter(admission_number=admission_number, is_active=True).first()
                        if not student:
                            errors.append(f"Row {index + 2}: Student '{admission_number}' not found")
                            continue
                        
                        if use_cbe:
                            # Get CBE grade
                            cursor.execute("""
                                SELECT id, points FROM digitallibrary_kneccbegrade 
                                WHERE min_score <= %s AND max_score >= %s
                                LIMIT 1
                            """, [score, score])
                            grade = cursor.fetchone()
                            if grade:
                                grade_id, points = grade
                                cursor.execute("""
                                    INSERT INTO digitallibrary_studentresult 
                                    (student_id, exam_id, subject_id, score, grade_id, points, entered_by_id, entered_at, updated_at)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                                    ON CONFLICT (student_id, exam_id, subject_id) 
                                    DO UPDATE SET 
                                        score = EXCLUDED.score,
                                        grade_id = EXCLUDED.grade_id,
                                        points = EXCLUDED.points,
                                        updated_at = NOW()
                                """, [student.id, exam.id, subject.id, score, grade_id, points, request.user.id])
                                results_processed += 1
                        else:
                            # Traditional grading
                            percentage = (score / max_score) * 100
                            if percentage >= 80: grade_name = 'A'; points = 12
                            elif percentage >= 75: grade_name = 'A-'; points = 11
                            elif percentage >= 70: grade_name = 'B+'; points = 10
                            elif percentage >= 65: grade_name = 'B'; points = 9
                            elif percentage >= 60: grade_name = 'B-'; points = 8
                            elif percentage >= 55: grade_name = 'C+'; points = 7
                            elif percentage >= 50: grade_name = 'C'; points = 6
                            elif percentage >= 45: grade_name = 'C-'; points = 5
                            elif percentage >= 40: grade_name = 'D+'; points = 4
                            elif percentage >= 35: grade_name = 'D'; points = 3
                            elif percentage >= 30: grade_name = 'D-'; points = 2
                            else: grade_name = 'E'; points = 1
                            
                            cursor.execute("""
                                SELECT id FROM digitallibrary_grade WHERE grade = %s LIMIT 1
                            """, [grade_name])
                            grade = cursor.fetchone()
                            grade_id = grade[0] if grade else None
                            
                            cursor.execute("""
                                INSERT INTO digitallibrary_studentresult 
                                (student_id, exam_id, subject_id, score, grade_id, points, entered_by_id, entered_at, updated_at)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                                ON CONFLICT (student_id, exam_id, subject_id) 
                                DO UPDATE SET 
                                    score = EXCLUDED.score,
                                    grade_id = EXCLUDED.grade_id,
                                    points = EXCLUDED.points,
                                    updated_at = NOW()
                            """, [student.id, exam.id, subject.id, score, grade_id, points, request.user.id])
                            results_processed += 1
                            
                    except Exception as e:
                        errors.append(f"Row {index + 2}: {str(e)}")
            
            if results_processed > 0:
                messages.success(request, f'✅ Successfully processed {results_processed} results for {exam.name} - {subject.name}')
            else:
                messages.warning(request, '⚠️ No valid results were found in the file.')
            
            if errors:
                for error in errors[:5]:
                    messages.warning(request, error)
                if len(errors) > 5:
                    messages.info(request, f'... and {len(errors) - 5} more errors')
                    
            return redirect('digitallibrary:exam_list')
            
        except Exam.DoesNotExist:
            messages.error(request, 'Selected exam not found')
        except Subject.DoesNotExist:
            messages.error(request, 'Selected subject not found')
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            print(f"❌ Exception: {e}")
        
        return redirect('digitallibrary:bulk_enter_results')
    
    # GET request - show the upload form
    context = {
        'exams': exams,
        'subjects': subjects,
        'title': 'Bulk Excel Upload',
        'school': SchoolSetting.objects.first(),
    }
    
    return render(request, 'performance/bulk_excel_upload.html', context)

@staff_member_required
def bulk_results_entry_by_class(request, exam_id, class_id):
    """
    Enter results for all students in a class for all subjects
    """
    from .models import Exam, Class, Subject, Student, StudentResult
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect, render
    
    # Get objects
    exam = get_object_or_404(Exam, id=exam_id)
    student_class = get_object_or_404(Class, id=class_id)
    
    # Get students in this class - try different field names
    students = Student.objects.filter(
        current_class=student_class,
        is_active=True
    ).order_by('first_name', 'last_name')
    
    # If no students found with 'current_class', try 'student_class'
    if not students.exists():
        students = Student.objects.filter(
            student_class=student_class,
            is_active=True
        ).order_by('first_name', 'last_name')
    
    # Get all active subjects
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    
    # Get existing results
    existing_results = {}
    if students.exists():
        results = StudentResult.objects.filter(
            exam=exam, 
            student__in=students
        )
        for r in results:
            key = f"{r.student_id}_{r.subject_id}"
            existing_results[key] = r
    
    # Handle POST request
    if request.method == 'POST':
        saved_count = 0
        for key, value in request.POST.items():
            if key.startswith('score_') and value.strip():
                parts = key.replace('score_', '').split('_')
                if len(parts) == 2:
                    student_id, subject_id = parts
                    try:
                        score = float(value)
                        student = Student.objects.get(id=student_id)
                        subject = Subject.objects.get(id=subject_id)
                        
                        StudentResult.objects.update_or_create(
                            student=student,
                            exam=exam,
                            subject=subject,
                            defaults={'score': score, 'entered_by': request.user}
                        )
                        saved_count += 1
                    except (ValueError, Student.DoesNotExist, Subject.DoesNotExist):
                        continue
        
        messages.success(request, f'Successfully saved {saved_count} results for {student_class.name}')
        return redirect('digitallibrary:bulk_results_entry_by_class', exam_id=exam.id, class_id=class_id)
    
    context = {
        'exam': exam,
        'class': student_class,
        'students': students,
        'subjects': subjects,
        'existing_results': existing_results,
        'student_count': students.count(),
        'subject_count': subjects.count(),
        'title': f'Bulk Results - {exam.name} - {student_class.name}',
    }
    return render(request, 'performance/bulk_results_entry_by_class.html', context)


@staff_member_required
def bulk_excel_upload(request):
    """
    Upload Excel file with results for bulk entry
    """
    from .models import Exam, Subject, Student, StudentResult, SchoolSetting
    import openpyxl
    
    if request.method == 'POST':
        exam_id = request.POST.get('exam')
        subject_id = request.POST.get('subject')
        excel_file = request.FILES.get('excel_file')
        
        if not all([exam_id, subject_id, excel_file]):
            messages.error(request, 'Please select exam, subject and upload an Excel file')
            return redirect('digitallibrary:bulk_excel_upload')
        
        try:
            exam = Exam.objects.get(id=exam_id)
            subject = Subject.objects.get(id=subject_id)
            
            # Load workbook
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            # Find column indices
            admission_col = None
            score_col = None
            
            for idx, cell in enumerate(sheet[1], 1):
                header = str(cell.value).lower().strip() if cell.value else ''
                if 'admission' in header or 'adm' in header or 'reg' in header:
                    admission_col = idx
                elif 'score' in header or 'mark' in header or 'result' in header:
                    score_col = idx
            
            if admission_col is None or score_col is None:
                messages.error(request, 'Excel file must have "Admission Number" and "Score" columns')
                return redirect('digitallibrary:bulk_excel_upload')
            
            results_processed = 0
            errors = []
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row:
                    continue
                
                admission_number = str(row[admission_col - 1]).strip() if row[admission_col - 1] else None
                score_value = row[score_col - 1] if score_col - 1 < len(row) else None
                
                if not admission_number or score_value is None:
                    continue
                
                try:
                    score = float(score_value)
                    max_score = float(exam.max_score) if exam.max_score else 100.0
                    
                    if score < 0 or score > max_score:
                        errors.append(f"Row {row_idx}: Score {score} is outside valid range (0-{max_score})")
                        continue
                    
                    student = Student.objects.get(admission_number=admission_number, is_active=True)
                    
                    StudentResult.objects.update_or_create(
                        student=student,
                        exam=exam,
                        subject=subject,
                        defaults={'score': score, 'entered_by': request.user}
                    )
                    results_processed += 1
                    
                except Student.DoesNotExist:
                    errors.append(f"Row {row_idx}: Student with admission number '{admission_number}' not found")
                except ValueError:
                    errors.append(f"Row {row_idx}: Invalid score value '{score_value}'")
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
            
            if results_processed > 0:
                messages.success(request, f'Successfully processed {results_processed} results for {exam.name} - {subject.name}')
            else:
                messages.warning(request, 'No valid results were found in the file.')
            
            if errors:
                for error in errors[:5]:
                    messages.warning(request, error)
                if len(errors) > 5:
                    messages.warning(request, f'... and {len(errors) - 5} more errors')
                    
            return redirect('digitallibrary:exam_list')
            
        except Exam.DoesNotExist:
            messages.error(request, 'Selected exam not found')
        except Subject.DoesNotExist:
            messages.error(request, 'Selected subject not found')
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
    
    exams = Exam.objects.all().order_by('-academic_year', '-created_at')
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    
    context = {
        'exams': exams,
        'subjects': subjects,
        'title': 'Bulk Excel Upload',
        'school': SchoolSetting.objects.first(),
    }
    return render(request, 'performance/bulk_excel_upload.html', context)


@staff_member_required
def exam_create(request):
    """
    Create a new exam
    """
    from .forms import ExamForm
    from .models import SchoolSetting
    
    if request.method == 'POST':
        form = ExamForm(request.POST)
        if form.is_valid():
            exam = form.save()
            messages.success(request, f'Exam "{exam.name}" created successfully!')
            return redirect('digitallibrary:exam_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ExamForm()
    
    context = {
        'form': form,
        'title': 'Create New Exam',
        'school': SchoolSetting.objects.first(),
    }
    return render(request, 'performance/exam_form.html', context)
@staff_member_required
def exam_performance_detail(request, exam_id):
    """View detailed performance analytics for a specific exam"""
    from .models import Exam, Subject, Student, StudentResult, Class
    from django.db.models import Avg, Sum, Count, Q
    from collections import defaultdict
    
    exam = get_object_or_404(Exam, id=exam_id)
    
    # Get students for this exam
    if exam.student_class:
        students = exam.student_class.students.filter(is_active=True)
    else:
        students = Student.objects.filter(is_active=True)
    
    total_students = students.count()
    
    # Get all subjects
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    
    # Get all results for this exam
    results = StudentResult.objects.filter(exam=exam).select_related('student', 'subject')
    
    # Calculate overall statistics
    overall_stats = results.aggregate(
        avg_score=Avg('score'),
        total_results=Count('id'),
        max_score=Avg('score'),
        min_score=Avg('score')
    )
    
    class_average = overall_stats['avg_score'] or 0
    total_results = overall_stats['total_results'] or 0
    
    # Calculate pass rate
    passed_results = results.filter(score__gte=50).count()
    pass_rate = (passed_results / total_results * 100) if total_results > 0 else 0
    
    # Grade distribution
    grade_distribution = {
        'A (80-100)': results.filter(score__gte=80).count(),
        'B (70-79)': results.filter(score__gte=70, score__lt=80).count(),
        'C (60-69)': results.filter(score__gte=60, score__lt=70).count(),
        'D (50-59)': results.filter(score__gte=50, score__lt=60).count(),
        'E (0-49)': results.filter(score__lt=50).count(),
    }
    
    # Subject performance
    subject_performance = []
    for subject in subjects:
        subject_results = results.filter(subject=subject)
        if subject_results.exists():
            avg = subject_results.aggregate(avg=Avg('score'))['avg'] or 0
            highest = subject_results.aggregate(max=Avg('score'))['max'] or 0
            lowest = subject_results.aggregate(min=Avg('score'))['min'] or 0
            passed = subject_results.filter(score__gte=50).count()
            
            # Determine grade
            if avg >= 80:
                grade = 'A'
            elif avg >= 70:
                grade = 'B'
            elif avg >= 60:
                grade = 'C'
            elif avg >= 50:
                grade = 'D'
            else:
                grade = 'E'
            
            subject_performance.append({
                'id': subject.id,
                'name': subject.name,
                'average': round(avg, 1),
                'highest': round(highest, 1),
                'lowest': round(lowest, 1),
                'passed': passed,
                'total_students': total_students,
                'pass_rate': round((passed / total_students * 100), 1),
                'grade': grade,
            })
    
    # Student rankings
    student_scores = defaultdict(lambda: {'total': 0, 'count': 0, 'scores': []})
    for result in results:
        student_scores[result.student_id]['total'] += result.score
        student_scores[result.student_id]['count'] += 1
        student_scores[result.student_id]['scores'].append(result.score)
    
    rankings = []
    for student_id, data in student_scores.items():
        student = Student.objects.filter(id=student_id).first()
        if student:
            average = data['total'] / data['count'] if data['count'] > 0 else 0
            
            # Determine grade
            if average >= 80:
                grade = 'A'
            elif average >= 75:
                grade = 'A-'
            elif average >= 70:
                grade = 'B+'
            elif average >= 65:
                grade = 'B'
            elif average >= 60:
                grade = 'B-'
            elif average >= 55:
                grade = 'C+'
            elif average >= 50:
                grade = 'C'
            elif average >= 45:
                grade = 'C-'
            elif average >= 40:
                grade = 'D+'
            else:
                grade = 'E'
            
            rankings.append({
                'student': student,
                'average': round(average, 1),
                'total': round(data['total'], 1),
                'grade': grade,
                'subjects_count': data['count'],
            })
    
    # Sort by average score (highest first)
    rankings.sort(key=lambda x: x['average'], reverse=True)
    
    # Add rank to each student
    for idx, ranking in enumerate(rankings, 1):
        ranking['rank'] = idx
    
    # Get top 10 students
    top_students = rankings[:10]
    
    # Find weakest subject
    weakest_subject = min(subject_performance, key=lambda x: x['average']) if subject_performance else None
    
    # Find strongest subject
    strongest_subject = max(subject_performance, key=lambda x: x['average']) if subject_performance else None
    
    # Calculate completion rate
    results_count = results.values('student').distinct().count()
    completion_rate = (results_count / total_students * 100) if total_students > 0 else 0
    
    # Get school settings
    school = SchoolSetting.objects.first()
    
    context = {
        'exam': exam,
        'total_students': total_students,
        'total_results': total_results,
        'class_average': round(class_average, 1),
        'pass_rate': round(pass_rate, 1),
        'completion_rate': round(completion_rate, 1),
        'grade_distribution': grade_distribution,
        'subject_performance': subject_performance,
        'rankings': rankings[:20],  # Top 20 for display
        'top_students': top_students,
        'weakest_subject': weakest_subject,
        'strongest_subject': strongest_subject,
        'school': school,
        'title': f'Performance - {exam.name}',
    }
    
    return render(request, 'performance/exam_performance_detail.html', context)
@staff_member_required
def subject_grading_config(request, subject_id):
    """Allow teachers to customize grading for a specific subject"""
    from .models import Subject, GradingSystem, SubjectGradingConfig, GradeScale
    
    subject = get_object_or_404(Subject, id=subject_id)
    school_system = GradingSystem.objects.filter(is_default=True, is_active=True).first()
    
    # Get or create config for current year/term
    academic_year = request.GET.get('year', str(timezone.now().year))
    term = request.GET.get('term', '1')
    
    config, created = SubjectGradingConfig.objects.get_or_create(
        subject=subject,
        grading_system=school_system,
        academic_year=academic_year,
        term=term,
        defaults={
            'max_score': 100,
            'passing_score': 50,
            'exam_weight': 70,
            'coursework_weight': 30,
            'is_active': True
        }
    )
    
    if request.method == 'POST':
        # Update grading configuration
        config.passing_score = request.POST.get('passing_score', 50)
        config.max_score = request.POST.get('max_score', 100)
        config.exam_weight = request.POST.get('exam_weight', 70)
        config.coursework_weight = request.POST.get('coursework_weight', 30)
        config.save()
        
        # Update grade scales for this subject
        grade_ids = request.POST.getlist('grade_id')
        grades = request.POST.getlist('grade[]')
        min_scores = request.POST.getlist('min_score[]')
        max_scores = request.POST.getlist('max_score[]')
        points = request.POST.getlist('points[]')
        
        for i, grade in enumerate(grades):
            if grade and min_scores[i] and max_scores[i]:
                GradeScale.objects.update_or_create(
                    id=grade_ids[i] if i < len(grade_ids) and grade_ids[i] else None,
                    defaults={
                        'grading_system': school_system,
                        'grade': grade,
                        'min_score': min_scores[i],
                        'max_score': max_scores[i],
                        'points': points[i] if i < len(points) else 0,
                        'is_active': True
                    }
                )
        
        messages.success(request, f'Grading configuration updated for {subject.name}')
        return redirect('digitallibrary:subject_grading_config', subject_id=subject.id)
    
    context = {
        'subject': subject,
        'config': config,
        'grade_scales': school_system.grades.all().order_by('-min_score'),
        'school_system': school_system,
        'academic_year': academic_year,
        'term': term,
    }
    return render(request, 'digitallibrary/subject_grading_config.html', context)
# digitallibrary/views.py

@staff_member_required
def knec_cbe_grading(request):
    """Display KNEC CBE Grading System"""
    from .models import KNECCBEGrade
    import json
    
    grades = KNECCBEGrade.objects.filter(is_active=True).order_by('order')
    
    # Convert grades to JSON for JavaScript
    grades_json = json.dumps([
        {
            'level': g.level,
            'level_name': g.level_name,
            'min_score': float(g.min_score),
            'max_score': float(g.max_score),
            'points': int(g.points),
            'placement': g.placement,
            'description': g.description
        }
        for g in grades
    ])
    
    context = {
        'grades': grades,
        'grades_json': grades_json,
        'title': 'KNEC CBE Grading System',
    }
    return render(request, 'digitallibrary/knec_cbe_grading.html', context)


# Initialize default grades on system start
def initialize_grading_system():
    """Run this in your management command or app ready signal"""
    from .models import KNECCBEGrade
    KNECCBEGrade.initialize_default_grades()

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse
from django.db import connection
from django_tenants.utils import get_tenant
from . import models

def is_admin_or_principal(user):
    """Check if user is admin or principal"""
    if user.is_authenticated and (user.is_superuser or user.is_staff):
        return True
    if hasattr(user, 'profile'):
        return user.profile.role in ['admin', 'principal']
    return False

@login_required
def submit_feedback(request):
    """Submit feedback form using raw SQL"""
    if request.method == 'POST':
        rating = request.POST.get('rating')
        message = request.POST.get('message')
        feedback_type = request.POST.get('feedback_type', 'general')
        subject = request.POST.get('subject', 'Feedback from user')
        
        if message:
            school = None
            school_name = None
            try:
                school = get_tenant(request)
                school_name = school.name if school else None
            except:
                pass
            
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO digitallibrary_feedback (
                        user_id, user_name, user_email, user_role,
                        feedback_type, priority, subject, message, rating,
                        school_name, status, created_at, updated_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                """, [
                    request.user.id if request.user.is_authenticated else None,
                    request.user.get_full_name() or request.user.username if request.user.is_authenticated else 'Anonymous',
                    request.user.email if request.user.is_authenticated else '',
                    getattr(request.user.profile, 'role', 'user') if hasattr(request.user, 'profile') else 'user',
                    feedback_type,
                    'medium',
                    subject,
                    message,
                    int(rating) if rating else 5,
                    school_name,
                    'pending'
                ])
            messages.success(request, "Thank you for your feedback!")
        else:
            messages.error(request, "Please enter a message.")
        
        return redirect(request.META.get('HTTP_REFERER', '/app/'))
    
    return render(request, 'digitallibrary/feedback_form.html')

@login_required
@user_passes_test(is_admin_or_principal)
def feedback_admin(request):
    """Admin view to manage all feedback using raw SQL"""
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    type_filter = request.GET.get('type', '')
    rating_filter = request.GET.get('rating', '')
    
    # Build SQL query
    sql = """
        SELECT id, user_name, user_email, feedback_type, priority, 
               subject, message, rating, status, school_name, 
               created_at, is_resolved, admin_response
        FROM digitallibrary_feedback
        WHERE 1=1
    """
    params = []
    
    if status_filter:
        sql += " AND status = %s"
        params.append(status_filter)
    
    if type_filter:
        sql += " AND feedback_type = %s"
        params.append(type_filter)
    
    if rating_filter:
        sql += " AND rating = %s"
        params.append(int(rating_filter))
    
    sql += " ORDER BY created_at DESC LIMIT 100"
    
    # Get feedback
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        feedbacks = cursor.fetchall()
    
    # Get statistics
    with connection.cursor() as cursor:
        cursor.execute("SELECT COUNT(*) FROM digitallibrary_feedback")
        total = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM digitallibrary_feedback WHERE status = 'pending'")
        pending = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM digitallibrary_feedback WHERE status = 'resolved'")
        resolved = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM digitallibrary_feedback WHERE status = 'reviewing'")
        reviewing = cursor.fetchone()[0]
        
        cursor.execute("SELECT COALESCE(AVG(rating), 0) FROM digitallibrary_feedback WHERE rating > 0")
        avg_rating = cursor.fetchone()[0]
        
        # Get by type
        cursor.execute("""
            SELECT feedback_type, COUNT(*) 
            FROM digitallibrary_feedback 
            GROUP BY feedback_type
        """)
        by_type = cursor.fetchall()
        
        # Get by rating
        cursor.execute("""
            SELECT rating, COUNT(*) 
            FROM digitallibrary_feedback 
            WHERE rating > 0
            GROUP BY rating 
            ORDER BY rating DESC
        """)
        by_rating = cursor.fetchall()
    
    feedback_types = [
        ('bug', 'Bug Report'), ('feature', 'Feature Request'),
        ('improvement', 'Improvement'), ('general', 'General Feedback'),
        ('issue', 'System Issue'), ('training', 'Training Request'),
        ('suggestion', 'Suggestion'), ('complaint', 'Complaint'),
        ('inquiry', 'Inquiry'),
    ]
    
    status_choices = [
        ('pending', 'Pending'), ('reviewing', 'Under Review'),
        ('in_progress', 'In Progress'), ('resolved', 'Resolved'),
        ('closed', 'Closed'), ('rejected', 'Rejected'),
    ]
    
    context = {
        'feedbacks': feedbacks,
        'stats': {
            'total': total,
            'pending': pending,
            'resolved': resolved,
            'reviewing': reviewing,
            'average_rating': round(float(avg_rating), 1) if avg_rating else 0,
            'by_type': by_type,
            'by_rating': by_rating,
        },
        'current_filter': {
            'status': status_filter,
            'type': type_filter,
            'rating': rating_filter,
        },
        'feedback_types': feedback_types,
        'status_choices': status_choices,
    }
    
    return render(request, 'digitallibrary/feedback_admin.html', context)

@login_required
@user_passes_test(is_admin_or_principal)
def resolve_feedback(request, feedback_id):
    """Mark feedback as resolved using raw SQL"""
    if request.method == 'POST':
        admin_response = request.POST.get('admin_response', '')
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE digitallibrary_feedback 
                SET status = 'resolved', is_resolved = TRUE, 
                    resolved_at = NOW(), admin_response = %s
                WHERE id = %s
            """, [admin_response, feedback_id])
        messages.success(request, "Feedback marked as resolved!")
    
    return redirect('digitallibrary:feedback_admin')

@login_required
@user_passes_test(is_admin_or_principal)
def delete_feedback(request, feedback_id):
    """Delete feedback using raw SQL"""
    if request.method == 'POST':
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM digitallibrary_feedback WHERE id = %s", [feedback_id])
        messages.success(request, "Feedback deleted successfully!")
    
    return redirect('digitallibrary:feedback_admin')

@login_required
def feedback_list(request):
    """Public feedback list view using raw SQL"""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, user_name, feedback_type, rating, message, 
                   created_at, admin_response, status
            FROM digitallibrary_feedback 
            WHERE is_public = TRUE AND status = 'resolved'
            ORDER BY created_at DESC 
            LIMIT 20
        """)
        feedbacks = cursor.fetchall()
    
    return render(request, 'digitallibrary/feedback_list.html', {'feedbacks': feedbacks})
# Add at the top with other imports
from django.contrib.auth import logout
from django.shortcuts import redirect

# Add this function at the end of the file
def logout_view(request):
    """Handle logout with both GET and POST"""
    logout(request)
    return redirect('/login/')
    
@csrf_exempt
def simple_login(request, tenant_schema=None):
    """Ultra-simple test login view"""
    
    # If tenant_schema is not in the URL, try to get it from the path
    if not tenant_schema:
        import re
        match = re.match(r'^/tenant/([^/]+)/', request.path)
        if match:
            tenant_schema = match.group(1)
            print(f"🔍 Extracted tenant from path: {tenant_schema}")
    
    print(f"🔐 SIMPLE LOGIN - User request, Tenant: {tenant_schema}")
    
    if request.method == 'POST':
        # Handle JSON or form data
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            username = data.get('username')
            password = data.get('password')
        else:
            username = request.POST.get('username')
            password = request.POST.get('password')
        
        print(f"🔐 SIMPLE LOGIN - User: {username}, Tenant: {tenant_schema}")
        
        # Authenticate
        user = authenticate(request, username=username, password=password)
        
        if user:
            login(request, user)
            if tenant_schema:
                request.session['tenant_schema'] = tenant_schema
            request.session.save()
            
            print(f"✅ LOGIN SUCCESS! Session: {request.session.session_key}")
            
            # Redirect to the correct tenant dashboard
            redirect_url = f'/tenant/{tenant_schema}/app/' if tenant_schema else '/app/'
            
            if request.content_type == 'application/json':
                return JsonResponse({
                    'success': True,
                    'user': username,
                    'session_key': request.session.session_key,
                    'redirect_url': redirect_url
                })
            else:
                from django.shortcuts import redirect
                return redirect(redirect_url)
        else:
            print(f"❌ LOGIN FAILED for {username}")
            if request.content_type == 'application/json':
                return JsonResponse({'success': False, 'error': 'Invalid credentials'}, status=400)
            else:
                return HttpResponse(f'<h2>Login Failed</h2><p>Invalid credentials for {username}</p><a href="/tenant/{tenant_schema}/app/simple-login/">Try again</a>', status=401)
    
    # GET request - show a simple form
    return HttpResponse(f'''
        <!DOCTYPE html>
        <html>
        <head>
            <title>Simple Login - {tenant_schema}</title>
            <style>
                body {{ font-family: Arial; padding: 50px; }}
                input {{ padding: 8px; margin: 5px; width: 200px; }}
                button {{ padding: 8px 20px; background: green; color: white; border: none; cursor: pointer; }}
            </style>
        </head>
        <body>
            <h2>Simple Login for {tenant_schema}</h2>
            <form method="post">
                <input type="text" name="username" placeholder="Username" required><br>
                <input type="password" name="password" placeholder="Password" required><br>
                <button type="submit">Login</button>
            </form>
            <p><strong>Test credentials:</strong> admin / admin123</p>
            <hr>
            <p><a href="/tenant/{tenant_schema}/app/login/">Go to regular login page</a></p>
        </body>
        </html>
    ''')
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required

@login_required
def debug_session(request, tenant_schema=None):
    """Debug view to check authentication status"""
    return JsonResponse({
        'authenticated': request.user.is_authenticated,
        'username': request.user.username,
        'role': request.user.profile.role if hasattr(request.user, 'profile') else None,
        'session_key': request.session.session_key,
        'tenant': request.session.get('tenant_schema'),
    })
